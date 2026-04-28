"""
=============================================================
  REPORTE DE COSTOS - VERSIÓN WEB (Render.com)
  Este archivo es el punto de entrada para el servidor web.
  Lee el Excel desde la misma carpeta del proyecto.
=============================================================
"""

import pandas as pd
import numpy as np
import os
from datetime import datetime
import plotly.graph_objects as go
from dash import Dash, html, dcc, Input, Output, dash_table, State, callback_context
import io
import base64

# ─── CONFIGURACIÓN ─────────────────────────────────────────
ARCHIVO_DATOS    = "Analisis de costos_PY.xlsx"
HOJA_EXPLOSION   = "Explosión"
HOJA_TIEMPOS     = "Tiempos"
PREFIJO_FABRIC   = "231"
PROCESOS_EXCLUIR = []
# ───────────────────────────────────────────────────────────

# ── Cargar datos ────────────────────────────────────────────
import sys
print(f"📂 Directorio actual: {os.getcwd()}")
print(f"📄 Archivos disponibles: {os.listdir('.')}")
if not os.path.exists(ARCHIVO_DATOS):
    print(f"❌ ERROR: No se encontró {ARCHIVO_DATOS}")
    sys.exit(1)
print(f"✅ Excel encontrado: {ARCHIVO_DATOS}")
df_exp = pd.read_excel(ARCHIVO_DATOS, sheet_name=HOJA_EXPLOSION)
df_tie = pd.read_excel(ARCHIVO_DATOS, sheet_name=HOJA_TIEMPOS)
try:
    df_mat = pd.read_excel(ARCHIVO_DATOS, sheet_name="Materiales")
    df_mat.columns = df_mat.columns.str.strip()
    df_mat["Codigo"] = df_mat["Codigo"].astype(str).str.strip()
    print("✅ Hoja Materiales cargada")
except:
    df_mat = pd.DataFrame(columns=["Codigo","Descripción","UM","TIPO DE COMPRA","MOQ","LT-días","Tipo"])
    print("⚠️ Hoja Materiales no encontrada, usando vacío")

df_exp.columns = df_exp.columns.str.strip()
df_tie.columns = df_tie.columns.str.strip()

# ── Datos históricos SAP ─────────────────────────────────────
ARCHIVO_HISTORICO = "cubo_costos.xlsx"
if os.path.exists(ARCHIVO_HISTORICO):
    df_cubo = pd.read_excel(ARCHIVO_HISTORICO, sheet_name="Cubo_Costos")
    df_cubo.columns = df_cubo.columns.str.strip()
    df_cubo["Cod_PT"] = df_cubo["Cod_PT"].astype(str).str.strip()
    def clean_cod(x):
        try:
            s = str(x).strip()
            if s in ("nan","None",""): return ""
            return str(int(float(s)))
        except: return str(x).strip()
    df_cubo["Cod_Insumo"] = df_cubo["Cod_Insumo"].apply(clean_cod)
    for col in ["VOL_PRODUCIDOS","CANT_REAL","COSTO_REAL"]:
        df_cubo[col] = pd.to_numeric(df_cubo[col], errors="coerce").fillna(0)
    PERIODOS_ORDENADOS = sorted(df_cubo["Periodo"].unique())
    pts_cubo = df_cubo[["Cod_PT"]].drop_duplicates()
    pts_cubo["Desc"] = pts_cubo["Cod_PT"].map(
        df_cubo[["Cod_PT","Desc_PT"]].drop_duplicates().set_index("Cod_PT")["Desc_PT"])
    def get_linea_h(c):
        c = str(c)
        if c.startswith("214011"): return "Cuadernos"
        if len(c)>=7 and not c[6].isdigit(): return "Nuevos Desarrollos"
        return "Útiles"
    pts_cubo["Linea"] = pts_cubo["Cod_PT"].apply(get_linea_h)
    pts_cubo = pts_cubo.sort_values("Cod_PT")
else:
    df_cubo = pd.DataFrame(); PERIODOS_ORDENADOS = []; pts_cubo = pd.DataFrame()


for col in ["Código PT", "Código Semi", "Componente", "Familia"]:
    if col in df_exp.columns:
        df_exp[col] = df_exp[col].astype(str).str.strip()
df_tie["Código Semi"] = df_tie["Código Semi"].astype(str).str.strip()

for col in ["Cantidad Total Requerida", "Cantidad Base", "Costo estandar"]:
    df_exp[col] = pd.to_numeric(df_exp[col], errors="coerce").fillna(0)
for col in ["Cantidad Base", "T.MO", "T.Maq", "Tarifa MO", "Tarifa Maquina"]:
    if col in df_tie.columns:
        df_tie[col] = pd.to_numeric(df_tie[col], errors="coerce").fillna(0)

# ── Funciones de cálculo ────────────────────────────────────
def es_fabricado(familia):
    return str(familia).strip().startswith(PREFIJO_FABRIC)

def get_tiempos(codigo, df_t):
    row = df_t[df_t["Código Semi"] == str(codigo)]
    return row.iloc[0] if not row.empty else None

def calcular_semi(codigo_semi, cantidad_req, df_e, df_t, cache, resumen_global, codigo_pt):
    cache_key = f"{codigo_semi}_{cantidad_req}"
    if cache_key in cache:
        return cache[cache_key]["costo_x_und"], []

    hijos = df_e[(df_e["Código Semi"] == str(codigo_semi)) & (df_e["Código PT"] == str(codigo_pt))].copy()
    if hijos.empty:
        return 0, []

    desc_semi     = hijos["Descripción Semi"].iloc[0] if "Descripción Semi" in hijos.columns else ""
    t             = get_tiempos(codigo_semi, df_t)
    proceso       = str(t["Proceso"]).strip().upper() if t is not None else "SIN PROCESO"
    cant_base_t   = float(t["Cantidad Base"])          if t is not None else 1
    tarifa_maq    = float(t["Tarifa Maquina"])         if t is not None else 0
    tarifa_mo     = float(t["Tarifa MO"])              if t is not None else 0
    t_maq         = float(t["T.Maq"])                  if t is not None else 0
    t_mo          = float(t["T.MO"])                   if t is not None else 0
    if cant_base_t == 0:
        cant_base_t = 1

    cif = (t_maq / cant_base_t) * cantidad_req * tarifa_maq
    mod = (t_mo  / cant_base_t) * cantidad_req * tarifa_mo

    detalle      = []
    cm_total     = 0
    cm_comprados = 0

    for _, row in hijos.iterrows():
        componente = str(row["Componente"])
        desc_comp  = str(row.get("Descripción Componente", ""))
        cantidad   = float(row["Cantidad Total Requerida"])
        costo_std  = float(row["Costo estandar"])
        familia    = str(row.get("Familia", componente[:3])).strip()

        if es_fabricado(familia):
            costo_calc, sub_det = calcular_semi(componente, cantidad, df_e, df_t, cache, resumen_global, codigo_pt)
            detalle.extend(sub_det)
            cm_comp = cantidad * costo_calc
        else:
            costo_calc   = costo_std
            cm_comp      = cantidad * costo_calc
            cm_comprados += cm_comp

        cm_total += cm_comp
        detalle.append({
            "Código Semi": codigo_semi, "Descripción Semi": desc_semi,
            "Componente": componente,   "Descripción Componente": desc_comp,
            "Familia": familia, "Tipo": "FABRICADO" if es_fabricado(familia) else "COMPRADO",
            "Proceso": proceso, "Cantidad Total Req": cantidad,
            "Costo Calculado": costo_calc, "CM": cm_comp,
            "CIF": 0, "MOD": 0, "Total": cm_comp,
        })

    total_semi  = cm_total + cif + mod
    costo_x_und = total_semi / cantidad_req if cantidad_req != 0 else 0

    if proceso not in PROCESOS_EXCLUIR:
        if proceso not in resumen_global:
            resumen_global[proceso] = {"CM": 0, "CIF": 0, "MOD": 0}
        resumen_global[proceso]["CM"]  += cm_comprados
        resumen_global[proceso]["CIF"] += cif
        resumen_global[proceso]["MOD"] += mod

    detalle.append({
        "Código Semi": codigo_semi, "Descripción Semi": desc_semi,
        "Componente": f"[PROCESO] {codigo_semi}",
        "Descripción Componente": f"{proceso} — CIF + MOD",
        "Familia": PREFIJO_FABRIC, "Tipo": "PROCESO", "Proceso": proceso,
        "Cantidad Total Req": cantidad_req, "Costo Calculado": costo_x_und,
        "CM": cm_total, "CIF": cif, "MOD": mod, "Total": cm_total + cif + mod,
    })

    cache[cache_key] = {"costo_x_und": costo_x_und}
    return costo_x_und, detalle


def explotar_pt(codigo_pt, df_e, df_t):
    nivel1 = df_e[df_e["Código Semi"] == str(codigo_pt)].copy()
    if nivel1.empty:
        return {}, [], 0

    cant_base_pt = float(nivel1["Cantidad Base"].iloc[0])
    if cant_base_pt == 0:
        cant_base_pt = 1
    desc_pt = nivel1["Descripción Semi"].iloc[0] if "Descripción Semi" in nivel1.columns else ""

    t           = get_tiempos(codigo_pt, df_t)
    proceso_pt  = str(t["Proceso"]).strip().upper() if t is not None else "ENCAJADO"
    cant_base_t = float(t["Cantidad Base"])          if t is not None else 1
    tarifa_maq  = float(t["Tarifa Maquina"])         if t is not None else 0
    tarifa_mo   = float(t["Tarifa MO"])              if t is not None else 0
    t_maq       = float(t["T.Maq"])                  if t is not None else 0
    t_mo        = float(t["T.MO"])                   if t is not None else 0
    if cant_base_t == 0:
        cant_base_t = 1

    cif_pt = (t_maq / cant_base_t) * cant_base_pt * tarifa_maq
    mod_pt = (t_mo  / cant_base_t) * cant_base_pt * tarifa_mo

    cache          = {}
    detalle        = []
    resumen_global = {}
    cm_total       = 0
    cm_comprados   = 0

    for _, row in nivel1.iterrows():
        componente = str(row["Componente"])
        desc_comp  = str(row.get("Descripción Componente", ""))
        cantidad   = float(row["Cantidad Total Requerida"])
        costo_std  = float(row["Costo estandar"])
        familia    = str(row.get("Familia", componente[:3])).strip()

        if es_fabricado(familia):
            costo_calc, sub_det = calcular_semi(componente, cantidad, df_e, df_t, cache, resumen_global, codigo_pt)
            detalle.extend(sub_det)
            cm_comp = cantidad * costo_calc
        else:
            costo_calc   = costo_std
            cm_comp      = cantidad * costo_calc
            cm_comprados += cm_comp

        cm_total += cm_comp
        detalle.append({
            "Código Semi": codigo_pt, "Descripción Semi": desc_pt,
            "Componente": componente, "Descripción Componente": desc_comp,
            "Familia": familia, "Tipo": "FABRICADO" if es_fabricado(familia) else "COMPRADO",
            "Proceso": proceso_pt, "Cantidad Total Req": cantidad,
            "Costo Calculado": costo_calc, "CM": cm_comp,
            "CIF": 0, "MOD": 0, "Total": cm_comp,
        })

    if proceso_pt not in resumen_global:
        resumen_global[proceso_pt] = {"CM": 0, "CIF": 0, "MOD": 0}
    resumen_global[proceso_pt]["CM"]  += cm_comprados
    resumen_global[proceso_pt]["CIF"] += cif_pt
    resumen_global[proceso_pt]["MOD"] += mod_pt

    total_pt    = cm_total + cif_pt + mod_pt
    costo_x_und = total_pt / cant_base_pt
    return resumen_global, detalle, costo_x_und


# ── Generar resumen global ──────────────────────────────────
lista_pt      = df_exp["Código PT"].unique()
filas_resumen = []
filas_detalle = []

for codigo_pt in lista_pt:
    df_pt_rows = df_exp[df_exp["Código PT"] == codigo_pt]
    if df_pt_rows.empty:
        continue
    desc_pt = df_pt_rows["Descripción PT"].iloc[0]
    resumen, detalle, _ = explotar_pt(codigo_pt, df_exp, df_tie)
    total_general = sum(v["CM"] + v["CIF"] + v["MOD"] for v in resumen.values())
    if total_general == 0:
        continue
    cant_base_pt = float(df_exp[df_exp["Código Semi"] == codigo_pt]["Cantidad Base"].iloc[0]) \
                   if not df_exp[df_exp["Código Semi"] == codigo_pt].empty else 1
    if cant_base_pt == 0:
        cant_base_pt = 1
    for proceso, valores in resumen.items():
        for tipo, monto in [("CM", valores["CM"]), ("CIF", valores["CIF"]), ("MOD", valores["MOD"])]:
            if monto > 0:
                filas_resumen.append({
                    "Código PT": codigo_pt, "Descripción PT": desc_pt,
                    "Proceso": proceso, "Tipo de Costo": f"{tipo} {proceso}",
                    "Costo Unitario": monto / cant_base_pt, "Total PT": total_general,
                })
    for d in detalle:
        d["Código PT"]      = codigo_pt
        d["Descripción PT"] = desc_pt
        filas_detalle.append(d)

df_resumen = pd.DataFrame(filas_resumen)
df_detalle = pd.DataFrame(filas_detalle)
if not df_resumen.empty:
    df_resumen["% del Total"] = (
        df_resumen["Costo Unitario"] /
        df_resumen.groupby("Código PT")["Costo Unitario"].transform("sum")
    )


# ── Procesos con detalle completo ──────────────────────────
PROCESOS_CON_DETALLE = ["INYECCIÓN","INYECCION","ENSAMBLE","ENCAJADO",
                        "TROQUELADO","SOPLADO","DOSIFICADO","TERMOEN",
                        "TERMOENCOGIDO"]

def construir_tabla_cascada(codigo_pt, df_e, df_t, resumen_sim, cant_base_pt):
    """
    Construye tabla detallada por proceso:
    - Procesos con detalle: muestra componentes + MOD + CIF
    - Otros procesos: solo una fila con el costo total
    - Fila final: TOTAL PT
    """
    filas = []
    total_general = 0

    # Reconstruir detalle para este PT
    _, detalle_sim, _ = explotar_pt(codigo_pt, df_e, df_t)

    # Agrupar detalle por proceso
    det_df = pd.DataFrame(detalle_sim) if detalle_sim else pd.DataFrame()

    for proceso, valores in resumen_sim.items():
        total_proceso = (valores["CM"] + valores["CIF"] + valores["MOD"]) / cant_base_pt
        proc_upper = proceso.strip().upper()
        tiene_detalle = any(p in proc_upper for p in PROCESOS_CON_DETALLE)

        if tiene_detalle and not det_df.empty:
            # Filas de componentes del proceso
            det_proc = det_df[det_df["Proceso"].str.upper().str.strip() == proc_upper]

# Componentes de inyección: agrupar por nombre Componente de Tiempos
            # CM = suma total de CM de todos los semis de ese componente / cant_base_pt
            if "INYEC" in proc_upper:
                grupos_comp = {}
                codigos = det_proc["Código Semi"].unique()
                for cod in codigos:
                    t_row = df_t[df_t["Código Semi"] == str(cod)]
                    if not t_row.empty:
                        comp_nombre = str(t_row.iloc[0].get("Componente", cod))
                        # Suma todo el CM de los componentes comprados de este semi
                        filas_semi = det_proc[
                            (det_proc["Código Semi"] == cod) & 
                            (det_proc["Tipo"] == "COMPRADO")
                        ]
                        cm_total_semi = filas_semi["CM"].sum() if not filas_semi.empty else 0
                        if cm_total_semi > 0:
                            grupos_comp[comp_nombre] = grupos_comp.get(comp_nombre, 0) + cm_total_semi
                for comp_nombre, cm_total in grupos_comp.items():
                    # Dividir entre cant_base_pt igual que CIF y MOD
                    costo_und = cm_total / cant_base_pt
                    filas.append({
                        "Proceso": proceso,
                        "Componente": comp_nombre,
                        "Costo": costo_und,
                    })
            else:
                # Componentes comprados (no 231)
                comp_comprados = det_proc[det_proc["Tipo"] == "COMPRADO"]
                grupos = comp_comprados.groupby(["Componente","Descripción Componente"])["CM"].sum()
                for (cod, desc), cm in grupos.items():
                    costo_und = cm / cant_base_pt
                    if costo_und > 0:
                        filas.append({
                            "Proceso": proceso,
                            "Componente": desc if desc and desc != "nan" else cod,
                            "Costo": costo_und,
                        })

            # MOD y CIF
            mod = valores["MOD"] / cant_base_pt
            cif = valores["CIF"] / cant_base_pt
            if mod > 0:
                filas.append({"Proceso": proceso, "Componente": "MOD", "Costo": mod})
            if cif > 0:
                filas.append({"Proceso": proceso, "Componente": "CIF", "Costo": cif})
        else:
            # Proceso sin detalle — solo una fila con total
            filas.append({"Proceso": proceso, "Componente": "—", "Costo": total_proceso})

        total_general += total_proceso

    # Fila total
    filas.append({"Proceso": "TOTAL PT", "Componente": "", "Costo": total_general})

    # Agregar %TG
    for f in filas:
        if total_general > 0 and f["Proceso"] != "TOTAL PT":
            f["%TG"] = f"{f['Costo']/total_general*100:.1f}%"
        elif f["Proceso"] == "TOTAL PT":
            f["%TG"] = "100.0%"
        else:
            f["%TG"] = "0.0%"
        f["Costo"] = f"S/ {f['Costo']:.6f}"

    return filas

# ── Funciones Histórico ─────────────────────────────────────
def get_costo_teorico(codigo_pt):
    """Retorna (cm_unit, cif_unit, mod_unit, total_unit) teórico."""
    resumen, _, _ = explotar_pt(codigo_pt, df_exp, df_tie)
    cant_base = float(df_exp[df_exp["Código Semi"]==codigo_pt]["Cantidad Base"].iloc[0]) \
                if not df_exp[df_exp["Código Semi"]==codigo_pt].empty else 1
    if cant_base == 0: cant_base = 1
    cm  = sum(v["CM"]  for v in resumen.values()) / cant_base
    cif = sum(v["CIF"] for v in resumen.values()) / cant_base
    mod = sum(v["MOD"] for v in resumen.values()) / cant_base
    return cm, cif, mod, cm+cif+mod

# ── Costo real mensual con fallback ──────────────────────────
PERIODOS_ORDENADOS = sorted(df_cubo["Periodo"].unique())

def _get_costo_sap(codigo, periodo):
    """
    Obtiene costo unitario del PT desde SAP usando promedio ponderado por volumen.
    Cuando hay múltiples OPs en el mes, pondera cada CU por su volumen.
    """
    sub = df_cubo[(df_cubo["Cod_PT"]==str(codigo)) & (df_cubo["Periodo"]==periodo)]
    if sub.empty: return None

    # Volumen total del mes (suma de todas las OPs)
    vol_total = sub[sub["TIPO_COSTO_DIR"]=="Volumen"]["VOL_PRODUCIDOS"].sum()
    if vol_total == 0: return None

    # Costo Unitario: promedio ponderado por volumen de cada OP
    cu_rows = sub[sub["TIPO_COSTO_DIR"]=="Costo Unitario"][["OP","COSTO_REAL"]].copy()
    vol_rows = sub[sub["TIPO_COSTO_DIR"]=="Volumen"][["OP","VOL_PRODUCIDOS"]].copy()
    if cu_rows.empty: return None

    # Merge CU con volumen por OP
    merged = cu_rows.merge(vol_rows, on="OP", how="left")
    merged["VOL_PRODUCIDOS"] = merged["VOL_PRODUCIDOS"].fillna(0)
    vol_cu = merged["VOL_PRODUCIDOS"].sum()
    if vol_cu > 0:
        # Promedio ponderado
        cu_ponderado = (merged["COSTO_REAL"] * merged["VOL_PRODUCIDOS"]).sum() / vol_cu
    else:
        # Sin volumen por OP, usar promedio simple
        cu_ponderado = cu_rows["COSTO_REAL"].mean()

    # CM, CIF, MOD: también ponderados por volumen
    cm_dir  = sub[sub["TIPO_COSTO_DIR"]=="Materiales"]["COSTO_REAL"].sum() / vol_total
    cif_dir = sub[(sub["TIPO_COSTO_DIR"]=="Costo de Conversión") &
                  (sub["CLASE_COSTO"]=="Gastos Indirectos")]["COSTO_REAL"].sum() / vol_total
    mod_dir = sub[(sub["TIPO_COSTO_DIR"]=="Costo de Conversión") &
                  (sub["CLASE_COSTO"]=="Mano de obra directa")]["COSTO_REAL"].sum() / vol_total

    return {"total": cu_ponderado, "cm": cm_dir, "cif": cif_dir,
            "mod": mod_dir, "vol": vol_total}


def _get_teo_semi(codigo, codigo_pt):
    """Costo teórico unitario (CM+CIF+MOD) de un semi para usar como fallback."""
    hijos = df_exp[(df_exp["Código Semi"]==str(codigo)) &
                   (df_exp["Código PT"]==str(codigo_pt))].copy()
    if hijos.empty: return 0
    t  = get_tiempos(codigo, df_tie)
    cb = float(t["Cantidad Base"]) if t is not None else 1
    if cb == 0: cb = 1
    cif_u = (float(t["T.Maq"]) if t is not None else 0) / cb * (float(t["Tarifa Maquina"]) if t is not None else 0)
    mod_u = (float(t["T.MO"])  if t is not None else 0) / cb * (float(t["Tarifa MO"])       if t is not None else 0)
    cache = {}
    def _cm(cod, cant, pt):
        key = f"{cod}_{cant}"
        if key in cache: return cache[key]
        h = df_exp[(df_exp["Código Semi"]==cod)&(df_exp["Código PT"]==pt)]
        if h.empty: return 0
        t2 = get_tiempos(cod, df_tie)
        cb2= float(t2["Cantidad Base"]) if t2 is not None else 1
        if cb2==0: cb2=1
        cif2=(float(t2["T.Maq"]) if t2 is not None else 0)/cb2*(float(t2["Tarifa Maquina"]) if t2 is not None else 0)
        mod2=(float(t2["T.MO"])  if t2 is not None else 0)/cb2*(float(t2["Tarifa MO"])       if t2 is not None else 0)
        cm2 = sum(
            float(r["Cantidad Total Requerida"]) * (
                _cm(str(r["Componente"]), float(r["Cantidad Total Requerida"]), pt)
                if str(r.get("Familia",str(r["Componente"])[:3])).startswith("231")
                else float(r["Costo estandar"])
            ) for _,r in h.iterrows()
        )
        cu2 = (cm2 + (cif2+mod2)*cb2) / cb2
        cache[key] = cu2; return cu2
    cm_total = sum(
        float(r["Cantidad Total Requerida"]) * (
            _cm(str(r["Componente"]), float(r["Cantidad Total Requerida"]), codigo_pt)
            if str(r.get("Familia",str(r["Componente"])[:3])).startswith("231")
            else float(r["Costo estandar"])
        ) for _,r in hijos.iterrows()
    )
    return (cm_total + (cif_u + mod_u) * cb) / cb


def get_costo_real_mensual(codigo_pt):
    """
    Costo real mensual del PT usando directamente el campo Costo Unitario de SAP.
    SAP ya acumula todos los niveles de la BOM en ese campo.
    Si un mes no tiene datos → usar costo teórico como fallback.
    CM, CIF, MOD del gráfico se toman del nivel directo del PT.
    """
    # Costo teórico total para fallback
    cm_t, cif_t, mod_t, total_t = get_costo_teorico(codigo_pt)

    # Obtener factores de los semis directos del PT para desglosar CM/CIF/MOD
    nivel1 = df_exp[df_exp["Código Semi"]==str(codigo_pt)].copy()
    cant_base_pt = float(nivel1["Cantidad Base"].iloc[0]) if not nivel1.empty else 1
    if cant_base_pt == 0: cant_base_pt = 1

    result = []
    for per in PERIODOS_ORDENADOS:
        r = _get_costo_sap(str(codigo_pt), per)
        if r and r["total"] > 0:
            # Datos reales SAP: usar Costo Unitario para el total
            # Para CM/CIF/MOD usar los valores directos del nivel PT
            # y sumar CIF+MOD de los semis no incluidos en el direct
            total    = r["total"]
            vol_pt   = r["vol"]
            tiene    = vol_pt > 0
            # Aproximar CM/CIF/MOD proporcional al teórico
            if total_t > 0:
                cm_r  = total * (cm_t  / total_t)
                cif_r = total * (cif_t / total_t)
                mod_r = total * (mod_t / total_t)
            else:
                cm_r = cif_r = mod_r = total / 3
        else:
            # Fallback: usar costo teórico
            total  = total_t
            vol_pt = 0
            tiene  = False
            cm_r   = cm_t
            cif_r  = cif_t
            mod_r  = mod_t

        result.append({
            "Periodo":   per,
            "Volumen":   vol_pt,
            "CM":        round(cm_r,  6),
            "CIF":       round(cif_r, 6),
            "MOD":       round(mod_r, 6),
            "Total":     round(total, 6),
            "Con_Datos": tiene,
        })

    return pd.DataFrame(result)

# ── Lista PTs disponibles en cubo ────────────────────────────
pts_cubo = df_cubo[["Cod_PT"]].drop_duplicates()
pts_cubo["Desc"] = pts_cubo["Cod_PT"].map(
    df_cubo[["Cod_PT","Desc_PT"]].drop_duplicates().set_index("Cod_PT")["Desc_PT"]
)

def get_linea(codigo_pt):
    c = str(codigo_pt)
    if c.startswith("214011"): return "Cuadernos"
    if len(c) >= 7 and not c[6].isdigit(): return "Nuevos Desarrollos"
    return "Útiles"

pts_cubo["Linea"] = pts_cubo["Cod_PT"].apply(get_linea)
pts_cubo = pts_cubo.sort_values("Cod_PT")


# ── Dashboard ───────────────────────────────────────────────
app    = Dash(__name__)
server = app.server

# ── Seguridad: solo redes Layconsa ──────────────────────────
from flask import request, abort

IPS_PERMITIDAS = {
    "143.208.132.11",  # WiFi 1 Layconsa
    # "200.xxx.xxx.xxx",  # WiFi 2 — agrega cuando tengas la IP real
}

@server.before_request
def verificar_ip():
    forwarded = request.headers.get("X-Forwarded-For", "")
    remote    = request.remote_addr
    todas_ips = [ip.strip() for ip in forwarded.split(",")]
    todas_ips.append(remote)
    print(f"IPs detectadas: {todas_ips}")
    # Filtro IP temporalmente desactivado — revisar logs para obtener IP real
    # if not any(ip in IPS_PERMITIDAS for ip in todas_ips):
    #     abort(403)

# ── Autenticación ────────────────────────────────────────────
import dash_auth

VALID_USERNAME_PASSWORD_PAIRS = {
    'pablo': '491iqwbCVOR',
    'admin': 'otro_password',
}

auth = dash_auth.BasicAuth(app, VALID_USERNAME_PASSWORD_PAIRS)

COLORES = {
    "bg": "#0F1923", "card": "#1A2633", "text": "#E8EDF2",
    "accent": "#00C8FF", "CM": "#2196F3", "MOD": "#4CAF50",
    "CIF": "#FF9800", "TOTAL": "#E91E63",
    "red": "#FF5252", "green": "#4CAF50", "orange": "#FF9800",
    "cm": "#2196F3", "cif": "#FF9800", "mod": "#4CAF50",
    "teorico": "#FF5252",
}

def empty_fig():
    import plotly.graph_objects as go
    fig = go.Figure()
    fig.update_layout(
        template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)", font_color="#E8EDF2",
    )
    return fig


lista_pt_dd = df_resumen[["Código PT", "Descripción PT"]].drop_duplicates()

def get_linea(codigo_pt):
    c = str(codigo_pt)
    if c.startswith("214011"):
        return "Cuadernos"
    # 7th digit (index 6) is a letter → Nuevos Desarrollos
    if len(c) >= 7 and not c[6].isdigit():
        return "Nuevos Desarrollos"
    return "Útiles"

lista_pt_dd["Línea"] = lista_pt_dd["Código PT"].apply(get_linea)


def get_maquinas_inyeccion(codigo_pt):
    """Máquinas de INYECCIÓN con T.Ciclo y Cav.Oper editables."""
    visitados = set()
    maquinas  = {}

    def buscar(codigo):
        if codigo in visitados:
            return
        visitados.add(codigo)
        hijos = df_exp[df_exp["Código Semi"] == codigo]
        for _, row in hijos.iterrows():
            comp    = str(row["Componente"])
            familia = str(row.get("Familia", comp[:3])).strip()
            t       = df_tie[df_tie["Código Semi"] == comp]
            if not t.empty:
                proc = str(t.iloc[0].get("Proceso", "")).strip().upper()
                if "INYEC" in proc:
                    t_row = t.iloc[0]
                    maq   = str(t_row.get("Maquina", comp))
                    if maq not in maquinas:
                        maquinas[maq] = {
                            "Maquina":   maq,
                            "T.Ciclo":   float(t_row.get("T.ciclo",        0) or 0),
                            "Cav.Oper":  float(t_row.get("Cav. Oper",      0) or 0),
                            "Cav.Tot":   float(t_row.get("Cav. Tot",       0) or 0),
                            "Tarifa Maq":float(t_row.get("Tarifa Maquina", 0) or 0),
                            "Tarifa MO": float(t_row.get("Tarifa MO",      0) or 0),
                        }
            if familia.startswith("231"):
                buscar(comp)

    buscar(codigo_pt)
    return list(maquinas.values())


def get_semis_otros_procesos(codigo_pt):
    """Otros procesos agrupados por máquina — incluye PT (ENCAJADO) y semis."""
    visitados  = set()
    maquinas   = {}
    excluidos  = ["INYEC", "MYT", "M&T", "MASAS"]

    def agregar_si_aplica(codigo):
        """Agrega el código a la tabla si su proceso no está excluido."""
        t = df_tie[df_tie["Código Semi"] == codigo]
        if not t.empty:
            proc = str(t.iloc[0].get("Proceso", "")).strip().upper()
            if not any(ex in proc for ex in excluidos) and proc != "SIN PROCESO":
                t_row = t.iloc[0]
                maq   = str(t_row.get("Maquina", codigo))
                key   = f"{proc}_{maq}"
                if key not in maquinas:
                    maquinas[key] = {
                        "Proceso":      proc,
                        "Maquina":      maq,
                        "Cantidad Base":float(t_row.get("Cantidad Base", 0) or 0),
                        "T.MO":         float(t_row.get("T.MO",          0) or 0),
                        "T.Maq":        float(t_row.get("T.Maq",         0) or 0),
                        "Cant.Opr":     float(t_row.get("Cant.Opr",      0) or 0),
                        "Tarifa Maq":   float(t_row.get("Tarifa Maquina",0) or 0),
                        "Tarifa MO":    float(t_row.get("Tarifa MO",     0) or 0),
                    }

    def buscar(codigo):
        if codigo in visitados:
            return
        visitados.add(codigo)
        # Verificar el propio código (para capturar ENCAJADO del PT)
        agregar_si_aplica(codigo)
        hijos = df_exp[df_exp["Código Semi"] == codigo]
        for _, row in hijos.iterrows():
            comp    = str(row["Componente"])
            familia = str(row.get("Familia", comp[:3])).strip()
            agregar_si_aplica(comp)
            if familia.startswith("231"):
                buscar(comp)

    buscar(codigo_pt)
    return list(maquinas.values())


TAB_COSTOS = html.Div(
    style={"backgroundColor": COLORES["bg"], "minHeight": "100vh",
           "fontFamily": "'Segoe UI', sans-serif",
           "color": COLORES["text"], "padding": "20px"},
    children=[
        html.H1("📦 Reporte de Costos por Proceso",
                style={"color": COLORES["accent"], "textAlign": "center", "marginBottom": "5px"}),
        html.P(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
               style={"color": "#7A9BBF", "textAlign": "center", "marginBottom": "25px"}),

        html.Div(style={"marginBottom": "25px"}, children=[
            html.Label("Selecciona un Producto Terminado:",
                       style={"color": COLORES["accent"], "fontWeight": "bold"}),
            html.Div(style={"display": "flex", "gap": "10px", "margin": "8px 0"}, children=[
                html.Button("Todos",              id="linea-todos",      n_clicks=0,
                    style={"backgroundColor": COLORES["accent"], "color": "#000",
                           "border": "none", "borderRadius": "6px", "padding": "6px 16px",
                           "cursor": "pointer", "fontWeight": "bold", "fontSize": "13px"}),
                html.Button("Útiles",             id="linea-utiles",     n_clicks=0,
                    style={"backgroundColor": "#4A5568", "color": "white",
                           "border": "none", "borderRadius": "6px", "padding": "6px 16px",
                           "cursor": "pointer", "fontSize": "13px"}),
                html.Button("Cuadernos",          id="linea-cuadernos",  n_clicks=0,
                    style={"backgroundColor": "#4A5568", "color": "white",
                           "border": "none", "borderRadius": "6px", "padding": "6px 16px",
                           "cursor": "pointer", "fontSize": "13px"}),
                html.Button("Nuevos Desarrollos", id="linea-nuevos",     n_clicks=0,
                    style={"backgroundColor": "#4A5568", "color": "white",
                           "border": "none", "borderRadius": "6px", "padding": "6px 16px",
                           "cursor": "pointer", "fontSize": "13px"}),
            ]),
            dcc.Dropdown(
                id="selector-pt",
                options=[{"label": f"{r['Código PT']} — {r['Descripción PT']}",
                          "value": r["Código PT"]}
                         for _, r in lista_pt_dd.iterrows()],
                value=lista_pt_dd["Código PT"].iloc[0],
                style={"marginTop": "8px", "color": "#000"}
            ),
        ]),

        html.Div(id="kpis", style={"display": "flex", "gap": "15px",
                                    "marginBottom": "25px", "flexWrap": "wrap"}),

        # Simulador
        html.Div(style={"backgroundColor": COLORES["card"], "borderRadius": "12px",
                        "padding": "15px", "marginBottom": "20px",
                        "border": "1px solid #00C8FF"}, children=[
            html.H3("🔧 Simulador de Inyección — Modifica T.Ciclo y Cav.Oper por Máquina",
                    style={"color": COLORES["accent"], "fontSize": "16px",
                           "marginTop": 0, "marginBottom": "10px"}),
            html.P("Edita los valores en la tabla y presiona Recalcular.",
                   style={"color": "#7A9BBF", "fontSize": "12px", "marginBottom": "10px"}),
            dash_table.DataTable(
                id="tabla-simulador",
                columns=[
                    {"name": "Máquina",        "id": "Maquina",    "editable": False},
                    {"name": "T.Ciclo (s)",    "id": "T.Ciclo",    "editable": True,  "type": "numeric"},
                    {"name": "Cav.Oper",       "id": "Cav.Oper",   "editable": True,  "type": "numeric"},
                    {"name": "Cav.Tot",        "id": "Cav.Tot",    "editable": False},
                    {"name": "Cant.Base Calc", "id": "Cant.Base",  "editable": False},
                    {"name": "Tarifa Maq",     "id": "Tarifa Maq", "editable": True, "type": "numeric"},
                    {"name": "Tarifa MO",      "id": "Tarifa MO",  "editable": False},
                ],
                style_header={"backgroundColor": "#1F3864", "color": "white", "fontWeight": "bold"},
                style_cell={"backgroundColor": "#1E2D3D", "color": COLORES["text"],
                            "border": "1px solid #2A3F54", "padding": "8px", "textAlign": "center"},
                style_data_conditional=[
                    {"if": {"column_editable": True},
                     "backgroundColor": "#0D2137", "border": "1px solid #00C8FF"},
                    {"if": {"row_index": "odd"}, "backgroundColor": "#162030"},
                ],
                editable=True, page_action="none",
            ),
        ]),

        # Botón recalcular — aplica a AMBOS simuladores
        html.Div(style={"textAlign": "center", "margin": "15px 0"}, children=[
            html.Button("🔄 Recalcular Todos los Procesos", id="btn-recalcular",
                style={"backgroundColor": COLORES["accent"], "color": "#000",
                       "fontWeight": "bold", "border": "none", "borderRadius": "8px",
                       "padding": "12px 40px", "cursor": "pointer", "fontSize": "15px",
                       "boxShadow": "0 0 15px rgba(0,200,255,0.4)"}),
            html.Button("⬇️ Descargar Excel", id="btn-descargar",
                style={"backgroundColor": "#4CAF50", "color": "#000",
                       "fontWeight": "bold", "border": "none", "borderRadius": "8px",
                       "padding": "12px 30px", "cursor": "pointer", "fontSize": "15px",
                       "boxShadow": "0 0 15px rgba(76,175,80,0.4)"}),
            html.Button("📄 Descargar PT Actual", id="btn-descargar-pt",
                style={"backgroundColor": "#FF9800", "color": "#000",
                       "fontWeight": "bold", "border": "none", "borderRadius": "8px",
                       "padding": "12px 30px", "cursor": "pointer", "fontSize": "15px",
                       "boxShadow": "0 0 15px rgba(255,152,0,0.4)"}),
            dcc.Download(id="descarga-excel"),
            dcc.Download(id="descarga-pt"),
            html.Div(id="msg-simulador",
                     style={"color": "#4CAF50", "fontSize": "13px", "marginTop": "8px"}),
        ]),

        # Simulador otros procesos
        html.Div(style={"backgroundColor": COLORES["card"], "borderRadius": "12px",
                        "padding": "15px", "marginBottom": "20px",
                        "border": "1px solid #4CAF50"}, children=[
            html.H3("⚙️ Simulador Otros Procesos — Modifica Cantidad Base, T.MO, T.Maq",
                    style={"color": "#4CAF50", "fontSize": "16px",
                           "marginTop": 0, "marginBottom": "10px"}),
            html.P("Edita los valores y presiona Recalcular para ver el impacto.",
                   style={"color": "#7A9BBF", "fontSize": "12px", "marginBottom": "10px"}),
            dash_table.DataTable(
                id="tabla-simulador-otros",
                columns=[
                    {"name": "Proceso",        "id": "Proceso",       "editable": False},
                    {"name": "Máquina",        "id": "Maquina",       "editable": False},
                    {"name": "Cantidad Base",  "id": "Cantidad Base", "editable": True,  "type": "numeric"},
                    {"name": "T.MO",           "id": "T.MO",          "editable": True,  "type": "numeric"},
                    {"name": "T.Maq",          "id": "T.Maq",         "editable": True,  "type": "numeric"},
                    {"name": "Cant.Opr",       "id": "Cant.Opr",      "editable": False},
                    {"name": "Tarifa Maq",     "id": "Tarifa Maq",    "editable": True, "type": "numeric"},
                    {"name": "Tarifa MO",      "id": "Tarifa MO",     "editable": False},
                ],
                style_header={"backgroundColor": "#1F3864", "color": "white", "fontWeight": "bold"},
                style_cell={"backgroundColor": "#1E2D3D", "color": COLORES["text"],
                            "border": "1px solid #2A3F54", "padding": "8px", "textAlign": "center"},
                style_data_conditional=[
                    {"if": {"column_editable": True},
                     "backgroundColor": "#0D2137", "border": "1px solid #4CAF50"},
                    {"if": {"row_index": "odd"}, "backgroundColor": "#162030"},
                ],
                editable=True, page_action="none",
            ),
        ]),

        html.Div(style={"display": "grid", "gridTemplateColumns": "1fr 1fr",
                        "gap": "20px", "marginBottom": "20px"}, children=[
            html.Div(style={"backgroundColor": COLORES["card"],
                            "borderRadius": "12px", "padding": "15px"}, children=[
                html.H3("Cascada de Costos (S/)", style={"color": COLORES["accent"],
                        "fontSize": "16px", "marginTop": 0}),
                dcc.Graph(id="grafico-cascada")
            ]),
            html.Div(style={"backgroundColor": COLORES["card"],
                            "borderRadius": "12px", "padding": "15px"}, children=[
                html.H3("Cascada de Costos (%)", style={"color": COLORES["accent"],
                        "fontSize": "16px", "marginTop": 0}),
                dcc.Graph(id="grafico-cascada-pct")
            ]),
        ]),

        html.Div(style={"display": "grid", "gridTemplateColumns": "1fr 1fr",
                        "gap": "20px", "marginBottom": "20px"}, children=[
            html.Div(style={"backgroundColor": COLORES["card"],
                            "borderRadius": "12px", "padding": "15px"}, children=[
                html.H3("Costo por Proceso (%)", style={"color": COLORES["accent"],
                        "fontSize": "16px", "marginTop": 0}),
                dcc.Graph(id="grafico-donut")
            ]),
            html.Div(style={"backgroundColor": COLORES["card"],
                            "borderRadius": "12px", "padding": "15px"}, children=[
                html.H3("Costo por Proceso (S/)", style={"color": COLORES["accent"],
                        "fontSize": "16px", "marginTop": 0}),
                dcc.Graph(id="grafico-donut-soles")
            ]),
        ]),

        # ── Pareto reemplaza tabla resumen ────────────────────
        html.Div(style={"backgroundColor": COLORES["card"], "borderRadius": "12px",
                        "padding": "15px", "marginBottom": "20px"}, children=[
            html.H3("📊 Pareto de Costos por Tipo",
                    style={"color": COLORES["accent"], "fontSize": "16px", "marginTop": 0}),
            dcc.Graph(id="grafico-pareto")
        ]),

        # ── Tabla cascada detallada + dona ─────────────────────
        html.Div(style={"display": "grid", "gridTemplateColumns": "1fr 1fr",
                        "gap": "20px", "marginBottom": "20px"}, children=[
            html.Div(style={"backgroundColor": COLORES["card"], "borderRadius": "12px",
                            "padding": "15px"}, children=[
                html.H3("📋 Cascada de Costos Detallada por Proceso",
                        style={"color": COLORES["accent"], "fontSize": "16px", "marginTop": 0}),
                dash_table.DataTable(
                    id="tabla-cascada-detalle",
                    columns=[
                        {"name": "Proceso",    "id": "Proceso"},
                        {"name": "Componente", "id": "Componente"},
                        {"name": "Costo",      "id": "Costo"},
                        {"name": "%TG",        "id": "%TG"},
                    ],
                    style_header={"backgroundColor": "#1F3864", "color": "white", "fontWeight": "bold"},
                    style_cell={"backgroundColor": "#1E2D3D", "color": COLORES["text"],
                                "border": "1px solid #2A3F54", "padding": "8px", "textAlign": "center"},
                    style_cell_conditional=[
                        {"if": {"column_id": "Proceso"},    "width": "150px", "minWidth": "150px"},
                        {"if": {"column_id": "Componente"}, "width": "150px", "minWidth": "150px"},
                        {"if": {"column_id": "Costo"},      "width": "150px", "minWidth": "150px"},
                        {"if": {"column_id": "%TG"},        "width": "80px",  "minWidth": "80px"},
                    ],
                    style_data_conditional=[
                        {"if": {"row_index": "odd"}, "backgroundColor": "#162030"},
                        {"if": {"filter_query": '{Componente} = "MOD"'},
                         "color": "#4CAF50", "fontWeight": "bold"},
                        {"if": {"filter_query": '{Componente} = "CIF"'},
                         "color": "#FF9800", "fontWeight": "bold"},
                        {"if": {"filter_query": '{Proceso} = "TOTAL PT"'},
                         "backgroundColor": "#1F3864", "color": "#00C8FF",
                         "fontWeight": "bold", "fontSize": "14px"},
                    ],
                    page_action="none",
                )
            ]),
            html.Div(style={"backgroundColor": COLORES["card"], "borderRadius": "12px",
                            "padding": "15px"}, children=[
                html.H3("🥧 Distribución por Componente (%)",
                        style={"color": COLORES["accent"], "fontSize": "16px", "marginTop": 0}),
                dcc.Graph(id="grafico-dona-cascada")
            ]),
        ]),

        # ── Materiales comprados reemplaza detalle componentes ──
        html.Div(style={"backgroundColor": COLORES["card"], "borderRadius": "12px",
                        "padding": "15px"}, children=[
            html.H3("🛒 Materiales Comprados — Precio Editable para Simular",
                    style={"color": "#4CAF50", "fontSize": "16px", "marginTop": 0}),
            html.P("Modifica el precio de cualquier material y presiona Recalcular.",
                   style={"color": "#7A9BBF", "fontSize": "12px", "marginBottom": "10px"}),
            dash_table.DataTable(
                id="tabla-materiales",
                columns=[
                    {"name": "Tipo",             "id": "Tipo",           "editable": False},
                    {"name": "Componente",        "id": "Componente",     "editable": False},
                    {"name": "Descripción",       "id": "Descripción",    "editable": False},
                    {"name": "Consumo x Und",     "id": "Consumo",        "editable": False},
                    {"name": "Precio (S/)",       "id": "Precio",         "editable": True, "type": "numeric"},
                    {"name": "CM (S/)",           "id": "CM",             "editable": False},
                    {"name": "Tipo de Compra",    "id": "Tipo de Compra", "editable": False},
                    {"name": "MOQ",               "id": "MOQ",            "editable": False},
                    {"name": "LT-días",           "id": "LT-días",        "editable": False},
                ],
                style_header={"backgroundColor": "#1F3864", "color": "white", "fontWeight": "bold"},
                style_cell={"backgroundColor": "#1E2D3D", "color": COLORES["text"],
                            "border": "1px solid #2A3F54", "padding": "8px", "textAlign": "center"},
                style_data_conditional=[
                    {"if": {"column_editable": True},
                     "backgroundColor": "#0D2137", "border": "1px solid #4CAF50"},
                    {"if": {"row_index": "odd"}, "backgroundColor": "#162030"},
                ],
                editable=True, page_size=20,
                filter_action="native", sort_action="native",
            )
        ]),
    ]
)

TAB_HISTORICO = html.Div(
    style={"backgroundColor": COLORES["bg"], "minHeight": "100vh",
           "fontFamily": "'Segoe UI', sans-serif",
           "color": COLORES["text"], "padding": "20px"},
    children=[
        # Header
        html.Div(style={"textAlign": "center", "marginBottom": "25px"}, children=[
            html.H1("📊 Histórico de Costos — Layconsa",
                    style={"color": COLORES["accent"], "margin": 0, "fontSize": "24px"}),
            html.P("Costo real SAP vs Costo teórico por PT",
                   style={"color": "#7A9BBF", "margin": "5px 0 0 0"}),
        ]),

        # Selector
        html.Div(style={"backgroundColor": COLORES["card"], "borderRadius": "12px",
                        "padding": "15px", "marginBottom": "20px"}, children=[
            html.H3("🔍 Seleccionar Producto Terminado",
                    style={"color": COLORES["accent"], "fontSize": "15px", "marginTop": 0}),
            # Filtros línea
            html.Div(style={"display": "flex", "gap": "10px", "marginBottom": "10px"}, children=[
                html.Button("Todos",              id="h-linea-todos",     n_clicks=0,
                    style={"backgroundColor": COLORES["accent"], "color": "#000",
                           "border": "none", "borderRadius": "6px", "padding": "6px 16px",
                           "cursor": "pointer", "fontWeight": "bold", "fontSize": "13px"}),
                html.Button("Útiles",             id="h-linea-utiles",    n_clicks=0,
                    style={"backgroundColor": "#4A5568", "color": "white",
                           "border": "none", "borderRadius": "6px", "padding": "6px 16px",
                           "cursor": "pointer", "fontSize": "13px"}),
                html.Button("Cuadernos",          id="h-linea-cuadernos", n_clicks=0,
                    style={"backgroundColor": "#4A5568", "color": "white",
                           "border": "none", "borderRadius": "6px", "padding": "6px 16px",
                           "cursor": "pointer", "fontSize": "13px"}),
                html.Button("Nuevos Des.",         id="h-linea-nuevos",    n_clicks=0,
                    style={"backgroundColor": "#4A5568", "color": "white",
                           "border": "none", "borderRadius": "6px", "padding": "6px 16px",
                           "cursor": "pointer", "fontSize": "13px"}),
            ]),
            dcc.Dropdown(id="h-selector-pt", placeholder="Selecciona un PT...",
                         style={"color": "#000"}),
        ]),

        # KPIs
        html.Div(id="h-kpis", style={"display": "flex", "gap": "15px",
                                      "flexWrap": "wrap", "marginBottom": "20px"}),

        # Gráfico principal
        html.Div(style={"backgroundColor": COLORES["card"], "borderRadius": "12px",
                        "padding": "15px", "marginBottom": "20px"}, children=[
            html.H3("📈 Costo Unitario Real vs Teórico por Mes",
                    style={"color": COLORES["accent"], "fontSize": "15px", "marginTop": 0}),
            html.P("💡 Click en una barra para ver el desglose por proceso",
                   style={"color": "#7A9BBF", "fontSize": "11px", "margin": "0 0 8px 0"}),
            dcc.Graph(id="h-grafico-apilado", figure=empty_fig()),
        ]),

        # ── Modal desglose por proceso ────────────────────────
        html.Div(id="h-modal-overlay", style={"display": "none"}, children=[
            html.Div(style={
                "position": "fixed", "top": 0, "left": 0, "right": 0, "bottom": 0,
                "backgroundColor": "rgba(0,0,0,0.75)", "zIndex": 9999,
                "display": "flex", "alignItems": "center", "justifyContent": "center",
                "padding": "16px",
            }, children=[
                html.Div(style={
                    "backgroundColor": COLORES["card"], "borderRadius": "16px",
                    "padding": "20px", "width": "100%", "maxWidth": "460px",
                    "border": "1px solid #2A3F54",
                    "boxShadow": "0 20px 60px rgba(0,0,0,0.6)",
                    "maxHeight": "85vh", "overflowY": "auto",
                }, children=[
                    html.Div(style={"display": "flex", "justifyContent": "space-between",
                                    "alignItems": "flex-start", "marginBottom": "16px"}, children=[
                        html.Div(children=[
                            html.H3(id="h-modal-titulo",
                                    style={"color": COLORES["accent"], "margin": 0,
                                           "fontSize": "15px"}),
                            html.P(id="h-modal-subtitulo",
                                   style={"color": "#7A9BBF", "margin": "4px 0 0 0",
                                          "fontSize": "11px"}),
                        ]),
                        html.Button("✕", id="h-modal-cerrar", n_clicks=0,
                            style={"backgroundColor": "transparent", "color": "#7A9BBF",
                                   "border": "none", "fontSize": "20px",
                                   "cursor": "pointer", "lineHeight": "1"}),
                    ]),
                    html.Div(id="h-modal-contenido"),
                ])
            ])
        ]),

        # Gráfico desviación %
        html.Div(style={"backgroundColor": COLORES["card"], "borderRadius": "12px",
                        "padding": "15px", "marginBottom": "20px"}, children=[
            html.H3("📉 Desviación % Real vs Teórico por Mes",
                    style={"color": COLORES["accent"], "fontSize": "15px", "marginTop": 0}),
            dcc.Graph(id="h-grafico-desviacion", figure=empty_fig()),
        ]),

        # Grid: tabla mensual + tabla drivers
        html.Div(style={"display": "grid", "gridTemplateColumns": "1fr 1fr",
                        "gap": "20px", "marginBottom": "20px"}, children=[
            html.Div(style={"backgroundColor": COLORES["card"], "borderRadius": "12px",
                            "padding": "15px"}, children=[
                html.H3("📋 Detalle Mensual de Costos",
                        style={"color": COLORES["accent"], "fontSize": "15px", "marginTop": 0}),
                dash_table.DataTable(
                    id="h-tabla-mensual",
                    columns=[
                        {"name": "Periodo",  "id": "Periodo"},
                        {"name": "Volumen",  "id": "Volumen"},
                        {"name": "CM",       "id": "CM"},
                        {"name": "CIF",      "id": "CIF"},
                        {"name": "MOD",      "id": "MOD"},
                        {"name": "Total Real","id": "Total"},
                        {"name": "Teórico",  "id": "Teorico"},
                        {"name": "Desv. S/", "id": "Desv_Soles"},
                        {"name": "Desv. %",  "id": "Desv_Pct"},
                    ],
                    data=[],
                    style_header={"backgroundColor": "#1F3864", "color": "white",
                                  "fontWeight": "bold", "textAlign": "center"},
                    style_cell={"backgroundColor": "#1E2D3D", "color": COLORES["text"],
                                "border": "1px solid #2A3F54", "padding": "8px",
                                "textAlign": "center", "fontSize": "12px"},
                    style_data_conditional=[
                        {"if": {"row_index": "odd"}, "backgroundColor": "#162030"},
                        {"if": {"filter_query": "{Desv_Pct} contains '+'"},
                         "color": COLORES["red"], "fontWeight": "bold"},
                        {"if": {"filter_query": "{Desv_Pct} contains '-'"},
                         "color": COLORES["green"], "fontWeight": "bold"},
                    ],
                    page_action="none", sort_action="native",
                ),
            ]),
            html.Div(style={"backgroundColor": COLORES["card"], "borderRadius": "12px",
                            "padding": "15px"}, children=[
                html.H3("⚖️ Comparativo Teórico vs Real (Último Mes con Datos)",
                        style={"color": COLORES["accent"], "fontSize": "15px", "marginTop": 0}),
                dash_table.DataTable(
                    id="h-tabla-comparativo",
                    columns=[
                        {"name": "Driver",   "id": "Driver"},
                        {"name": "Teórico",  "id": "Teorico"},
                        {"name": "Real",     "id": "Real"},
                        {"name": "Desv. S/", "id": "Desv_Soles"},
                        {"name": "Desv. %",  "id": "Desv_Pct"},
                    ],
                    data=[],
                    style_header={"backgroundColor": "#1F3864", "color": "white",
                                  "fontWeight": "bold", "textAlign": "center"},
                    style_cell={"backgroundColor": "#1E2D3D", "color": COLORES["text"],
                                "border": "1px solid #2A3F54", "padding": "10px",
                                "textAlign": "center"},
                    style_data_conditional=[
                        {"if": {"row_index": "odd"}, "backgroundColor": "#162030"},
                        {"if": {"filter_query": "{Desv_Pct} contains '+'"},
                         "color": COLORES["red"], "fontWeight": "bold"},
                        {"if": {"filter_query": "{Desv_Pct} contains '-'"},
                         "color": COLORES["green"], "fontWeight": "bold"},
                    ],
                    page_action="none",
                ),
            ]),
        ]),

        # Volumen producido por mes
        html.Div(style={"backgroundColor": COLORES["card"], "borderRadius": "12px",
                        "padding": "15px", "marginBottom": "20px"}, children=[
            html.H3("📦 Volumen Producido por Mes",
                    style={"color": COLORES["accent"], "fontSize": "15px", "marginTop": 0}),
            dcc.Graph(id="h-grafico-volumen", figure=empty_fig()),
        ]),

        # Botón descarga
        html.Div(style={"textAlign": "center", "marginBottom": "20px"}, children=[
            html.Button("⬇️ Descargar Análisis", id="h-btn-descargar",
                style={"backgroundColor": COLORES["orange"], "color": "#000",
                       "fontWeight": "bold", "border": "none", "borderRadius": "8px",
                       "padding": "12px 30px", "cursor": "pointer", "fontSize": "15px"}),
            dcc.Download(id="h-descarga"),
        ]),
    ]
)

app.layout = html.Div(
    style={"backgroundColor": COLORES["bg"], "minHeight": "100vh",
           "fontFamily": "'Segoe UI', sans-serif",
           "color": COLORES["text"], "padding": "0"},
    children=[
        dcc.Tabs(
            id="tabs-main", value="tab-costos",
            colors={"border": "#2A3F54", "primary": COLORES["accent"],
                    "background": COLORES["card"]},
            parent_style={"backgroundColor": COLORES["bg"]},
            children=[
                dcc.Tab(label="📦 Reporte de Costos", value="tab-costos",
                    style={"backgroundColor": COLORES["card"],
                           "color": COLORES["text"], "padding": "10px 20px"},
                    selected_style={"backgroundColor": COLORES["bg"],
                        "color": COLORES["accent"], "fontWeight": "bold",
                        "padding": "10px 20px",
                        "borderTop": "3px solid " + COLORES["accent"]}),
                dcc.Tab(label="📊 Histórico SAP", value="tab-historico",
                    style={"backgroundColor": COLORES["card"],
                           "color": COLORES["text"], "padding": "10px 20px"},
                    selected_style={"backgroundColor": COLORES["bg"],
                        "color": COLORES["accent"], "fontWeight": "bold",
                        "padding": "10px 20px",
                        "borderTop": "3px solid " + COLORES["accent"]}),
            ]
        ),
        html.Div(id="tab-content")
    ]
)


@app.callback(
    Output("tabla-simulador", "data"),
    Output("tabla-simulador-otros", "data"),
    Input("selector-pt", "value"),
)
def cargar_simuladores(codigo_pt):
    # Inyección
    maquinas = get_maquinas_inyeccion(codigo_pt)
    rows_iny = []
    for m in maquinas:
        cant_base = round((3600 / m["T.Ciclo"]) * m["Cav.Oper"] * 24, 2)                     if m["T.Ciclo"] > 0 else 0
        rows_iny.append({
            "Maquina":   m["Maquina"],   "T.Ciclo":   m["T.Ciclo"],
            "Cav.Oper":  m["Cav.Oper"],  "Cav.Tot":   m["Cav.Tot"],
            "Cant.Base": cant_base,       "Tarifa Maq":m["Tarifa Maq"],
            "Tarifa MO": m["Tarifa MO"],
        })
    # Otros procesos
    otros    = get_semis_otros_procesos(codigo_pt)
    rows_otros = []
    for s in otros:
        rows_otros.append({
            "Proceso":      s["Proceso"],
            "Maquina":      s["Maquina"],
            "Cantidad Base":s["Cantidad Base"],
            "T.MO":         s["T.MO"],
            "T.Maq":        s["T.Maq"],
            "Cant.Opr":     s["Cant.Opr"],
            "Tarifa Maq":   s["Tarifa Maq"],
            "Tarifa MO":    s["Tarifa MO"],
        })
    return rows_iny, rows_otros


# ── Filtro por línea ────────────────────────────────────────
@app.callback(
    Output("selector-pt", "options"),
    Input("linea-todos",     "n_clicks"),
    Input("linea-utiles",    "n_clicks"),
    Input("linea-cuadernos", "n_clicks"),
    Input("linea-nuevos",    "n_clicks"),
    prevent_initial_call=False,
)
def filtrar_linea(n_todos, n_utiles, n_cuadernos, n_nuevos):
    from dash import callback_context
    ctx     = callback_context
    trigger = ctx.triggered[0]["prop_id"].split(".")[0] if ctx.triggered else "linea-todos"
    if trigger == "linea-utiles":
        df_f = lista_pt_dd[lista_pt_dd["Línea"] == "Útiles"]
    elif trigger == "linea-cuadernos":
        df_f = lista_pt_dd[lista_pt_dd["Línea"] == "Cuadernos"]
    elif trigger == "linea-nuevos":
        df_f = lista_pt_dd[lista_pt_dd["Línea"] == "Nuevos Desarrollos"]
    else:
        df_f = lista_pt_dd
    return [{"label": f"{r['Código PT']} — {r['Descripción PT']}", "value": r["Código PT"]}
            for _, r in df_f.iterrows()]


@app.callback(
    Output("tabla-materiales", "data"),
    Input("selector-pt", "value"),
)
def cargar_materiales(codigo_pt):
    """Carga todos los materiales COMPRADOS del PT en todos los niveles."""
    visitados = set()
    materiales = {}  # key=componente para evitar duplicados

    def buscar(codigo):
        if codigo in visitados:
            return
        visitados.add(codigo)
        hijos = df_exp[(df_exp["Código Semi"] == codigo) &
                       (df_exp["Código PT"] == codigo_pt)]
        for _, row in hijos.iterrows():
            comp    = str(row["Componente"])
            familia = str(row.get("Familia", comp[:3])).strip()
            if not es_fabricado(familia):
                if comp not in materiales:
                    # Cruzar con hoja Materiales
                    mat_row = df_mat[df_mat["Codigo"] == comp]
                    tipo_compra = str(mat_row["TIPO DE COMPRA"].iloc[0]) if not mat_row.empty and "TIPO DE COMPRA" in mat_row.columns else ""
                    moq         = mat_row["MOQ"].iloc[0]    if not mat_row.empty and "MOQ" in mat_row.columns else ""
                    lt_dias     = mat_row["LT-días"].iloc[0] if not mat_row.empty and "LT-días" in mat_row.columns else ""
                    tipo        = str(mat_row["Tipo"].iloc[0]) if not mat_row.empty and "Tipo" in mat_row.columns else ""
                    materiales[comp] = {
                        "Tipo":          tipo,
                        "Componente":    comp,
                        "Descripción":   str(row.get("Descripción Componente", "")),
                        "Precio":        float(row.get("Costo estandar", 0)),
                        "Tipo de Compra":tipo_compra,
                        "MOQ":           moq,
                        "LT-días":       lt_dias,
                    }
            if es_fabricado(familia):
                buscar(comp)

    buscar(codigo_pt)

    # Calcular consumo x unidad = Cantidad Total Requerida acumulada / cant_base_pt
    cant_base_pt = float(df_exp[df_exp["Código Semi"] == codigo_pt]["Cantidad Base"].iloc[0])                    if not df_exp[df_exp["Código Semi"] == codigo_pt].empty else 1
    if cant_base_pt == 0: cant_base_pt = 1

    # Acumular cantidades de cada componente en toda la explosión del PT
    filas_exp_pt = df_exp[df_exp["Código PT"] == codigo_pt]
    consumo_map  = filas_exp_pt[filas_exp_pt["Familia"].apply(lambda f: not es_fabricado(str(f)))]                   .groupby("Componente")["Cantidad Total Requerida"].sum() / cant_base_pt

    for comp in materiales:
        consumo = float(consumo_map.get(comp, 0))
        precio  = float(materiales[comp].get("Precio", 0))
        materiales[comp]["Consumo"] = round(consumo, 6)
        materiales[comp]["CM"]      = round(consumo * precio, 4)

    return sorted(materiales.values(), key=lambda x: x["Tipo"])


@app.callback(
    Output("kpis",                "children"),
    Output("grafico-cascada",     "figure"),
    Output("grafico-cascada-pct", "figure"),
    Output("grafico-donut",       "figure"),
    Output("grafico-donut-soles", "figure"),
    Output("grafico-pareto",      "figure"),
    Output("tabla-cascada-detalle", "data"),
    Output("grafico-dona-cascada", "figure"),
    Output("msg-simulador",       "children"),
    Input("selector-pt",          "value"),
    Input("btn-recalcular",       "n_clicks"),
    State("tabla-simulador",      "data"),
    State("tabla-simulador-otros","data"),
    State("tabla-materiales",     "data"),
    prevent_initial_call=True,
)
def actualizar(codigo_pt, n_clicks, datos_simulador, datos_otros, datos_materiales):
    df_tie_sim = df_tie.copy()
    # Aplicar cambios de inyección por máquina
    if datos_simulador:
        for row in datos_simulador:
            maquina  = str(row.get("Maquina", ""))
            t_ciclo  = float(row.get("T.Ciclo", 0) or 0)
            cav_oper = float(row.get("Cav.Oper", 0) or 0)
            if t_ciclo > 0 and cav_oper > 0 and maquina:
                nueva_base = (3600 / t_ciclo) * cav_oper * 24
                mask = df_tie_sim["Maquina"].astype(str).str.strip() == maquina
                df_tie_sim.loc[mask, "Cantidad Base"] = nueva_base
                tarifa_maq_sim = float(row.get("Tarifa Maq", 0) or 0)
                if tarifa_maq_sim > 0: df_tie_sim.loc[mask, "Tarifa Maquina"] = tarifa_maq_sim
    # Aplicar cambios de otros procesos por máquina
    if datos_otros:
        for row in datos_otros:
            maquina    = str(row.get("Maquina", ""))
            nueva_base = float(row.get("Cantidad Base", 0) or 0)
            nuevo_tmo  = float(row.get("T.MO",          0) or 0)
            nuevo_tmaq = float(row.get("T.Maq",         0) or 0)
            if maquina and nueva_base > 0:
                # Aplica a todos los semis que usan esta máquina
                mask = df_tie_sim["Maquina"].astype(str).str.strip() == maquina
                df_tie_sim.loc[mask, "Cantidad Base"] = nueva_base
                tarifa_maq_sim = float(row.get("Tarifa Maq", 0) or 0)
                if tarifa_maq_sim > 0: df_tie_sim.loc[mask, "Tarifa Maquina"] = tarifa_maq_sim
                if nuevo_tmo  > 0: df_tie_sim.loc[mask, "T.MO"]  = nuevo_tmo
                if nuevo_tmaq > 0: df_tie_sim.loc[mask, "T.Maq"] = nuevo_tmaq

    # Aplicar precios modificados de materiales
    df_exp_sim = df_exp.copy()
    if datos_materiales:
        for row in datos_materiales:
            comp  = str(row.get("Componente", ""))
            precio = float(row.get("Precio", 0) or 0)
            if comp and precio > 0:
                mask = df_exp_sim["Componente"] == comp
                df_exp_sim.loc[mask, "Costo estandar"] = precio

    filas_pt   = df_exp_sim[df_exp_sim["Código PT"]   == str(codigo_pt)]
    filas_semi = df_exp_sim[df_exp_sim["Código Semi"] == str(codigo_pt)]
    print(f"DEBUG codigo_pt={codigo_pt}")
    print(f"DEBUG filas donde Código PT={codigo_pt}: {len(filas_pt)}")
    print(f"DEBUG filas donde Código Semi={codigo_pt}: {len(filas_semi)}")
    print(f"DEBUG primeros Código Semi únicos: {df_exp_sim['Código Semi'].unique()[:5]}")
    resumen_sim, detalle_sim, _ = explotar_pt(codigo_pt, df_exp_sim, df_tie_sim)
    print(f"DEBUG resumen_sim={resumen_sim}")

    cant_base_pt = float(df_exp[df_exp["Código Semi"] == codigo_pt]["Cantidad Base"].iloc[0]) \
                   if not df_exp[df_exp["Código Semi"] == codigo_pt].empty else 1
    if cant_base_pt == 0:
        cant_base_pt = 1

    filas = []
    for proceso, valores in resumen_sim.items():
        for tipo, monto in [("CM", valores["CM"]), ("CIF", valores["CIF"]), ("MOD", valores["MOD"])]:
            if monto > 0:
                filas.append({
                    "Proceso": proceso, "Tipo de Costo": f"{tipo} {proceso}",
                    "Costo Unitario": monto / cant_base_pt,
                })

    df_pt = pd.DataFrame(filas)
    if df_pt.empty:
        df_pt = df_resumen[df_resumen["Código PT"] == codigo_pt].copy()
    else:
        total = df_pt["Costo Unitario"].sum()
        df_pt["% del Total"] = df_pt["Costo Unitario"] / total if total > 0 else 0

    msg    = f"✅ Recalculado — {datetime.now().strftime('%H:%M:%S')}" if n_clicks else ""
    df_det = df_detalle[df_detalle["Código PT"] == codigo_pt].copy()
    total  = df_pt["Costo Unitario"].sum()
    tot_cm  = df_pt[df_pt["Tipo de Costo"].str.startswith("CM")]["Costo Unitario"].sum()
    tot_mod = df_pt[df_pt["Tipo de Costo"].str.startswith("MOD")]["Costo Unitario"].sum()
    tot_cif = df_pt[df_pt["Tipo de Costo"].str.startswith("CIF")]["Costo Unitario"].sum()

    def kpi(titulo, valor, color, sub=None):
        return html.Div(
            style={"backgroundColor": COLORES["card"], "borderLeft": f"4px solid {color}",
                   "borderRadius": "10px", "padding": "15px 20px",
                   "flex": "1", "minWidth": "160px"},
            children=[
                html.P(titulo, style={"margin": 0, "fontSize": "12px", "color": "#7A9BBF"}),
                html.H2(valor, style={"margin": "5px 0 0 0", "color": color, "fontSize": "18px"}),
                html.P(sub or "", style={"margin": "3px 0 0 0", "fontSize": "11px",
                                         "color": "#7A9BBF"}),
            ]
        )

    kpis_elem = [
        kpi("💰 Costo x Und", total,   COLORES["accent"]),
        kpi("🧱 CM Total",    tot_cm,  COLORES["CM"]),
        kpi("👷 MOD Total",   tot_mod, COLORES["MOD"]),
        kpi("⚙️ CIF Total",   tot_cif, COLORES["CIF"]),
    ]

    labels       = list(df_pt["Tipo de Costo"]) + ["TOTAL"]
    valores      = list(df_pt["Costo Unitario"]) + [total]
    measures     = ["relative"] * len(df_pt) + ["total"]

    fig_cas = go.Figure(go.Waterfall(
        x=labels, y=valores, measure=measures,
        text=[f"S/ {v:.4f}" for v in valores], textposition="outside",
        increasing=dict(marker_color=COLORES["CM"]),
        totals=dict(marker_color=COLORES["TOTAL"]),
        connector=dict(line=dict(color="#4A5568", width=1)),
        hovertemplate="<b>%{x}</b><br>S/ %{y:.6f}<extra></extra>"
    ))
    fig_cas.update_layout(template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)",
                          plot_bgcolor="rgba(0,0,0,0)",
                          margin=dict(l=10, r=10, t=30, b=80),
                          xaxis_tickangle=-35, showlegend=False)

    pcts         = list(df_pt["% del Total"] * 100) + [100.0]
    measures_pct = ["relative"] * len(df_pt) + ["total"]
    fig_cas_pct  = go.Figure(go.Waterfall(
        x=labels, y=pcts, measure=measures_pct,
        text=[f"{v:.1f}%" for v in pcts], textposition="outside",
        increasing=dict(marker_color=COLORES["MOD"]),
        totals=dict(marker_color=COLORES["TOTAL"]),
        connector=dict(line=dict(color="#4A5568", width=1)),
        hovertemplate="<b>%{x}</b><br>%{y:.1f}%<extra></extra>"
    ))
    fig_cas_pct.update_layout(template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)",
                              plot_bgcolor="rgba(0,0,0,0)", yaxis=dict(ticksuffix="%"),
                              margin=dict(l=10, r=10, t=30, b=80),
                              xaxis_tickangle=-35, showlegend=False)

    resumen_proc = df_pt.groupby("Proceso")["Costo Unitario"].sum().reset_index()
    paleta       = ["#2196F3", "#4CAF50", "#FF9800", "#E91E63", "#9C27B0", "#00BCD4", "#FF5722"]

    # Dona en porcentaje
    fig_don = go.Figure(go.Pie(
        labels=resumen_proc["Proceso"], values=resumen_proc["Costo Unitario"],
        hole=0.55, marker_colors=paleta[:len(resumen_proc)],
        textinfo="label+percent",
        hovertemplate="<b>%{label}</b><br>S/ %{value:.6f}<br>%{percent}<extra></extra>"
    ))
    fig_don.update_layout(template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)",
                          margin=dict(l=10, r=10, t=10, b=10))

    # Dona CM + CIF + MOD agrupados
    resumen_tipo = df_pt.copy()
    resumen_tipo["Tipo"] = resumen_tipo["Tipo de Costo"].str.split(" ").str[0]
    resumen_tipo = resumen_tipo.groupby("Tipo")["Costo Unitario"].sum().reset_index()
    fig_don_soles = go.Figure(go.Pie(
        labels=resumen_tipo["Tipo"],
        values=resumen_tipo["Costo Unitario"],
        hole=0.55,
        marker_colors=["#2196F3", "#FF9800", "#4CAF50"],
        textinfo="label+percent",
        texttemplate="<b>%{label}</b><br>%{percent}",
        hovertemplate="<b>%{label}</b><br>S/ %{value:.6f}<br>%{percent}<extra></extra>"
    ))
    fig_don_soles.update_layout(
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=10, r=10, t=10, b=10)
    )

    # ── Pareto ─────────────────────────────────────────────
    df_pareto = df_pt[["Tipo de Costo","Costo Unitario"]].copy()
    df_pareto = df_pareto.sort_values("Costo Unitario", ascending=False).reset_index(drop=True)
    df_pareto["Acumulado %"] = (df_pareto["Costo Unitario"].cumsum() /
                                 df_pareto["Costo Unitario"].sum() * 100)

    fig_pareto = go.Figure()
    fig_pareto.add_trace(go.Bar(
        x=df_pareto["Tipo de Costo"], y=df_pareto["Costo Unitario"],
        name="Costo Unitario", marker_color=COLORES["CM"],
        text=[f"S/ {v:.4f}" for v in df_pareto["Costo Unitario"]],
        textposition="outside",
        hovertemplate="<b>%{x}</b><br>S/ %{y:.6f}<extra></extra>"
    ))
    fig_pareto.add_trace(go.Scatter(
        x=df_pareto["Tipo de Costo"], y=df_pareto["Acumulado %"],
        name="% Acumulado", yaxis="y2", mode="lines+markers",
        line=dict(color=COLORES["TOTAL"], width=2),
        marker=dict(size=6),
        hovertemplate="<b>%{x}</b><br>%{y:.1f}%<extra></extra>"
    ))
    fig_pareto.update_layout(
        template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        yaxis=dict(title="Costo Unitario (S/)"),
        yaxis2=dict(title="% Acumulado", overlaying="y", side="right",
                    range=[0, 110], ticksuffix="%"),
        margin=dict(l=10, r=60, t=30, b=100),
        xaxis_tickangle=-35, showlegend=True,
        legend=dict(orientation="h", y=1.1)
    )

    # ── Tabla cascada detallada ────────────────────────────
    filas_cascada = construir_tabla_cascada(
        codigo_pt, df_exp_sim, df_tie_sim, resumen_sim, cant_base_pt
    )

    # Dona cascada — excluye fila TOTAL PT
    cascada_sin_total = [f for f in filas_cascada if f["Proceso"] != "TOTAL PT"]
    labels_don_cas = [f"{f['Proceso']} — {f['Componente']}" for f in cascada_sin_total]
    values_don_cas = [float(str(f["Costo"]).replace("S/ ", "")) for f in cascada_sin_total]
    paleta_ext     = paleta * 5
    fig_dona_cascada = go.Figure(go.Pie(
        labels=labels_don_cas, values=values_don_cas,
        hole=0.55, marker_colors=paleta_ext[:len(labels_don_cas)],
        textinfo="percent",
        hovertemplate="<b>%{label}</b><br>%{percent}<extra></extra>"
    ))
    fig_dona_cascada.update_layout(
        template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=10, r=10, t=10, b=10)
    )

    return (kpis_elem, fig_cas, fig_cas_pct, fig_don, fig_don_soles,
            fig_pareto, filas_cascada, fig_dona_cascada, msg)


# ── Descarga PT actual ──────────────────────────────────────
@app.callback(
    Output("descarga-pt",    "data"),
    Input("btn-descargar-pt","n_clicks"),
    State("selector-pt",     "value"),
    State("tabla-materiales","data"),
    prevent_initial_call=True,
)
def descargar_pt(n_clicks, codigo_pt, datos_mat):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    if not codigo_pt: return None

    # Info del PT
    row_pt    = df_resumen[df_resumen["Código PT"] == codigo_pt]
    desc_pt   = row_pt["Descripción PT"].iloc[0] if not row_pt.empty else codigo_pt
    cant_base = float(df_exp[df_exp["Código Semi"]==codigo_pt]["Cantidad Base"].iloc[0])                 if not df_exp[df_exp["Código Semi"]==codigo_pt].empty else 1

    # Cascada del PT
    resumen_pt, _, _ = explotar_pt(codigo_pt, df_exp, df_tie)
    filas_cas = construir_tabla_cascada(codigo_pt, df_exp, df_tie, resumen_pt, cant_base)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Hoja 1: Resumen KPIs
        total = row_pt["Costo Unitario"].sum() if not row_pt.empty else 0
        cm    = row_pt[row_pt["Tipo de Costo"].str.startswith("CM")]["Costo Unitario"].sum()  if not row_pt.empty else 0
        cif   = row_pt[row_pt["Tipo de Costo"].str.startswith("CIF")]["Costo Unitario"].sum() if not row_pt.empty else 0
        mod   = row_pt[row_pt["Tipo de Costo"].str.startswith("MOD")]["Costo Unitario"].sum() if not row_pt.empty else 0
        df_kpi = pd.DataFrame([
            {"Concepto": "Código PT",       "Valor": codigo_pt},
            {"Concepto": "Descripción",     "Valor": desc_pt},
            {"Concepto": "Cantidad Base",   "Valor": int(cant_base)},
            {"Concepto": "Costo Unitario",  "Valor": round(total, 6)},
            {"Concepto": "CM Unitario",     "Valor": round(cm,    6)},
            {"Concepto": "CIF Unitario",    "Valor": round(cif,   6)},
            {"Concepto": "MOD Unitario",    "Valor": round(mod,   6)},
        ])
        df_kpi.to_excel(writer, sheet_name="Resumen", index=False)

        # Hoja 2: Cascada
        pd.DataFrame(filas_cas).to_excel(writer, sheet_name="Cascada Detallada", index=False)

        # Hoja 3: Materiales Comprados
        if datos_mat:
            pd.DataFrame(datos_mat).to_excel(writer, sheet_name="Materiales Comprados", index=False)

        # Formato
        hf    = Font(bold=True, color="FFFFFF")
        fills = {
            "Resumen":             PatternFill("solid", fgColor="1F3864"),
            "Cascada Detallada":   PatternFill("solid", fgColor="1F5C2E"),
            "Materiales Comprados":PatternFill("solid", fgColor="2E75B6"),
        }
        for sn, ws in writer.sheets.items():
            for cell in ws[1]:
                cell.font = hf
                cell.fill = fills.get(sn, fills["Resumen"])
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                ml = max((len(str(c.value)) for c in col if c.value), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml+4, 50)

    output.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return dcc.send_bytes(output.read(), filename=f"PT_{codigo_pt}_{ts}.xlsx")


@app.callback(
    Output("descarga-excel", "data"),
    Input("btn-descargar", "n_clicks"),
    prevent_initial_call=True,
)
def descargar_excel(n_clicks):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # ── Hoja 1: Resumen Global ──────────────────────────
        resumen_global = df_resumen[
            ["Código PT", "Descripción PT", "Proceso", "Tipo de Costo", "Costo Unitario", "% del Total"]
        ].copy()
        totales = df_resumen.groupby(["Código PT", "Descripción PT"])["Costo Unitario"].sum().reset_index()
        totales["Proceso"]       = "TOTAL"
        totales["Tipo de Costo"] = "TOTAL PT"
        totales["% del Total"]   = 1.0
        resumen_exp = pd.concat([resumen_global, totales], ignore_index=True)
        resumen_exp = resumen_exp.sort_values(["Código PT", "Tipo de Costo"])
        resumen_exp["% del Total"]    = resumen_exp["% del Total"].map("{:.1%}".format)
        resumen_exp["Costo Unitario"] = resumen_exp["Costo Unitario"].map("{:.6f}".format)
        resumen_exp.to_excel(writer, sheet_name="Resumen Global", index=False)

        # ── Hoja 2: Detalle Componentes ─────────────────────
        cols_det = ["Código PT", "Descripción PT", "Código Semi", "Descripción Semi",
                    "Componente", "Descripción Componente", "Familia", "Tipo",
                    "Proceso", "Cantidad Total Req", "Costo Calculado",
                    "CM", "CIF", "MOD", "Total"]
        cols_det = [c for c in cols_det if c in df_detalle.columns]
        df_detalle[cols_det].to_excel(writer, sheet_name="Detalle Componentes", index=False)

        # ── Hoja 3: Cascada Detallada por PT ────────────────
        filas_cascada_total = []
        for codigo_pt in df_resumen["Código PT"].unique():
            desc_pt = df_resumen[df_resumen["Código PT"] == codigo_pt]["Descripción PT"].iloc[0]
            cant_base_pt = float(df_exp[df_exp["Código Semi"] == codigo_pt]["Cantidad Base"].iloc[0])                            if not df_exp[df_exp["Código Semi"] == codigo_pt].empty else 1
            if cant_base_pt == 0:
                cant_base_pt = 1
            resumen_pt, _, _ = explotar_pt(codigo_pt, df_exp, df_tie)
            filas_pt = construir_tabla_cascada(codigo_pt, df_exp, df_tie, resumen_pt, cant_base_pt)
            for f in filas_pt:
                filas_cascada_total.append({
                    "Código PT":      codigo_pt,
                    "Descripción PT": desc_pt,
                    "Proceso":        f["Proceso"],
                    "Componente":     f["Componente"],
                    "Costo":          f["Costo"],
                    "%TG":            f["%TG"],
                })
        df_cascada = pd.DataFrame(filas_cascada_total)
        df_cascada.to_excel(writer, sheet_name="Cascada Detallada", index=False)

        # ── Formato ─────────────────────────────────────────
        header_font = Font(bold=True, color="FFFFFF")
        fills = {
            "Resumen Global":      PatternFill("solid", fgColor="1F3864"),
            "Detalle Componentes": PatternFill("solid", fgColor="2E75B6"),
            "Cascada Detallada":   PatternFill("solid", fgColor="1F5C2E"),
        }
        for sheet_name, ws in writer.sheets.items():
            fill = fills.get(sheet_name, PatternFill("solid", fgColor="1F3864"))
            for cell in ws[1]:
                cell.font      = header_font
                cell.fill      = fill
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)

    output.seek(0)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return dcc.send_bytes(output.read(), filename=f"Reporte_Costos_{timestamp}.xlsx")


# ── Render pestaña activa ───────────────────────────────────
@app.callback(
    Output("tab-content", "children"),
    Input("tabs-main", "value"),
)
def render_tab(tab):
    if tab == "tab-historico":
        return TAB_HISTORICO
    return TAB_COSTOS


# ── Callbacks Histórico ──────────────────────────────────────
@app.callback(
    Output("h-selector-pt", "options"),
    Input("h-linea-todos",     "n_clicks"),
    Input("h-linea-utiles",    "n_clicks"),
    Input("h-linea-cuadernos", "n_clicks"),
    Input("h-linea-nuevos",    "n_clicks"),
    prevent_initial_call=False,
)
def filtrar_pts(n1, n2, n3, n4):
    from dash import callback_context
    ctx     = callback_context
    trigger = ctx.triggered[0]["prop_id"].split(".")[0] if ctx.triggered else "h-linea-todos"
    if trigger == "h-linea-utiles":
        df_f = pts_cubo[pts_cubo["Linea"]=="Útiles"]
    elif trigger == "h-linea-cuadernos":
        df_f = pts_cubo[pts_cubo["Linea"]=="Cuadernos"]
    elif trigger == "h-linea-nuevos":
        df_f = pts_cubo[pts_cubo["Linea"]=="Nuevos Desarrollos"]
    else:
        df_f = pts_cubo
    return [{"label": f"{r['Cod_PT']} — {r['Desc']}", "value": r["Cod_PT"]}
            for _, r in df_f.iterrows()]


# ── Callback 2: Actualizar dashboard ────────────────────────
@app.callback(
    Output("h-kpis",               "children"),
    Output("h-grafico-apilado",    "figure"),
    Output("h-grafico-desviacion", "figure"),
    Output("h-grafico-volumen",    "figure"),
    Output("h-tabla-mensual",      "data"),
    Output("h-tabla-comparativo",  "data"),
    Input("h-selector-pt",         "value"),
    prevent_initial_call=True,
)
def actualizar(codigo_pt):
    if not codigo_pt:
        return [], empty_fig(), empty_fig(), empty_fig(), [], []

    df_hist = get_costo_real_mensual(codigo_pt)
    cm_t, cif_t, mod_t, total_t = get_costo_teorico(codigo_pt)

    # ── KPIs ────────────────────────────────────────────────
    def kpi(titulo, valor, color, sub=None):
        return html.Div(
            style={"backgroundColor": COLORES["card"],
                   "borderLeft": f"4px solid {color}",
                   "borderRadius": "10px", "padding": "12px 18px",
                   "flex": "1", "minWidth": "150px"},
            children=[
                html.P(titulo, style={"margin": 0, "fontSize": "11px", "color": "#7A9BBF"}),
                html.H3(valor, style={"margin": "4px 0 0 0", "color": color, "fontSize": "16px"}),
                html.P(sub or "", style={"margin": 0, "fontSize": "10px", "color": "#7A9BBF"}),
            ]
        )

    if not df_hist.empty:
        ultimo = df_hist[df_hist["Con_Datos"]].iloc[-1] if df_hist["Con_Datos"].any() else df_hist.iloc[-1]
        real_total = ultimo["Total"]
        desv_s = real_total - total_t
        desv_p = (desv_s/total_t*100) if total_t > 0 else 0
        vol_total = df_hist["Volumen"].sum()
        color_desv = COLORES["red"] if desv_s > 0 else COLORES["green"]
        kpis = [
            kpi("💰 Costo Teórico",    f"S/ {total_t:,.6f}", COLORES["accent"], "por unidad"),
            kpi("📊 Costo Real (últ.)", f"S/ {real_total:,.6f}", COLORES["orange"],
                f"Período: {ultimo['Periodo']}"),
            kpi("📉 Desviación",        f"{desv_p:+.1f}%", color_desv,
                f"S/ {desv_s:+,.6f}/und"),
            kpi("🧱 CM Teórico",  f"S/ {cm_t:,.6f}",  COLORES["cm"],
                f"{cm_t/total_t*100:.1f}% del total" if total_t>0 else ""),
            kpi("⚙️ CIF Teórico", f"S/ {cif_t:,.6f}", COLORES["cif"],
                f"{cif_t/total_t*100:.1f}% del total" if total_t>0 else ""),
            kpi("👷 MOD Teórico", f"S/ {mod_t:,.6f}", COLORES["mod"],
                f"{mod_t/total_t*100:.1f}% del total" if total_t>0 else ""),
            kpi("📦 Vol. Total",         f"{vol_total:,.0f}", "#9C27B0", "unidades producidas"),
        ]
    else:
        kpis = [kpi("💰 Costo Teórico", f"S/ {total_t:,.6f}", COLORES["accent"],
                    "Sin datos reales")]

    # ── Gráfico apilado ──────────────────────────────────────
    if not df_hist.empty:
        periodos = df_hist["Periodo"].tolist()
        marker_syms = ["circle" if c else "diamond" for c in df_hist["Con_Datos"]]

        totales   = df_hist["Total"].values
        cm_vals   = df_hist["CM"].values
        cif_vals  = df_hist["CIF"].values
        mod_vals  = df_hist["MOD"].values

        # customdata: [total, pct_cm, pct_cif, pct_mod] por periodo
        cm_pcts  = [round(cm/t*100,1) if t>0 else 0 for cm,t in zip(cm_vals,totales)]
        cif_pcts = [round(cif/t*100,1) if t>0 else 0 for cif,t in zip(cif_vals,totales)]
        mod_pcts = [round(mod/t*100,1) if t>0 else 0 for mod,t in zip(mod_vals,totales)]

        fig_ap = go.Figure()
        fig_ap.add_trace(go.Bar(
            name="CM Real", x=periodos, y=cm_vals,
            marker_color=COLORES["cm"], opacity=0.85,
            text=[f"S/ {v:.4f}" for v in cm_vals],
            textposition="inside", textfont=dict(color="white", size=10),
            customdata=list(zip([0]*len(periodos), cm_pcts)),
            hovertemplate="<b>CM</b>: S/ %{y:.4f} (%{customdata[1]:.1f}%)<extra></extra>",
        ))
        fig_ap.add_trace(go.Bar(
            name="CIF Real", x=periodos, y=cif_vals,
            marker_color=COLORES["cif"], opacity=0.85,
            text=[f"S/ {v:.4f}" for v in cif_vals],
            textposition="inside", textfont=dict(color="white", size=10),
            customdata=list(zip([1]*len(periodos), cif_pcts)),
            hovertemplate="<b>CIF</b>: S/ %{y:.4f} (%{customdata[1]:.1f}%)<extra></extra>",
        ))
        fig_ap.add_trace(go.Bar(
            name="MOD Real", x=periodos, y=mod_vals,
            marker_color=COLORES["mod"], opacity=0.85,
            text=[f"S/ {v:.4f}" for v in mod_vals],
            textposition="inside", textfont=dict(color="white", size=10),
            customdata=list(zip([2]*len(periodos), mod_pcts)),
            hovertemplate="<b>MOD</b>: S/ %{y:.4f} (%{customdata[1]:.1f}%)<extra></extra>",
        ))
        # Traza invisible para mostrar total encima
        fig_ap.add_trace(go.Bar(
            name="Total", x=periodos, y=[0]*len(periodos),
            marker_color="rgba(0,0,0,0)",
            text=[f"S/ {t:.4f}" for t in totales],
            textposition="outside",
            textfont=dict(color="#00C8FF", size=11),
            showlegend=False,
            hoverinfo="skip",
        ))
        # Línea costo teórico
        fig_ap.add_trace(go.Scatter(
            name="Costo Teórico", x=periodos,
            y=[total_t]*len(periodos),
            mode="lines", line=dict(color=COLORES["teorico"], width=2.5, dash="dash"),
            hovertemplate="<b>Teórico</b>: S/ %{y:.4f}<extra></extra>",
        ))
        # Puntos sin producción
        sin_prod = df_hist[~df_hist["Con_Datos"]]
        if not sin_prod.empty:
            fig_ap.add_trace(go.Scatter(
                name="Sin producción (fallback)",
                x=sin_prod["Periodo"], y=sin_prod["Total"],
                mode="markers",
                marker=dict(symbol="diamond", size=10, color="#9C27B0",
                            line=dict(width=1.5, color="white")),
            ))
        fig_ap.update_layout(
            template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)", barmode="stack",
            legend=dict(orientation="h", y=1.1),
            yaxis_title="Costo Unitario (S/)",
            margin=dict(l=10,r=10,t=50,b=10),
            hovermode="x unified",
            hoverlabel=dict(bgcolor="#1A2633", font_color="#00C8FF",
                           font_size=13, bordercolor="#00C8FF"),
        )
    else:
        fig_ap = empty_fig()

    # ── Gráfico desviación % ─────────────────────────────────
    if not df_hist.empty and total_t > 0:
        desv_pct = ((df_hist["Total"] - total_t) / total_t * 100).round(1)
        colores_barras = [COLORES["red"] if v > 0 else COLORES["green"] for v in desv_pct]
        fig_desv = go.Figure()
        fig_desv.add_trace(go.Bar(
            x=df_hist["Periodo"], y=desv_pct,
            marker_color=colores_barras, name="Desviación %",
            text=[f"{v:+.1f}%" for v in desv_pct], textposition="outside",
        ))
        fig_desv.add_hline(y=0, line_color="white", line_width=1, opacity=0.5)
        fig_desv.update_layout(
            template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            yaxis_title="Desviación (%)",
            margin=dict(l=10,r=10,t=30,b=10),
        )
    else:
        fig_desv = empty_fig()

    # ── Gráfico volumen ──────────────────────────────────────
    if not df_hist.empty:
        fig_vol = go.Figure()
        fig_vol.add_trace(go.Bar(
            x=df_hist["Periodo"],
            y=df_hist["Volumen"],
            marker_color=COLORES["accent"], opacity=0.8,
            text=[f"{int(v):,}" if v > 0 else "—" for v in df_hist["Volumen"]],
            textposition="outside",
        ))
        fig_vol.update_layout(
            template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            yaxis_title="Unidades",
            margin=dict(l=10,r=10,t=30,b=10),
        )
    else:
        fig_vol = empty_fig()

    # ── Tabla mensual ────────────────────────────────────────
    tabla_data = []
    if not df_hist.empty:
        for _, row in df_hist.iterrows():
            desv_s = row["Total"] - total_t
            desv_p = (desv_s/total_t*100) if total_t > 0 else 0
            tabla_data.append({
                "Periodo":    row["Periodo"],
                "Volumen":    f"{int(row['Volumen']):,}" if row["Volumen"] > 0 else "—",
                "CM":         f"{row['CM']:,.6f}",
                "CIF":        f"{row['CIF']:,.6f}",
                "MOD":        f"{row['MOD']:,.6f}",
                "Total":      f"{row['Total']:,.6f}",
                "Teorico":    f"{total_t:,.6f}",
                "Desv_Soles": f"{desv_s:+,.6f}",
                "Desv_Pct":   f"{desv_p:+.1f}%",
            })

    # ── Tabla comparativo ────────────────────────────────────
    comp_data = []
    if not df_hist.empty:
        ultimo = df_hist[df_hist["Con_Datos"]].iloc[-1] if df_hist["Con_Datos"].any() else df_hist.iloc[-1]
        for driver, teo, real_v in [
            ("CM",    cm_t,    ultimo["CM"]),
            ("CIF",   cif_t,   ultimo["CIF"]),
            ("MOD",   mod_t,   ultimo["MOD"]),
            ("TOTAL", total_t, ultimo["Total"]),
        ]:
            ds = real_v - teo
            dp = (ds/teo*100) if teo > 0 else 0
            comp_data.append({
                "Driver":    driver,
                "Teorico":   f"S/ {teo:,.6f}",
                "Real":      f"S/ {real_v:,.6f}",
                "Desv_Soles":f"S/ {ds:+,.6f}",
                "Desv_Pct":  f"{dp:+.1f}%",
            })

    return kpis, fig_ap, fig_desv, fig_vol, tabla_data, comp_data


# ── Callback 3: Descarga ─────────────────────────────────────
@app.callback(
    Output("h-descarga",     "data"),
    Input("h-btn-descargar", "n_clicks"),
    State("h-selector-pt",   "value"),
    prevent_initial_call=True,
)
def descargar(n_clicks, codigo_pt):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
    if not codigo_pt: return None

    df_hist  = get_costo_real_mensual(codigo_pt)
    cm_t, cif_t, mod_t, total_t = get_costo_teorico(codigo_pt)

    # Descripción del PT
    row_pt  = df_cubo[df_cubo["Cod_PT"]==codigo_pt]
    desc_pt = row_pt["Desc_PT"].iloc[0] if not row_pt.empty else codigo_pt

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # ── Hoja 1: RESUMEN EJECUTIVO ────────────────────────
        if not df_hist.empty:
            ultimo  = df_hist[df_hist["Con_Datos"]].iloc[-1] if df_hist["Con_Datos"].any() else df_hist.iloc[-1]
            real_cu = ultimo["Total"]
            desv_s  = real_cu - total_t
            desv_p  = desv_s / total_t * 100 if total_t > 0 else 0
            mejor   = df_hist[df_hist["Con_Datos"]]["Total"].min() if df_hist["Con_Datos"].any() else 0
            peor    = df_hist[df_hist["Con_Datos"]]["Total"].max() if df_hist["Con_Datos"].any() else 0
            prom    = df_hist[df_hist["Con_Datos"]]["Total"].mean() if df_hist["Con_Datos"].any() else 0

            exec_rows = [
                ["RESUMEN EJECUTIVO — ANÁLISIS DE COSTOS", ""],
                ["Código PT", codigo_pt],
                ["Descripción", desc_pt],
                ["Fecha análisis", datetime.now().strftime("%d/%m/%Y %H:%M")],
                ["Períodos analizados", f"{df_hist['Periodo'].min()} a {df_hist['Periodo'].max()}"],
                ["", ""],
                ["COSTO TEÓRICO (S/ x und)", ""],
                ["CM Teórico", round(cm_t, 6)],
                ["CIF Teórico", round(cif_t, 6)],
                ["MOD Teórico", round(mod_t, 6)],
                ["TOTAL Teórico", round(total_t, 6)],
                ["", ""],
                ["COSTO REAL — ÚLTIMO MES CON DATOS", ""],
                ["Período", ultimo["Periodo"]],
                ["CM Real", round(ultimo["CM"], 6)],
                ["CIF Real", round(ultimo["CIF"], 6)],
                ["MOD Real", round(ultimo["MOD"], 6)],
                ["TOTAL Real", round(real_cu, 6)],
                ["", ""],
                ["DESVIACIÓN vs TEÓRICO", ""],
                ["Desviación (S/)", round(desv_s, 6)],
                ["Desviación (%)", f"{desv_p:+.1f}%"],
                ["Señal", "🔴 SOBRE COSTO" if desv_s > 0 else "🟢 BAJO COSTO"],
                ["", ""],
                ["ESTADÍSTICAS HISTÓRICAS", ""],
                ["Costo mínimo real", round(mejor, 6)],
                ["Costo máximo real", round(peor, 6)],
                ["Costo promedio real", round(prom, 6)],
                ["Volatilidad (%)", f"{df_hist[df_hist['Con_Datos']]['Total'].std()/prom*100:.1f}%" if prom > 0 else "—"],
                ["", ""],
                ["RECOMENDACIONES", ""],
            ]
            # Recomendaciones dinámicas
            recs = []
            if desv_p > 10:
                recs.append(["⚠️", f"Costo real supera el teórico en {desv_p:.1f}%. Revisar eficiencia de procesos."])
            if desv_p < -10:
                recs.append(["✅", f"Costo real {abs(desv_p):.1f}% por debajo del teórico. Validar calidad del producto."])
            vol_trend = df_hist[df_hist["Con_Datos"]]["Volumen"].tail(3).mean() if df_hist["Con_Datos"].any() else 0
            if vol_trend > 0:
                cif_pct = ultimo["CIF"] / real_cu * 100 if real_cu > 0 else 0
                if cif_pct > 40:
                    recs.append(["⚠️", f"CIF representa {cif_pct:.0f}% del costo real. Analizar eficiencia de maquinaria."])
            cm_pct = ultimo["CM"] / real_cu * 100 if real_cu > 0 else 0
            if cm_pct > 70:
                recs.append(["📦", f"CM representa {cm_pct:.0f}% del costo. Enfocarse en negociación de materiales."])
            if not recs:
                recs.append(["✅", "Costo dentro de rango esperado. Mantener monitoreo mensual."])
            for r in recs:
                exec_rows.append(r)

            df_exec = pd.DataFrame(exec_rows, columns=["Concepto", "Valor"])
            df_exec.to_excel(writer, sheet_name="1. Resumen Ejecutivo", index=False)

        # ── Hoja 2: HISTÓRICO MENSUAL ────────────────────────
        if not df_hist.empty:
            df_h = df_hist.copy()
            df_h["Teórico"]      = round(total_t, 6)
            df_h["Desv. S/"]     = (df_h["Total"] - total_t).round(6)
            df_h["Desv. %"]      = ((df_h["Desv. S/"] / total_t) * 100).round(1).map(lambda x: f"{x:+.1f}%")
            df_h["CM %"]         = (df_h["CM"]  / df_h["Total"] * 100).round(1).map(lambda x: f"{x:.0f}%")
            df_h["CIF %"]        = (df_h["CIF"] / df_h["Total"] * 100).round(1).map(lambda x: f"{x:.0f}%")
            df_h["MOD %"]        = (df_h["MOD"] / df_h["Total"] * 100).round(1).map(lambda x: f"{x:.0f}%")
            df_h["Fuente"]       = df_h["Con_Datos"].map({True: "SAP Real", False: "Teórico (fallback)"})
            df_h["Volumen"]      = df_h["Volumen"].apply(lambda x: int(x) if x > 0 else "—")
            cols = ["Periodo","Volumen","CM","CM %","CIF","CIF %","MOD","MOD %",
                    "Total","Teórico","Desv. S/","Desv. %","Fuente"]
            df_h[cols].to_excel(writer, sheet_name="2. Histórico Mensual", index=False)

        # ── Hoja 3: DETALLE CÁLCULO ──────────────────────────
        calc_rows = [
            ["METODOLOGÍA DE CÁLCULO", ""],
            ["", ""],
            ["FUENTE DE DATOS", ""],
            ["Costo teórico", "Calculado desde la BOM (explosión de materiales) + tiempos de proceso"],
            ["Costo real", "Extraído del cubo SAP — campo TIPO_COSTO_DIR = 'Costo Unitario'"],
            ["", ""],
            ["FÓRMULAS", ""],
            ["CM Unitario (teórico)", "Σ (Cantidad_componente × Precio_estándar) / Cantidad_base"],
            ["CIF Unitario (teórico)", "(T.Maq / Cant.Base) × Tarifa_Maquina"],
            ["MOD Unitario (teórico)", "(T.MO / Cant.Base) × Tarifa_MO"],
            ["Costo Real SAP", "Promedio ponderado por volumen cuando hay múltiples OPs en el mes"],
            ["Fallback (sin producción)", "Se usa el costo teórico cuando el código no tuvo OPs ese mes"],
            ["", ""],
            ["DRIVERS DE COSTO", ""],
            ["CM (Costo de Materiales)", "Insumos consumidos — materiales directos e indirectos"],
            ["CIF (Costo Indirecto de Fabricación)", "Horas máquina × tarifa del puesto de trabajo"],
            ["MOD (Mano de Obra Directa)", "Horas hombre × tarifa de mano de obra"],
            ["", ""],
            ["DESGLOSE TEÓRICO", ""],
        ]
        calc_rows.append(["Driver", "Costo Unitario (S/)", "% del Total"])
        for drv, val, total in [("CM", cm_t, total_t), ("CIF", cif_t, total_t), ("MOD", mod_t, total_t)]:
            pct = val/total*100 if total > 0 else 0
            calc_rows.append([drv, round(val, 6), f"{pct:.1f}%"])
        calc_rows.append(["TOTAL", round(total_t, 6), "100.0%"])

        pd.DataFrame(calc_rows, columns=["Campo", "Descripción / Valor", "Notas"]).to_excel(
            writer, sheet_name="3. Metodología", index=False)

        # ── Formato general ──────────────────────────────────
        from openpyxl.styles import Font, PatternFill, Alignment
        fills = {
            "1. Resumen Ejecutivo": PatternFill("solid", fgColor="1F3864"),
            "2. Histórico Mensual": PatternFill("solid", fgColor="1F5C2E"),
            "3. Metodología":       PatternFill("solid", fgColor="2E4A7A"),
        }
        hf = Font(bold=True, color="FFFFFF")
        for sn, ws in writer.sheets.items():
            # Header row
            for cell in ws[1]:
                cell.font = hf
                cell.fill = fills.get(sn, fills["1. Resumen Ejecutivo"])
                cell.alignment = Alignment(horizontal="center")
            # Color key rows in Resumen
            if sn == "1. Resumen Ejecutivo":
                for row in ws.iter_rows(min_row=2):
                    val = str(row[0].value or "")
                    if val.isupper() and val != "":
                        for cell in row:
                            cell.font  = Font(bold=True, color="00C8FF")
                            cell.fill  = PatternFill("solid", fgColor="0D1F2D")
            # Color desviation in Histórico
            if sn == "2. Histórico Mensual":
                desv_col = None
                for i, cell in enumerate(ws[1]):
                    if cell.value == "Desv. %":
                        desv_col = i + 1
                if desv_col:
                    for row in ws.iter_rows(min_row=2, min_col=desv_col, max_col=desv_col):
                        for cell in row:
                            try:
                                v = float(str(cell.value or "0").replace("%","").replace("+",""))
                                cell.font = Font(bold=True,
                                    color="FF5252" if v > 0 else ("4CAF50" if v < 0 else "FFFFFF"))
                            except: pass
            # Col widths
            for col in ws.columns:
                ml = max((len(str(c.value)) for c in col if c.value), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml+4, 55)

    output.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return dcc.send_bytes(output.read(), filename=f"Analisis_Costos_{codigo_pt}_{ts}.xlsx")



# ── Desglose por proceso — SAP recursivo ────────────────────
def _clean_cod_h(x):
    try: return str(int(float(str(x).strip())))
    except: return str(x).strip()

def _get_sap(cod, periodo):
    sub = df_cubo[(df_cubo["Cod_PT"]==str(cod))&(df_cubo["Periodo"]==periodo)]
    if sub.empty: return None
    vol = sub[sub["TIPO_COSTO_DIR"]=="Volumen"]["VOL_PRODUCIDOS"].sum()
    if vol == 0: return None
    mats = sub[sub["TIPO_COSTO_DIR"]=="Materiales"].copy()
    mats["_c"] = mats["Cod_Insumo"].apply(_clean_cod_h)
    cm  = mats[~mats["_c"].str.startswith("231")]["COSTO_REAL"].sum() / vol
    cif = sub[(sub["TIPO_COSTO_DIR"]=="Costo de Conversión")&
              (sub["CLASE_COSTO"]=="Gastos Indirectos")]["COSTO_REAL"].sum() / vol
    mod = sub[(sub["TIPO_COSTO_DIR"]=="Costo de Conversión")&
              (sub["CLASE_COSTO"]=="Mano de obra directa")]["COSTO_REAL"].sum() / vol
    semis = mats[mats["_c"].str.startswith("231")][["_c","CANT_REAL"]].copy()
    semis["f"] = semis["CANT_REAL"] / vol
    return {"cm":cm,"cif":cif,"mod":mod,"vol":vol,
            "semis":semis[["_c","f"]].to_dict("records")}

def _sap_fallback(cod, periodo, n=3):
    r = _get_sap(cod, periodo)
    if r: return r
    idx = PERIODOS_ORDENADOS.index(periodo) if periodo in PERIODOS_ORDENADOS else -1
    for p in reversed(PERIODOS_ORDENADOS[max(0,idx-n):idx]):
        r = _get_sap(cod, p)
        if r:
            r["vol"] = 0
            return r
    return None

def _desglose_recursivo(cod, periodo, factor, result, visited=None, depth=0):
    if visited is None: visited = set()
    if cod in visited or depth > 6: return
    visited.add(cod)
    r = _sap_fallback(cod, periodo)
    if not r: return
    t = df_tie[df_tie["Código Semi"]==str(cod)]
    proc = str(t.iloc[0]["Proceso"]).strip() if not t.empty else "Otros"
    if proc not in result:
        result[proc] = {"CM": 0, "CIF": 0, "MOD": 0}
    result[proc]["CM"]  += factor * r["cm"]
    result[proc]["CIF"] += factor * r["cif"]
    result[proc]["MOD"] += factor * r["mod"]
    for s in r.get("semis", []):
        _desglose_recursivo(s["_c"], periodo, factor * s["f"],
                            result, visited.copy(), depth+1)

def get_desglose_proceso(codigo_pt, periodo):
    result = {}
    _desglose_recursivo(str(codigo_pt), periodo, 1.0, result)
    total = sum(v["CM"]+v["CIF"]+v["MOD"] for v in result.values())
    return {k: {kk: round(vv, 6) for kk,vv in v.items()} for k,v in result.items()}, total


# ── Modal: click en barra → desglose por proceso ─────────────
@app.callback(
    Output("h-modal-overlay",   "style"),
    Output("h-modal-titulo",    "children"),
    Output("h-modal-subtitulo", "children"),
    Output("h-modal-contenido", "children"),
    Input("h-grafico-apilado",  "clickData"),
    Input("h-modal-cerrar",     "n_clicks"),
    State("h-selector-pt",      "value"),
    prevent_initial_call=True,
)
def toggle_modal(clickData, n_cerrar, codigo_pt):
    from dash import callback_context
    ctx     = callback_context
    trigger = ctx.triggered[0]["prop_id"].split(".")[0] if ctx.triggered else ""
    HIDDEN  = {"display": "none"}
    SHOWN   = {"display": "block"}

    if trigger == "h-modal-cerrar" or not clickData or not codigo_pt:
        return HIDDEN, "", "", []

    point    = clickData["points"][0]
    periodo  = str(point.get("x",""))[:7]
    # Use customdata[0] which stores 0=CM,1=CIF,2=MOD explicitly
    custom   = point.get("customdata", None)
    if isinstance(custom, (list, tuple)) and len(custom) > 0:
        idx_map = {0:"CM", 1:"CIF", 2:"MOD"}
        try:
            driver = idx_map.get(int(custom[0]), "CM")
        except:
            driver = ["CM","CIF","MOD"][point.get("curveNumber",0)] if point.get("curveNumber",0)<3 else "CM"
    else:
        curve  = point.get("curveNumber", 0)
        driver = ["CM","CIF","MOD"][curve] if curve < 3 else "CM"

    row_pt  = df_cubo[df_cubo["Cod_PT"]==str(codigo_pt)]
    desc_pt = str(row_pt["Desc_PT"].iloc[0]) if not row_pt.empty else codigo_pt

    try:
        desglose, real_total = get_desglose_proceso(codigo_pt, periodo)
    except Exception as e:
        return SHOWN, "Error", str(e), [html.P(str(e), style={"color":"red"})]

    if not desglose:
        return SHOWN, f"Sin datos — {periodo}", desc_pt[:45], [
            html.P("No hay datos SAP para este período", style={"color":"#7A9BBF"})]

    titulo  = f"📊 Costos por Proceso (aprox.) — {periodo}"
    subtit  = f"{codigo_pt} · {desc_pt[:45]}"
    col_map = {"CM":COLORES["cm"],"CIF":COLORES["cif"],"MOD":COLORES["mod"]}
    total_driver = sum(v[driver] for v in desglose.values())

    rows = []
    # Header 3 columnas
    cm_tot  = sum(v["CM"]  for v in desglose.values())
    cif_tot = sum(v["CIF"] for v in desglose.values())
    mod_tot = sum(v["MOD"] for v in desglose.values())

    # Header 7 columnas: Proceso | CM | %CM | CIF | %CIF | MOD | %MOD
    col_grid = "2fr repeat(6, auto)"
    hdr_style = {"fontSize":"10px","fontWeight":"bold","textAlign":"right","padding":"0 2px"}
    rows.append(html.Div(style={
        "display":"grid","gridTemplateColumns":col_grid,
        "gap":"4px","marginBottom":"6px",
        "borderBottom":"1px solid #2A3F54","paddingBottom":"6px"
    }, children=[
        html.Span("Proceso",style={"color":"#7A9BBF","fontSize":"10px","fontWeight":"bold"}),
        html.Span("CM",    style={**hdr_style,"color":COLORES["cm"]}),
        html.Span("%CM",   style={**hdr_style,"color":COLORES["cm"],"opacity":"0.7"}),
        html.Span("CIF",   style={**hdr_style,"color":COLORES["cif"]}),
        html.Span("%CIF",  style={**hdr_style,"color":COLORES["cif"],"opacity":"0.7"}),
        html.Span("MOD",   style={**hdr_style,"color":COLORES["mod"]}),
        html.Span("%MOD",  style={**hdr_style,"color":COLORES["mod"],"opacity":"0.7"}),
    ]))

    for proc in sorted(desglose.keys(), key=lambda p: -(desglose[p]["CM"]+desglose[p]["CIF"]+desglose[p]["MOD"])):
        v = desglose[proc]
        if v["CM"]+v["CIF"]+v["MOD"] <= 0: continue
        pct_cm  = v["CM"]  / cm_tot  * 100 if cm_tot  > 0 else 0
        pct_cif = v["CIF"] / cif_tot * 100 if cif_tot > 0 else 0
        pct_mod = v["MOD"] / mod_tot * 100 if mod_tot > 0 else 0
        val_style = {"fontSize":"11px","textAlign":"right","padding":"0 2px"}
        pct_style = {"fontSize":"10px","textAlign":"right","padding":"0 2px","color":"#7A9BBF"}
        rows.append(html.Div(style={
            "display":"grid","gridTemplateColumns":col_grid,
            "gap":"4px","padding":"5px 0",
            "borderBottom":"1px solid #1A2633"
        }, children=[
            html.Span(proc, style={"color":COLORES["text"],"fontSize":"12px"}),
            html.Span(f"{v['CM']:.4f}",    style={**val_style,"color":COLORES["cm"]}),
            html.Span(f"{pct_cm:.0f}%",    style=pct_style),
            html.Span(f"{v['CIF']:.4f}",   style={**val_style,"color":COLORES["cif"]}),
            html.Span(f"{pct_cif:.0f}%",   style=pct_style),
            html.Span(f"{v['MOD']:.4f}",   style={**val_style,"color":COLORES["mod"]}),
            html.Span(f"{pct_mod:.0f}%",   style=pct_style),
        ]))

    # Totales
    rows.append(html.Div(style={
        "display":"grid","gridTemplateColumns":col_grid,
        "gap":"4px","marginTop":"8px","paddingTop":"8px",
        "borderTop":"1px solid #2A3F54"
    }, children=[
        html.Span("TOTAL", style={"color":"#7A9BBF","fontWeight":"bold","fontSize":"12px"}),
        html.Span(f"{cm_tot:.4f}",  style={"color":COLORES["cm"], "fontWeight":"bold","fontSize":"12px","textAlign":"right"}),
        html.Span("100%", style={"color":"#7A9BBF","fontSize":"10px","textAlign":"right"}),
        html.Span(f"{cif_tot:.4f}", style={"color":COLORES["cif"],"fontWeight":"bold","fontSize":"12px","textAlign":"right"}),
        html.Span("100%", style={"color":"#7A9BBF","fontSize":"10px","textAlign":"right"}),
        html.Span(f"{mod_tot:.4f}", style={"color":COLORES["mod"],"fontWeight":"bold","fontSize":"12px","textAlign":"right"}),
        html.Span("100%", style={"color":"#7A9BBF","fontSize":"10px","textAlign":"right"}),
    ]))
    rows.append(html.P(f"Total: S/ {real_total:.4f}",
        style={"color":COLORES["accent"],"fontWeight":"bold",
               "textAlign":"right","marginTop":"6px","fontSize":"14px"}))

    return SHOWN, titulo, subtit, rows


if __name__ == "__main__":
    app.run(debug=False, host="0.0.0.0", port=int(os.environ.get("PORT", 8050)))
