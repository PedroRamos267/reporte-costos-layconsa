"""
=============================================================
  SIMULADOR DE RECETA — Layconsa
  Usa la misma lógica de cálculo que reporte_costos_web.py
  Filtra df_exp por Código PT para evitar duplicados
=============================================================
"""

import pandas as pd
import os, io
from datetime import datetime
from dash import Dash, html, dcc, Input, Output, State, dash_table
import plotly.graph_objects as go

ARCHIVO_DATOS  = "Analisis de costos_PY.xlsx"
PREFIJO_FABRIC = "231"
PROCESOS_EXCLUIR = []

import sys
if not os.path.exists(ARCHIVO_DATOS):
    print(f"❌ ERROR: No se encontró {ARCHIVO_DATOS}"); sys.exit(1)

df_exp = pd.read_excel(ARCHIVO_DATOS, sheet_name="Explosión")
df_tie = pd.read_excel(ARCHIVO_DATOS, sheet_name="Tiempos")
df_cos = pd.read_excel(ARCHIVO_DATOS, sheet_name="Costos")
df_mat = pd.read_excel(ARCHIVO_DATOS, sheet_name="Materiales")
df_exp.columns = df_exp.columns.str.strip()
df_tie.columns = df_tie.columns.str.strip()
df_cos.columns = df_cos.columns.str.strip()
df_mat.columns = df_mat.columns.str.strip()
df_mat["Codigo"] = df_mat["Codigo"].astype(str).str.strip()

for col in ["Código PT","Código Semi","Componente","Familia"]:
    if col in df_exp.columns:
        df_exp[col] = df_exp[col].astype(str).str.strip()
df_tie["Código Semi"]        = df_tie["Código Semi"].astype(str).str.strip()
df_cos["Número de artículo"] = df_cos["Número de artículo"].astype(str).str.strip()

for col in ["Cantidad Total Requerida","Cantidad Base","Costo estandar"]:
    if col in df_exp.columns:
        df_exp[col] = pd.to_numeric(df_exp[col], errors="coerce").fillna(0)
for col in ["Cantidad Base","T.MO","T.Maq","Tarifa MO","Tarifa Maquina","Cant.Opr"]:
    if col in df_tie.columns:
        df_tie[col] = pd.to_numeric(df_tie[col], errors="coerce").fillna(0)
df_cos["Precio promedio"] = pd.to_numeric(df_cos["Precio promedio"], errors="coerce").fillna(0)

# ── Helpers ─────────────────────────────────────────────────
def es_fabricado(familia):
    return str(familia).strip().startswith(PREFIJO_FABRIC)

def get_tiempos(codigo, df_t):
    row = df_t[df_t["Código Semi"] == str(codigo)]
    return row.iloc[0] if not row.empty else None

def get_unidad(codigo):
    codigo = str(codigo).strip()
    row = df_cos[df_cos["Número de artículo"] == codigo]
    if not row.empty: return str(row.iloc[0]["UnidadMedida"])
    row = df_exp[df_exp["Componente"] == codigo]
    if not row.empty and "Unidad" in df_exp.columns: return str(row.iloc[0]["Unidad"])
    return "—"

def get_descripcion(codigo):
    codigo = str(codigo).strip()
    row = df_cos[df_cos["Número de artículo"] == codigo]
    if not row.empty: return str(row.iloc[0]["Descripción del artículo"])
    row = df_exp[df_exp["Componente"] == codigo]
    if not row.empty: return str(row.iloc[0].get("Descripción Componente",""))
    return "—"

def fmt_num(v, decimales=2):
    """Formatea número con separador de miles y decimales fijos."""
    try:
        val = float(v)
        return f"{val:,.{decimales}f}"
    except:
        return str(v)

def fmt_auto(v):
    """Separador de miles para valores > 1000, sin forzar decimales."""
    try:
        val = float(v)
        if val >= 1000 or val <= -1000:
            return f"{val:,.2f}"
        return f"{val:.6f}" if val != int(val) else str(int(val))
    except:
        return str(v)

def safe_cant(v):
    """Convierte string con separador de miles a float."""
    try:
        return float(str(v).replace(",",""))
    except:
        return 0.0

def get_tipo(codigo):
    """Busca el Tipo desde hoja Materiales."""
    codigo = str(codigo).strip()
    row = df_mat[df_mat["Codigo"] == codigo]
    if not row.empty: return str(row.iloc[0].get("Tipo","—"))
    return "—"

def get_costo_estandar_exp(codigo, codigo_pt=None):
    """Costo estándar desde Explosión (Costo estandar column)."""
    codigo = str(codigo).strip()
    if codigo_pt:
        row = df_exp[(df_exp["Componente"] == codigo) & (df_exp["Código PT"] == str(codigo_pt))]
        if not row.empty:
            c = float(row.iloc[0]["Costo estandar"])
            if c > 0: return c
    row = df_exp[df_exp["Componente"] == codigo]
    if not row.empty:
        c = float(row.iloc[0]["Costo estandar"])
        if c > 0: return c
    row = df_cos[df_cos["Número de artículo"] == codigo]
    if not row.empty: return float(row.iloc[0]["Precio promedio"])
    return 0.0

def get_familia(codigo):
    """
    Busca la familia del componente.
    Primero como Componente, luego como Código Semi (para semis que se usan como nivel 1).
    Si empieza con 231 en df_exp como Código Semi, es fabricado.
    """
    codigo = str(codigo).strip()
    # Buscar como componente dentro de una explosión
    row = df_exp[df_exp["Componente"] == codigo]
    if not row.empty:
        return str(row.iloc[0].get("Familia", codigo[:3]))
    # Buscar como semi terminado (existe como Código Semi en Tiempos o Explosión)
    row_semi = df_exp[df_exp["Código Semi"] == codigo]
    if not row_semi.empty:
        # Si existe en Explosión como semi, su familia se deduce del prefijo
        return codigo[:3]
    return codigo[:3]

def get_costo_estandar_comp(codigo, codigo_pt):
    """Costo estándar desde Explosión filtrando por PT, o desde Costos."""
    codigo = str(codigo).strip()
    # Primero buscar en explosión del PT específico
    row = df_exp[(df_exp["Componente"] == codigo) & (df_exp["Código PT"] == str(codigo_pt))]
    if not row.empty:
        c = float(row.iloc[0]["Costo estandar"])
        if c > 0: return c
    # Luego en Costos (precio promedio)
    row = df_cos[df_cos["Número de artículo"] == codigo]
    if not row.empty: return float(row.iloc[0]["Precio promedio"])
    # Finalmente en cualquier fila de Explosión
    row = df_exp[df_exp["Componente"] == codigo]
    if not row.empty:
        c = float(row.iloc[0]["Costo estandar"])
        if c > 0: return c
    return 0.0

# ── Lógica de cálculo idéntica al dashboard ─────────────────
def calcular_semi(codigo_semi, cantidad_req, df_e, df_t, cache, resumen_global, codigo_pt, factor_escala=1.0):
    """
    Igual que reporte_costos_web.py — filtra por codigo_pt para evitar duplicados.
    factor_escala: escala las cantidades de sub-semis proporcionalmente a cant_base_nueva.
    """
    cache_key = f"{codigo_semi}_{cantidad_req}_{factor_escala}"
    if cache_key in cache:
        return cache[cache_key]["costo_x_und"], []

    hijos = df_e[
        (df_e["Código Semi"] == str(codigo_semi)) &
        (df_e["Código PT"]   == str(codigo_pt))
    ].copy()

    # Si no hay hijos en el PT base, buscar en cualquier PT que tenga ese semi
    # Esto permite usar componentes de otros PTs al simular una receta nueva
    if hijos.empty:
        hijos = df_e[df_e["Código Semi"] == str(codigo_semi)].copy()
        if not hijos.empty:
            # Usar el primer PT que lo contenga como referencia
            pt_ref = hijos["Código PT"].iloc[0]
            hijos  = df_e[
                (df_e["Código Semi"] == str(codigo_semi)) &
                (df_e["Código PT"]   == str(pt_ref))
            ].copy()
    if hijos.empty:
        return 0, []

    desc_semi   = hijos["Descripción Semi"].iloc[0] if "Descripción Semi" in hijos.columns else ""
    t           = get_tiempos(codigo_semi, df_t)
    proceso     = str(t["Proceso"]).strip().upper() if t is not None else "SIN PROCESO"
    cant_base_t = float(t["Cantidad Base"])          if t is not None else 1
    tarifa_maq  = float(t["Tarifa Maquina"])         if t is not None else 0
    tarifa_mo   = float(t["Tarifa MO"])              if t is not None else 0
    t_maq       = float(t["T.Maq"])                  if t is not None else 0
    t_mo        = float(t["T.MO"])                   if t is not None else 0
    if cant_base_t == 0: cant_base_t = 1

    # CIF y MOD escalan con factor_escala para mantener costo unitario constante
    cif = (t_maq / cant_base_t) * cantidad_req * tarifa_maq
    mod = (t_mo  / cant_base_t) * cantidad_req * tarifa_mo

    cm_total = 0; cm_comprados = 0; detalle = []

    for _, row in hijos.iterrows():
        componente = str(row["Componente"])
        # Escalar cantidad del sub-semi con el mismo factor que nivel 1
        cantidad   = float(row["Cantidad Total Requerida"]) * factor_escala
        costo_std  = float(row["Costo estandar"])
        familia    = str(row.get("Familia", componente[:3])).strip()

        if es_fabricado(familia):
            costo_calc, sub_det = calcular_semi(componente, cantidad, df_e, df_t, cache, resumen_global, codigo_pt, factor_escala)
            cm_comp = cantidad * costo_calc
        else:
            costo_calc   = costo_std
            cm_comp      = cantidad * costo_calc
            cm_comprados += cm_comp

        cm_total += cm_comp

    total_semi  = cm_total + cif + mod
    costo_x_und = total_semi / cantidad_req if cantidad_req != 0 else 0

    if proceso not in PROCESOS_EXCLUIR:
        if proceso not in resumen_global:
            resumen_global[proceso] = {"CM": 0, "CIF": 0, "MOD": 0}
        resumen_global[proceso]["CM"]  += cm_comprados
        resumen_global[proceso]["CIF"] += cif
        resumen_global[proceso]["MOD"] += mod

    cache[cache_key] = {"costo_x_und": costo_x_und}
    return costo_x_und, detalle


def calcular_receta_simulada(filas_receta, cant_base_nueva, df_t_sim, codigo_pt_base, df_exp_sim=None):
    """
    Calcula el costo de una receta simulada.
    - df_exp_sim: versión modificada de df_exp con precios editados (igual que dashboard)
    - Si no se pasa, usa df_exp original
    """
    if df_exp_sim is None:
        df_exp_sim = df_exp
    cache          = {}
    resumen_global = {}
    cm_total       = 0
    cm_comprados   = 0

    # Cantidad base original del PT para calcular factor por componente
    cant_base_orig = get_cant_base_original(codigo_pt_base)

    # Mapa de cantidades originales del PT base: {codigo: cantidad_original}
    filas_orig     = get_nivel1_pt(codigo_pt_base)
    orig_cant_map  = {str(f["Código"]): float(f["Cantidad"]) for f in filas_orig}
    # filas_orig has raw floats from get_nivel1_pt (no formatting applied there)

    # Factor global (cant_base_nueva / cant_base_orig) para el proceso PT
    factor_global  = (cant_base_nueva / cant_base_orig) if cant_base_orig > 0 else 1.0

    # Proceso del PT base (DOSIFICADO, ENCAJADO, etc.)
    t_pt        = get_tiempos(codigo_pt_base, df_t_sim)
    proceso_pt  = str(t_pt["Proceso"]).strip().upper() if t_pt is not None else "PROCESO_PT"
    cant_base_t = float(t_pt["Cantidad Base"])          if t_pt is not None else 1
    tarifa_maq  = float(t_pt["Tarifa Maquina"])         if t_pt is not None else 0
    tarifa_mo   = float(t_pt["Tarifa MO"])              if t_pt is not None else 0
    t_maq       = float(t_pt["T.Maq"])                  if t_pt is not None else 0
    t_mo        = float(t_pt["T.MO"])                   if t_pt is not None else 0
    if cant_base_t == 0: cant_base_t = 1

    # CIF y MOD del proceso PT escalan con cant_base_nueva
    cif_pt = (t_maq / cant_base_t) * cant_base_nueva * tarifa_maq
    mod_pt = (t_mo  / cant_base_t) * cant_base_nueva * tarifa_mo

    def _safe(v):
        try: return float(str(v).replace(",",""))
        except: return 0.0

    for fila in filas_receta:
        codigo   = str(fila.get("Código", "")).strip()
        cantidad = _safe(fila.get("Cantidad", 0))
        if not codigo or cantidad == 0:
            continue
        familia = get_familia(codigo)

        # Factor por componente: cantidad editada / cantidad original
        # Si el código fue cambiado manualmente, usar factor global
        cant_orig_comp = orig_cant_map.get(codigo, 0)
        if cant_orig_comp > 0:
            factor_comp = cantidad / cant_orig_comp
        else:
            # Código nuevo o cambiado — usar factor global
            factor_comp = factor_global

        if es_fabricado(familia):
            # Si el semi no existe bajo el PT base, buscar el PT donde sí existe
            codigo_pt_para_semi = codigo_pt_base
            hijos_check = df_exp_sim[
                (df_exp_sim["Código Semi"] == str(codigo)) &
                (df_exp_sim["Código PT"]   == str(codigo_pt_base))
            ]
            if hijos_check.empty:
                # Código cambiado manualmente — buscar en cualquier PT
                alt = df_exp_sim[df_exp_sim["Código Semi"] == str(codigo)]
                if not alt.empty:
                    codigo_pt_para_semi = str(alt.iloc[0]["Código PT"])
            costo_calc, _ = calcular_semi(
                codigo, cantidad, df_exp_sim, df_t_sim,
                cache, resumen_global, codigo_pt_para_semi, factor_comp
            )
            cm_comp = cantidad * costo_calc
        else:
            # Leer costo desde df_exp_sim (incluye precios editados)
            row_exp = df_exp_sim[df_exp_sim["Componente"] == str(codigo)]
            if not row_exp.empty:
                costo_std = float(row_exp.iloc[0]["Costo estandar"])
                if costo_std == 0:
                    costo_std = get_costo_estandar_comp(codigo, codigo_pt_base)
            else:
                costo_std = get_costo_estandar_comp(codigo, codigo_pt_base)
            cm_comp      = cantidad * costo_std
            cm_comprados += cm_comp

        cm_total += cm_comp

    # Agregar proceso del PT con CIF y MOD
    if proceso_pt not in resumen_global:
        resumen_global[proceso_pt] = {"CM": 0, "CIF": 0, "MOD": 0}
    resumen_global[proceso_pt]["CM"]  += cm_comprados
    resumen_global[proceso_pt]["CIF"] += cif_pt
    resumen_global[proceso_pt]["MOD"] += mod_pt

    # Limpiar vacíos
    resumen_global = {k: v for k, v in resumen_global.items()
                      if v["CM"] + v["CIF"] + v["MOD"] > 0}

    total_pt    = cm_total + cif_pt + mod_pt
    costo_x_und = total_pt / cant_base_nueva if cant_base_nueva > 0 else 0
    return resumen_global, costo_x_und


def get_nivel1_pt(codigo_pt):
    """Nivel 1 usando Código PT como filtro principal."""
    nivel1 = df_exp[df_exp["Código Semi"] == str(codigo_pt)].copy()
    rows = []
    for _, row in nivel1.iterrows():
        comp     = str(row["Componente"])
        cantidad = float(row["Cantidad Total Requerida"])
        familia  = str(row.get("Familia", comp[:3])).strip()
        fab      = es_fabricado(familia)
        rows.append({
            "Código":      comp,
            "Descripción": str(row.get("Descripción Componente", get_descripcion(comp))),
            "Unidad":      get_unidad(comp),
            "Cantidad":    cantidad,       # raw float for calculations
            "Es231":       fab,
        })
    return rows


def get_materiales_comprados(filas_receta, cant_base_nueva, cant_base_orig, codigo_pt_base):
    """
    Obtiene materiales comprados sumando directamente desde df_exp.
    La Explosión ya tiene Cantidad Total Requerida para el lote completo,
    por lo que basta agrupar por Componente y sumar.
    El factor de escala = cant_base_nueva / cant_base_orig se aplica al total.
    """
    def safe_cant(v):
        try: return float(str(v).replace(",",""))
        except: return 0.0

    factor_escala = (cant_base_nueva / cant_base_orig) if cant_base_orig > 0 else 1.0

    # Códigos nivel 1 de la receta actual (para excluirlos del cuadro 2)
    codigos_nivel1 = {str(f.get("Código","")).strip() for f in filas_receta}

    # Mapa de factores por componente nivel 1:
    # si el usuario cambió la cantidad, usamos cantidad_editada / cantidad_original
    orig_cant_map = {str(f["Código"]): float(f["Cantidad"]) for f in get_nivel1_pt(codigo_pt_base)}
    factor_por_comp = {}
    for fila in filas_receta:
        codigo = str(fila.get("Código","")).strip()
        cant_edit = safe_cant(fila.get("Cantidad", 0))
        cant_orig = orig_cant_map.get(codigo, 0)
        if cant_orig > 0:
            factor_por_comp[codigo] = cant_edit / cant_orig
        else:
            factor_por_comp[codigo] = factor_escala

    # Para cada semi 231 en la receta, obtener sus comprados desde df_exp
    materiales = {}

    for fila in filas_receta:
        codigo_semi = str(fila.get("Código","")).strip()
        if not codigo_semi: continue
        familia     = get_familia(codigo_semi)
        factor_semi = factor_por_comp.get(codigo_semi, factor_escala)

        if es_fabricado(familia):
            # Buscar el PT correcto para este semi
            pt_usar = codigo_pt_base
            check   = df_exp[(df_exp["Código Semi"]==codigo_semi) &
                             (df_exp["Código PT"]==str(codigo_pt_base))]
            if check.empty:
                alt = df_exp[df_exp["Código Semi"]==codigo_semi]
                if not alt.empty: pt_usar = str(alt.iloc[0]["Código PT"])

            # Obtener todos los comprados de este PT (todos los niveles)
            filas_pt   = df_exp[df_exp["Código PT"] == str(pt_usar)]
            cant_base_pt_orig = float(df_exp[df_exp["Código Semi"]==codigo_semi]["Cantidad Base"].iloc[0])                                 if not df_exp[df_exp["Código Semi"]==codigo_semi].empty else cant_base_orig
            if cant_base_pt_orig == 0: cant_base_pt_orig = 1

            # Comprados: no fabricados (no empieza con 231)
            comprados_pt = filas_pt[~filas_pt["Familia"].astype(str).str.startswith("231")]
            # Sumar por componente
            grouped = comprados_pt.groupby("Componente").agg({
                "Cantidad Total Requerida": "sum",
                "Descripción Componente":  "first",
                "Unidad":                  "first",
                "Familia":                 "first",
                "Costo estandar":          "first",
            }).reset_index()

            for _, row in grouped.iterrows():
                comp    = str(row["Componente"])
                if comp in codigos_nivel1: continue
                familia2= str(row["Familia"]).strip()
                es_232  = familia2.startswith("232")
                # Escalar: qty_lote_orig * (factor_semi)
                # factor_semi = cant_editada_nivel1 / cant_orig_nivel1
                cant_escalada = float(row["Cantidad Total Requerida"]) * factor_semi / cant_base_pt_orig * cant_base_pt_orig
                # Simplificado: cant_lote_orig * factor_semi (factor ya incluye escala)
                cant_final = float(row["Cantidad Total Requerida"]) * factor_semi

                tipo  = get_tipo(comp)
                costo = float(row["Costo estandar"]) if float(row["Costo estandar"]) > 0                         else get_costo_estandar_exp(comp, pt_usar)
                desc  = str(row["Descripción Componente"]) or get_descripcion(comp)
                unidad= str(row["Unidad"]) or get_unidad(comp)

                if comp not in materiales:
                    materiales[comp] = {"Código": comp, "Descripción": desc,
                                        "Unidad": unidad, "Tipo": tipo,
                                        "Cantidad": 0, "Costo": costo, "Es232": es_232}
                materiales[comp]["Cantidad"] += cant_final
        else:
            # Comprado directo nivel 1 — excluido del cuadro 2
            pass

    comprados = [v for v in materiales.values() if not v["Es232"]]
    aqp       = [v for v in materiales.values() if v["Es232"]]
    return comprados, aqp


def get_cant_base_original(codigo_pt):
    """
    Obtiene la cantidad base original del PT desde Explosión.
    Usamos Explosión porque las cantidades de la receta están
    expresadas en esa misma escala.
    """
    row = df_exp[df_exp["Código Semi"] == str(codigo_pt)]
    if not row.empty:
        cant = float(row.iloc[0]["Cantidad Base"] or 0)
        if cant > 0:
            return cant
    # Fallback: Tiempos
    t = get_tiempos(codigo_pt, df_tie)
    if t is not None:
        return float(t["Cantidad Base"] or 0)
    return 0.0


def get_procesos_tiempos(filas_receta, codigo_pt):
    """Carga procesos del PT base y de todos los semis fabricados en la receta."""
    visitados = set()
    maquinas  = {}

    def registrar(codigo):
        rows = df_tie[df_tie["Código Semi"] == str(codigo)]
        for _, t_row in rows.iterrows():
            proceso = str(t_row.get("Proceso","")).strip()
            maq     = str(t_row.get("Maquina", codigo))
            key     = f"{proceso}_{maq}"
            if key not in maquinas and proceso:
                maquinas[key] = {
                    "Proceso":       proceso,
                    "Maquina":       maq,
                    "Cantidad Base": float(t_row.get("Cantidad Base", 0) or 0),
                    "T.MO":          float(t_row.get("T.MO",          0) or 0),
                    "T.Maq":         float(t_row.get("T.Maq",         0) or 0),
                    "Tarifa Maq":    float(t_row.get("Tarifa Maquina",0) or 0),
                    "Tarifa MO":     float(t_row.get("Tarifa MO",     0) or 0),
                }

    def buscar(codigo, codigo_pt_buscar=None):
        if codigo in visitados: return
        visitados.add(codigo)
        registrar(codigo)
        # Buscar hijos bajo el PT correcto
        pt_usar = codigo_pt_buscar or codigo_pt
        hijos = df_exp[(df_exp["Código Semi"] == str(codigo)) &
                       (df_exp["Código PT"]   == str(pt_usar))]
        if hijos.empty:
            # Código cambiado — buscar en cualquier PT
            hijos = df_exp[df_exp["Código Semi"] == str(codigo)]
        for _, row in hijos.iterrows():
            comp    = str(row["Componente"])
            familia = str(row.get("Familia", comp[:3])).strip()
            if es_fabricado(familia):
                buscar(comp)

    # Proceso del PT base
    registrar(str(codigo_pt))
    # Procesos de semis fabricados en la receta (con sus códigos actuales)
    for fila in filas_receta:
        codigo  = str(fila.get("Código","")).strip()
        familia = get_familia(codigo)
        if es_fabricado(familia):
            # Verificar si existe bajo el PT base, si no buscar en cualquier PT
            hijos_check = df_exp[
                (df_exp["Código Semi"] == str(codigo)) &
                (df_exp["Código PT"]   == str(codigo_pt))
            ]
            pt_correcto = codigo_pt
            if hijos_check.empty:
                alt = df_exp[df_exp["Código Semi"] == str(codigo)]
                if not alt.empty:
                    pt_correcto = str(alt.iloc[0]["Código PT"])
            buscar(codigo, pt_correcto)

    return list(maquinas.values())


# ── App ──────────────────────────────────────────────────────
app    = Dash(__name__)
server = app.server

COLORES = {
    "bg": "#0F1923", "card": "#1A2633", "text": "#E8EDF2",
    "accent": "#00C8FF", "green": "#4CAF50", "orange": "#FF9800",
}

lista_pt = df_exp[["Código PT","Descripción PT"]].drop_duplicates()
def get_linea_sim(codigo_pt):
    c = str(codigo_pt)
    if c.startswith("214011"):
        return "Cuadernos"
    if len(c) >= 7 and not c[6].isdigit():
        return "Nuevos Desarrollos"
    return "Útiles"

lista_pt["Division"] = lista_pt["Código PT"].apply(get_linea_sim)

app.layout = html.Div(
    style={"backgroundColor": COLORES["bg"], "minHeight": "100vh",
           "fontFamily": "'Segoe UI', sans-serif", "color": COLORES["text"], "padding": "20px"},
    children=[
        html.H1("🧪 Simulador de Receta", style={"color": COLORES["accent"],
                "textAlign": "center", "marginBottom": "5px"}),
        html.P("Arma un nuevo producto combinando componentes existentes",
               style={"color": "#7A9BBF", "textAlign": "center", "marginBottom": "25px"}),

        # ── Datos nuevo producto ────────────────────────────
        html.Div(style={"backgroundColor": COLORES["card"], "borderRadius": "12px",
                        "padding": "15px", "marginBottom": "20px"}, children=[
            html.H3("📝 Nuevo Producto", style={"color": COLORES["accent"],
                    "fontSize": "16px", "marginTop": 0}),
            html.Div(style={"display": "grid", "gridTemplateColumns": "1fr 1fr 1fr", "gap": "15px"},
                     children=[
                html.Div([
                    html.Label("Código:", style={"color": "#7A9BBF", "fontSize": "12px"}),
                    dcc.Input(id="input-codigo-nuevo", type="text", placeholder="Ej: 2110099999",
                              style={"width": "100%", "padding": "8px", "marginTop": "5px",
                                     "backgroundColor": "#0D2137", "color": COLORES["text"],
                                     "border": "1px solid #00C8FF", "borderRadius": "6px"}),
                ]),
                html.Div([
                    html.Label("Descripción:", style={"color": "#7A9BBF", "fontSize": "12px"}),
                    dcc.Input(id="input-desc-nuevo", type="text", placeholder="Ej: Tempera 125ml Rojo",
                              style={"width": "100%", "padding": "8px", "marginTop": "5px",
                                     "backgroundColor": "#0D2137", "color": COLORES["text"],
                                     "border": "1px solid #00C8FF", "borderRadius": "6px"}),
                ]),
                html.Div([
                    html.Label("Cantidad Base (lote):", style={"color": "#7A9BBF", "fontSize": "12px"}),
                    dcc.Input(id="input-cant-base", type="number", placeholder="Ej: 8000", value=8000,
                              step=1,
                              style={"width": "100%", "padding": "8px", "marginTop": "5px",
                                     "backgroundColor": "#0D2137", "color": COLORES["text"],
                                     "border": "1px solid #FF9800", "borderRadius": "6px"}),
                ]),
            ]),
        ]),

        # ── PT Base ────────────────────────────────────────
        html.Div(style={"backgroundColor": COLORES["card"], "borderRadius": "12px",
                        "padding": "15px", "marginBottom": "20px"}, children=[
            html.H3("📦 PT Base", style={"color": COLORES["accent"], "fontSize": "16px", "marginTop": 0}),
            html.Div(style={"display": "flex", "gap": "12px", "marginBottom": "10px"}, children=[
                html.Button("Todos",      id="filtro-todos",      n_clicks=0,
                    style={"backgroundColor": COLORES["accent"], "color": "#000",
                           "border": "none", "borderRadius": "6px", "padding": "6px 16px",
                           "cursor": "pointer", "fontWeight": "bold", "fontSize": "13px"}),
                html.Button("Útiles",     id="filtro-utiles",     n_clicks=0,
                    style={"backgroundColor": "#4A5568", "color": "white",
                           "border": "none", "borderRadius": "6px", "padding": "6px 16px",
                           "cursor": "pointer", "fontSize": "13px"}),
                html.Button("Cuadernos",  id="filtro-cuadernos",  n_clicks=0,
                    style={"backgroundColor": "#4A5568", "color": "white",
                           "border": "none", "borderRadius": "6px", "padding": "6px 16px",
                           "cursor": "pointer", "fontSize": "13px"}),
                html.Button("Nuevos Des.", id="filtro-nuevos",    n_clicks=0,
                    style={"backgroundColor": "#4A5568", "color": "white",
                           "border": "none", "borderRadius": "6px", "padding": "6px 16px",
                           "cursor": "pointer", "fontSize": "13px"}),
            ]),
            dcc.Dropdown(
                id="selector-pt-base",
                options=[{"label": f"{r['Código PT']} — {r['Descripción PT']}", "value": r["Código PT"]}
                         for _, r in lista_pt.iterrows()],
                placeholder="Selecciona un PT base...", style={"color": "#000"}
            ),
            html.Button("📥 Cargar Receta Base", id="btn-cargar",
                style={"marginTop": "12px", "backgroundColor": COLORES["accent"],
                       "color": "#000", "fontWeight": "bold", "border": "none",
                       "borderRadius": "8px", "padding": "10px 24px", "cursor": "pointer"}),
            # Almacena cant_base original del PT para calcular el factor
            dcc.Store(id="store-cant-base-orig", data=0),
            dcc.Store(id="store-receta-orig",    data=[]),
            dcc.Store(id="store-codigo-pt",      data=None),
            dcc.Store(id="store-costos-231",     data={}),
        ]),

        # ── Tabla receta editable ───────────────────────────
        html.Div(style={"backgroundColor": COLORES["card"], "borderRadius": "12px",
                        "padding": "15px", "marginBottom": "20px",
                        "border": "1px solid #00C8FF"}, children=[
            html.H3("✏️ Receta Editable — Nivel 1",
                    style={"color": COLORES["accent"], "fontSize": "16px", "marginTop": 0}),
            html.P("Modifica códigos y cantidades. Las cantidades se ajustan automáticamente con la cantidad base.",
                   style={"color": "#7A9BBF", "fontSize": "12px", "marginBottom": "10px"}),
            dash_table.DataTable(
                id="tabla-receta",
                columns=[
                    {"name": "Código",      "id": "Código",      "editable": True},
                    {"name": "Descripción", "id": "Descripción", "editable": True},
                    {"name": "Unidad",      "id": "Unidad",      "editable": False},
                    {"name": "Cantidad",    "id": "Cantidad",    "editable": True, "type": "numeric"},
                    {"name": "Costo Unit.", "id": "Costo Unit.", "editable": True, "type": "numeric"},
                    {"name": "Costo Total", "id": "Costo Total", "editable": False},
                ],
                data=[], editable=True, row_deletable=True,
                style_header={"backgroundColor": "#1F3864", "color": "white", "fontWeight": "bold"},
                style_cell={"backgroundColor": "#1E2D3D", "color": COLORES["text"],
                            "border": "1px solid #2A3F54", "padding": "8px", "textAlign": "center"},
                style_cell_conditional=[
                    {"if": {"column_id": "Código"},      "width": "130px"},
                    {"if": {"column_id": "Descripción"}, "width": "230px", "textAlign": "left"},
                    {"if": {"column_id": "Unidad"},      "width": "70px"},
                    {"if": {"column_id": "Cantidad"},    "width": "100px"},
                    {"if": {"column_id": "Costo Unit."}, "width": "110px"},
                    {"if": {"column_id": "Costo Total"}, "width": "110px"},
                ],
                style_data_conditional=[
                    {"if": {"column_editable": True},
                     "backgroundColor": "#0D2137", "border": "1px solid #00C8FF"},
                    {"if": {"row_index": "odd"}, "backgroundColor": "#162030"},
                    {"if": {"filter_query": '{Es231} = "true"'},
                     "color": "#00C8FF"},
                ],
                page_action="none",
            ),
            html.Div(style={"display": "flex", "gap": "10px", "marginTop": "10px"}, children=[
                html.Button("➕ Agregar Fila", id="btn-agregar-fila",
                    style={"backgroundColor": "#4A5568", "color": "white",
                           "border": "none", "borderRadius": "8px", "padding": "8px 20px",
                           "cursor": "pointer", "fontSize": "13px"}),
                html.Button("🔍 Buscar Descripción y Costo", id="btn-buscar-desc",
                    style={"backgroundColor": "#1F5C8B", "color": "white",
                           "border": "none", "borderRadius": "8px", "padding": "8px 20px",
                           "cursor": "pointer", "fontSize": "13px"}),
            ]),
        ]),

        # ── Procesos ────────────────────────────────────────
        html.Div(style={"backgroundColor": COLORES["card"], "borderRadius": "12px",
                        "padding": "15px", "marginBottom": "20px",
                        "border": "1px solid #4CAF50"}, children=[
            html.H3("⚙️ Procesos — Cantidad Base Editable",
                    style={"color": COLORES["green"], "fontSize": "16px", "marginTop": 0}),
            html.P("Modifica la Cantidad Base para recalcular CIF y MOD de cada proceso.",
                   style={"color": "#7A9BBF", "fontSize": "12px", "marginBottom": "10px"}),
            dash_table.DataTable(
                id="tabla-procesos",
                columns=[
                    {"name": "Proceso",       "id": "Proceso",       "editable": False},
                    {"name": "Máquina",       "id": "Maquina",       "editable": False},
                    {"name": "Cantidad Base", "id": "Cantidad Base", "editable": True},
                    {"name": "T.MO",          "id": "T.MO",          "editable": False},
                    {"name": "T.Maq",         "id": "T.Maq",         "editable": False},
                    {"name": "Tarifa Maq",    "id": "Tarifa Maq",    "editable": False},
                    {"name": "Tarifa MO",     "id": "Tarifa MO",     "editable": False},
                ],
                data=[], editable=True,
                style_header={"backgroundColor": "#1F3864", "color": "white", "fontWeight": "bold"},
                style_cell={"backgroundColor": "#1E2D3D", "color": COLORES["text"],
                            "border": "1px solid #2A3F54", "padding": "8px", "textAlign": "center"},
                style_data_conditional=[
                    {"if": {"column_editable": True},
                     "backgroundColor": "#0D2137", "border": "1px solid #4CAF50"},
                    {"if": {"row_index": "odd"}, "backgroundColor": "#162030"},
                ],
                page_action="none",
            ),
        ]),

        # ── Materiales Comprados ────────────────────────────
        html.Div(style={"backgroundColor": COLORES["card"], "borderRadius": "12px",
                        "padding": "15px", "marginBottom": "20px",
                        "border": "1px solid #2196F3"}, children=[
            html.H3("🛒 Materiales Comprados",
                    style={"color": "#2196F3", "fontSize": "16px", "marginTop": 0}),
            html.P("Todos los materiales comprados de la explosión completa. Costo editable para simular.",
                   style={"color": "#7A9BBF", "fontSize": "12px", "marginBottom": "10px"}),
            dash_table.DataTable(
                id="tabla-comprados",
                columns=[
                    {"name": "Tipo",        "id": "Tipo",        "editable": False},
                    {"name": "Código",      "id": "Código",      "editable": False},
                    {"name": "Descripción", "id": "Descripción", "editable": False},
                    {"name": "Unidad",      "id": "Unidad",      "editable": False},
                    {"name": "Cantidad",    "id": "Cantidad",    "editable": False},
                    {"name": "Costo",       "id": "Costo",       "editable": True, "type": "numeric"},
                    {"name": "Total",       "id": "Total",       "editable": False},
                ],
                data=[],
                style_header={"backgroundColor": "#1F3864", "color": "white", "fontWeight": "bold"},
                style_cell={"backgroundColor": "#1E2D3D", "color": COLORES["text"],
                            "border": "1px solid #2A3F54", "padding": "8px", "textAlign": "center"},
                style_cell_conditional=[
                    {"if": {"column_id": "Descripción"}, "width": "250px", "textAlign": "left"},
                    {"if": {"column_id": "Tipo"},        "width": "100px"},
                    {"if": {"column_id": "Código"},      "width": "120px"},
                    {"if": {"column_id": "Cantidad"},    "width": "110px"},
                    {"if": {"column_id": "Costo"},       "width": "110px"},
                    {"if": {"column_id": "Total"},       "width": "110px"},
                ],
                style_data_conditional=[
                    {"if": {"column_editable": True},
                     "backgroundColor": "#0D2137", "border": "1px solid #2196F3"},
                    {"if": {"row_index": "odd"}, "backgroundColor": "#162030"},
                ],
                editable=True, page_size=15,
                filter_action="native", sort_action="native",
                sort_mode="multi",
            ),
        ]),

        # ── Botones ─────────────────────────────────────────
        html.Div(style={"display": "flex", "gap": "15px", "justifyContent": "center",
                        "marginBottom": "20px"}, children=[
            html.Button("🔢 Calcular Costo", id="btn-calcular",
                style={"backgroundColor": COLORES["green"], "color": "#000", "fontWeight": "bold",
                       "border": "none", "borderRadius": "8px", "padding": "12px 30px",
                       "cursor": "pointer", "fontSize": "15px",
                       "boxShadow": "0 0 15px rgba(76,175,80,0.4)"}),
            html.Button("⬇️ Exportar Excel", id="btn-exportar",
                style={"backgroundColor": COLORES["orange"], "color": "#000", "fontWeight": "bold",
                       "border": "none", "borderRadius": "8px", "padding": "12px 30px",
                       "cursor": "pointer", "fontSize": "15px",
                       "boxShadow": "0 0 15px rgba(255,152,0,0.4)"}),
            dcc.Download(id="descarga-receta"),
            html.Div(id="msg-receta", style={"color": COLORES["green"],
                                              "fontSize": "13px", "alignSelf": "center"}),
        ]),

        html.Div(id="resultado-receta"),
    ]
)


# ── Callback 0: Filtrar PT por división ─────────────────────
@app.callback(
    Output("selector-pt-base", "options"),
    Input("filtro-todos",      "n_clicks"),
    Input("filtro-utiles",     "n_clicks"),
    Input("filtro-cuadernos",  "n_clicks"),
    Input("filtro-nuevos",     "n_clicks"),
    prevent_initial_call=False,
)
def filtrar_pt(n_todos, n_utiles, n_cuadernos, n_nuevos):
    from dash import callback_context
    ctx     = callback_context
    trigger = ctx.triggered[0]["prop_id"].split(".")[0] if ctx.triggered else "filtro-todos"
    if trigger == "filtro-cuadernos":
        df_filt = lista_pt[lista_pt["Division"] == "Cuadernos"]
    elif trigger == "filtro-utiles":
        df_filt = lista_pt[lista_pt["Division"] == "Útiles"]
    elif trigger == "filtro-nuevos":
        df_filt = lista_pt[lista_pt["Division"] == "Nuevos Desarrollos"]
    else:
        df_filt = lista_pt
    return [{"label": f"{r['Código PT']} — {r['Descripción PT']}", "value": r["Código PT"]}
            for _, r in df_filt.iterrows()]


# ── Callback 1: Cargar receta base ──────────────────────────
@app.callback(
    Output("tabla-receta",        "data"),
    Output("tabla-procesos",      "data"),
    Output("tabla-comprados",     "data"),
    Output("store-cant-base-orig","data"),
    Output("store-receta-orig",   "data"),
    Output("store-codigo-pt",     "data"),
    Input("btn-cargar",           "n_clicks"),
    State("selector-pt-base",     "value"),
    prevent_initial_call=True,
)
def cargar_receta(n_clicks, codigo_pt):
    if not codigo_pt:
        return [], [], [], [], 0, [], None
    filas_raw     = get_nivel1_pt(codigo_pt)
    cant_base_ori = get_cant_base_original(codigo_pt)
    # Add costo placeholder to nivel1 (will be filled after calcular)
    filas = []
    for f in filas_raw:
        row = dict(f)
        if not row.get("Es231", False):
            row["Costo Unit."] = get_costo_estandar_exp(row["Código"], codigo_pt)
        else:
            row["Costo Unit."] = 0.0  # calculated later
        costo_u = float(row["Costo Unit."])
        cant    = float(row["Cantidad"])
        row["Cantidad"]    = fmt_num(cant, 2)
        row["Costo Total"] = fmt_num(cant * costo_u, 2)
        row["Es231"] = "true" if row.get("Es231", False) else "false"
        filas.append(row)
    procesos = get_procesos_tiempos(filas_raw, codigo_pt)
    # Load materiales comprados and AQP
    comp, aqp = get_materiales_comprados(filas_raw, cant_base_ori, cant_base_ori, codigo_pt)
    filas_comp = [{"Tipo": m["Tipo"], "Código": m["Código"], "Descripción": m["Descripción"],
                   "Unidad": m["Unidad"],
                   "Cantidad": fmt_num(m["Cantidad"], 2),
                   "Costo":    round(m["Costo"], 6),
                   "Total":    fmt_num(m["Cantidad"] * m["Costo"], 2)} for m in comp]
    # Format Cantidad Base in procesos
    for p in procesos:
        cb = float(p.get("Cantidad Base", 0) or 0)
        p["Cantidad Base"] = f"{int(cb):,}" if cb == int(cb) else f"{cb:,.0f}"
    return filas, procesos, filas_comp, cant_base_ori, filas_raw, codigo_pt


# ── Callback 2: Actualizar cantidades al cambiar cantidad base
@app.callback(
    Output("tabla-receta",    "data", allow_duplicate=True),
    Output("tabla-comprados", "data", allow_duplicate=True),
    Input("input-cant-base",      "value"),
    State("store-cant-base-orig", "data"),
    State("store-receta-orig",    "data"),
    State("tabla-receta",         "data"),
    State("store-codigo-pt",      "data"),
    prevent_initial_call=True,
)
def actualizar_cantidades(cant_base_nueva, cant_base_orig, receta_orig, data_actual, codigo_pt):
    if not cant_base_nueva or not cant_base_orig or cant_base_orig == 0:
        return data_actual or [], [], []
    cant_base_nueva = float(cant_base_nueva)
    factor    = cant_base_nueva / float(cant_base_orig)
    orig_map  = {str(f["Código"]): float(f["Cantidad"]) for f in (receta_orig or [])}
    resultado = []
    for row in (data_actual or []):
        codigo    = str(row.get("Código","")).strip()
        # orig_map has raw floats; row["Cantidad"] may be formatted string
        cant_orig_raw = orig_map.get(codigo, 0)
        if cant_orig_raw == 0:
            try: cant_orig_raw = float(str(row.get("Cantidad",0)).replace(",",""))
            except: cant_orig_raw = 0
        row       = dict(row)
        nueva_cant = round(cant_orig_raw * factor, 2)
        costo_u    = float(str(row.get("Costo Unit.", 0) or 0).replace(",",""))
        row["Cantidad"]    = fmt_num(nueva_cant, 2)
        row["Costo Total"] = fmt_num(nueva_cant * costo_u, 2)
        resultado.append(row)

    # Actualizar cantidades en tablas comprados y AQP
    if not codigo_pt:
        return resultado, [], []
    comp, aqp = get_materiales_comprados(resultado, cant_base_nueva, float(cant_base_orig), codigo_pt)
    filas_comp = [{"Tipo": m["Tipo"], "Código": m["Código"], "Descripción": m["Descripción"],
                   "Unidad": m["Unidad"],
                   "Cantidad": fmt_num(m["Cantidad"], 2),
                   "Costo":    round(m["Costo"], 6),
                   "Total":    fmt_num(m["Cantidad"]*m["Costo"], 2)} for m in comp]
    return resultado, filas_comp


# descripcion y costo_total se actualizan en calcular()


# ── Callback 3: Agregar fila ────────────────────────────────
@app.callback(
    Output("tabla-receta", "data", allow_duplicate=True),
    Input("btn-agregar-fila", "n_clicks"),
    State("tabla-receta",     "data"),
    prevent_initial_call=True,
)
def agregar_fila(n_clicks, data):
    if data is None: data = []
    data.append({"Código": "", "Descripción": "", "Unidad": "—",
                 "Cantidad": 0, "Costo Unit.": 0, "Costo Total": "0.00", "Es231": "false"})
    return data


# ── Callback: Buscar descripción y costo ────────────────────
@app.callback(
    Output("tabla-receta", "data", allow_duplicate=True),
    Input("btn-buscar-desc",   "n_clicks"),
    State("tabla-receta",      "data"),
    State("store-codigo-pt",   "data"),
    State("store-costos-231",  "data"),
    prevent_initial_call=True,
)
def buscar_descripcion(n_clicks, data, codigo_pt_base, costos_231):
    """Actualiza descripción, unidad y costo para todos los códigos de la receta."""
    if not data:
        return data
    costos_231     = costos_231 or {}
    codigos_orig   = {str(f["Código"]) for f in get_nivel1_pt(codigo_pt_base)} if codigo_pt_base else set()
    resultado      = []
    for row in data:
        row    = dict(row)
        codigo = str(row.get("Código","")).strip()
        if not codigo:
            resultado.append(row)
            continue
        # Siempre actualizar descripción y unidad
        row["Descripción"] = get_descripcion(codigo)
        row["Unidad"]      = get_unidad(codigo)
        familia            = get_familia(codigo)
        es231              = es_fabricado(familia)
        row["Es231"]       = "true" if es231 else "false"
        try:
            cant = float(str(row.get("Cantidad", 0) or 0).replace(",",""))
        except:
            cant = 0.0
        if es231:
            # Usar costo del store si ya se calculó, si no mostrar 0
            cu = float(costos_231.get(codigo, 0))
            row["Costo Unit."] = round(cu, 6)
            row["Costo Total"] = fmt_num(cant * cu, 2)
        else:
            cu = get_costo_estandar_exp(codigo, codigo_pt_base)
            row["Costo Unit."] = round(cu, 6)
            row["Costo Total"] = fmt_num(cant * cu, 2)
        resultado.append(row)
    return resultado


# ── Callback: Actualizar Costo Total en vivo ────────────────
# Triggered ONLY by store-costos-231 changes (after Calcular)
# and by input-cant-base — NOT by tabla-receta to avoid loops
@app.callback(
    Output("tabla-receta", "data", allow_duplicate=True),
    Input("store-costos-231", "data"),
    State("tabla-receta",     "data"),
    prevent_initial_call=True,
)
def actualizar_costo_total(costos_231, data):
    """Actualiza Costo Unit. de 231 y Costo Total tras presionar Calcular."""
    if not data:
        return data
    costos_231 = costos_231 or {}
    resultado  = []
    for row in data:
        row    = dict(row)
        codigo = str(row.get("Código","")).strip()
        es231  = str(row.get("Es231","false")) == "true"
        try:
            cant = float(str(row.get("Cantidad", 0) or 0).replace(",",""))
        except:
            cant = 0.0
        if es231 and codigo in costos_231:
            cu = float(costos_231[codigo])
            row["Costo Unit."] = round(cu, 6)
            row["Costo Total"] = fmt_num(cant * cu, 2)
        else:
            try:
                cu = float(str(row.get("Costo Unit.", 0) or 0).replace(",",""))
            except:
                cu = 0.0
            row["Costo Total"] = fmt_num(cant * cu, 2)
        resultado.append(row)
    return resultado


# ── Callback 4: Calcular ────────────────────────────────────
@app.callback(
    Output("resultado-receta", "children"),
    Output("tabla-procesos",   "data",     allow_duplicate=True),
    Output("tabla-receta",     "data",     allow_duplicate=True),
    Output("tabla-comprados",  "data",     allow_duplicate=True),
    Output("msg-receta",       "children"),
    Output("store-costos-231", "data",     allow_duplicate=True),
    Input("btn-calcular",      "n_clicks"),
    State("tabla-receta",      "data"),
    State("tabla-procesos",    "data"),
    State("tabla-comprados",   "data"),
    State("input-codigo-nuevo","value"),
    State("input-desc-nuevo",  "value"),
    State("input-cant-base",   "value"),
    State("store-codigo-pt",   "data"),
    State("store-cant-base-orig","data"),
    prevent_initial_call=True,
)
def calcular(n_clicks, data, datos_procesos, datos_comprados,
             codigo_nuevo, desc_nuevo, cant_base, codigo_pt_base, cant_base_orig):
    if not data:
        return "", [], [], [], "⚠️ Agrega componentes primero", {}
    if not codigo_pt_base:
        return "", [], [], [], "⚠️ Selecciona un PT base primero", {}

    cant_base_pt = float(cant_base or 8000)

    def safe_float(v):
        try: return float(str(v).replace(",",""))
        except: return 0.0


    # Aplicar ediciones de Cantidad Base en Tiempos
    df_tie_sim = df_tie.copy()
    if datos_procesos:
        for row in datos_procesos:
            maq        = str(row.get("Maquina",""))
            try:
                nueva_base = float(str(row.get("Cantidad Base", 0) or 0).replace(",",""))
            except:
                nueva_base = 0
            if maq and nueva_base > 0:
                mask = df_tie_sim["Maquina"].astype(str).str.strip() == maq
                df_tie_sim.loc[mask, "Cantidad Base"] = nueva_base

    # Aplicar precios editados — igual que reporte_costos_web
    df_exp_sim = df_exp.copy()

    # 1. Precios editados en tabla comprados (niveles 2+)
    if datos_comprados:
        for r in datos_comprados:
            comp   = str(r.get("Código", "")).strip()
            precio = safe_float(r.get("Costo", 0))
            if comp and precio > 0:
                mask = df_exp_sim["Componente"] == comp
                if mask.any():
                    df_exp_sim.loc[mask, "Costo estandar"] = precio

    # 2. Precios manuales en tabla-receta nivel 1 (comprados)
    for row in data:
        codigo = str(row.get("Código","")).strip()
        es231  = str(row.get("Es231","false")) == "true"
        if not es231 and codigo:
            cu = safe_float(row.get("Costo Unit.", 0))
            if cu > 0:
                mask = df_exp_sim["Componente"] == codigo
                if mask.any():
                    df_exp_sim.loc[mask, "Costo estandar"] = cu

    # Re-explotar con los precios editados
    resumen_sim, costo_x_und = calcular_receta_simulada(
        data, cant_base_pt, df_tie_sim, codigo_pt_base, df_exp_sim=df_exp_sim
    )

    nombre  = f"{codigo_nuevo or 'NUEVO'} — {desc_nuevo or 'Sin descripción'}"
    paleta  = ["#2196F3","#4CAF50","#FF9800","#E91E63","#9C27B0","#00BCD4","#FF5722"]
    total   = sum(v["CM"]+v["CIF"]+v["MOD"] for v in resumen_sim.values()) / cant_base_pt
    tot_cm  = sum(v["CM"]  for v in resumen_sim.values()) / cant_base_pt
    tot_cif = sum(v["CIF"] for v in resumen_sim.values()) / cant_base_pt
    tot_mod = sum(v["MOD"] for v in resumen_sim.values()) / cant_base_pt

    def kpi(titulo, valor, color):
        return html.Div(
            style={"backgroundColor": COLORES["card"], "borderLeft": f"4px solid {color}",
                   "borderRadius": "10px", "padding": "15px 20px", "flex": "1", "minWidth": "160px"},
            children=[
                html.P(titulo, style={"margin": 0, "fontSize": "12px", "color": "#7A9BBF"}),
                html.H2(valor, style={"margin": "5px 0 0 0", "color": color, "fontSize": "18px"}),
            ]
        )

    kpis = html.Div(style={"display": "flex", "gap": "15px", "flexWrap": "wrap", "marginBottom": "20px"},
                    children=[
        kpi("💰 Costo Unitario",  f"S/ {total:,.6f}",              COLORES["accent"]),
        kpi("📦 Costo del Lote",  f"S/ {total*cant_base_pt:,.2f}", COLORES["orange"]),
        kpi("🧱 CM Unitario",     f"S/ {tot_cm:,.6f}",             "#2196F3"),
        kpi("⚙️ CIF Unitario",    f"S/ {tot_cif:,.6f}",            COLORES["orange"]),
        kpi("👷 MOD Unitario",    f"S/ {tot_mod:,.6f}",            COLORES["green"]),
        kpi("🔢 Cantidad Base",   f"{int(cant_base_pt):,}",        "#7A9BBF"),
    ])

    # Cascada
    filas_cas = []; tg = 0
    for proceso, v in resumen_sim.items():
        for tipo, monto in [("CM",v["CM"]),("CIF",v["CIF"]),("MOD",v["MOD"])]:
            if monto > 0:
                cu = monto / cant_base_pt; tg += cu
                filas_cas.append({"Proceso": proceso, "Tipo": tipo, "Costo": f"S/ {cu:,.6f}"})
    for f in filas_cas:
        v = float(f["Costo"].replace("S/ ","").replace(",",""))
        f["%TG"] = f"{v/tg*100:.1f}%" if tg > 0 else "0%"
    filas_cas.append({"Proceso": "TOTAL PT", "Tipo": "", "Costo": f"S/ {tg:,.6f}", "%TG": "100.0%"})

    tabla_cas = dash_table.DataTable(
        columns=[{"name": c, "id": c} for c in ["Proceso","Tipo","Costo","%TG"]],
        data=filas_cas,
        style_header={"backgroundColor": "#1F3864", "color": "white", "fontWeight": "bold"},
        style_cell={"backgroundColor": "#1E2D3D", "color": COLORES["text"],
                    "border": "1px solid #2A3F54", "padding": "8px", "textAlign": "center"},
        style_data_conditional=[
            {"if": {"row_index": "odd"}, "backgroundColor": "#162030"},
            {"if": {"filter_query": '{Tipo} = "MOD"'}, "color": "#4CAF50", "fontWeight": "bold"},
            {"if": {"filter_query": '{Tipo} = "CIF"'}, "color": "#FF9800", "fontWeight": "bold"},
            {"if": {"filter_query": '{Proceso} = "TOTAL PT"'},
             "backgroundColor": "#1F3864", "color": "#00C8FF", "fontWeight": "bold"},
        ],
        page_action="none",
    )

    labels_don = [f"{f['Proceso']} {f['Tipo']}" for f in filas_cas if f["Proceso"] != "TOTAL PT"]
    values_don = [float(f["Costo"].replace("S/ ","")) for f in filas_cas if f["Proceso"] != "TOTAL PT"]
    fig_don = go.Figure(go.Pie(
        labels=labels_don, values=values_don, hole=0.55,
        marker_colors=(paleta*5)[:len(labels_don)], textinfo="percent",
        hovertemplate="<b>%{label}</b><br>S/ %{value:.6f}<br>%{percent}<extra></extra>"
    ))
    fig_don.update_layout(template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)",
                          margin=dict(l=10,r=10,t=10,b=10))

    resultado = html.Div([
        html.Div(style={"backgroundColor": COLORES["card"], "borderRadius": "12px",
                        "padding": "15px", "marginBottom": "20px"}, children=[
            html.H3(f"📊 Resultado: {nombre}",
                    style={"color": COLORES["accent"], "fontSize": "16px", "marginTop": 0}),
            kpis,
        ]),
        html.Div(style={"display": "grid", "gridTemplateColumns": "1fr 1fr", "gap": "20px"}, children=[
            html.Div(style={"backgroundColor": COLORES["card"], "borderRadius": "12px", "padding": "15px"},
                     children=[
                html.H3("📋 Cascada por Proceso",
                        style={"color": COLORES["accent"], "fontSize": "16px", "marginTop": 0}),
                tabla_cas,
            ]),
            html.Div(style={"backgroundColor": COLORES["card"], "borderRadius": "12px", "padding": "15px"},
                     children=[
                html.H3("🥧 Distribución de Costos",
                        style={"color": COLORES["accent"], "fontSize": "16px", "marginTop": 0}),
                dcc.Graph(figure=fig_don),
            ]),
        ]),
    ])

    # Actualizar tabla de procesos con códigos actuales
    procesos_actualizados = get_procesos_tiempos(data, codigo_pt_base)

    # Variables necesarias para actualizar receta
    cant_base_orig_local = get_cant_base_original(codigo_pt_base)
    factor_global  = (cant_base_pt / cant_base_orig_local) if cant_base_orig_local > 0 else 1.0
    orig_cant_map  = {str(f["Código"]): float(f["Cantidad"]) for f in get_nivel1_pt(codigo_pt_base)}

    # Actualizar Costo Unit. en receta para componentes 231 (costo calculado)
    cache_costos = {}
    def get_costo_semi(codigo):
        if codigo in cache_costos: return cache_costos[codigo]
        fam = get_familia(codigo)
        if es_fabricado(fam):
            pt_usar = codigo_pt_base
            check   = df_exp[(df_exp["Código Semi"]==codigo)&(df_exp["Código PT"]==str(codigo_pt_base))]
            if check.empty:
                alt = df_exp[df_exp["Código Semi"]==codigo]
                if not alt.empty: pt_usar = str(alt.iloc[0]["Código PT"])
            c_temp = {}; r_temp = {}
            cu, _ = calcular_semi(codigo, float(data[0].get("Cantidad",1) or 1),
                                  df_exp, df_tie_sim, c_temp, r_temp, pt_usar,
                                  factor_global)
            cache_costos[codigo] = cu
            return cu
        return 0.0

    receta_actualizada = []
    for row in data:
        row  = dict(row)
        codigo = str(row.get("Código","")).strip()
        fam    = get_familia(codigo)
        if es_fabricado(fam):
            # Recalcular costo del semi
            pt_usar = codigo_pt_base
            check   = df_exp[(df_exp["Código Semi"]==codigo)&(df_exp["Código PT"]==str(codigo_pt_base))]
            if check.empty:
                alt = df_exp[df_exp["Código Semi"]==codigo]
                if not alt.empty: pt_usar = str(alt.iloc[0]["Código PT"])
            cant   = float(row.get("Cantidad", 0) or 0)
            orig_c = orig_cant_map.get(codigo, cant)
            f_comp = (cant / orig_c) if orig_c > 0 else factor_global
            c_temp = {}; r_temp = {}
            cu, _  = calcular_semi(codigo, cant, df_exp, df_tie_sim, c_temp, r_temp, pt_usar, f_comp)
            row["Costo Unit."] = round(cu, 6)
            row["Costo Total"] = fmt_num(cant * cu, 2)
            row["Cantidad"]    = fmt_num(cant, 2)
        else:
            costo_u = float(str(row.get("Costo Unit.", 0) or 0).replace(",",""))
            cant_v  = safe_cant(row.get("Cantidad", 0))
            row["Costo Total"] = fmt_num(cant_v * costo_u, 2)
            row["Cantidad"]    = fmt_num(cant_v, 2)
        receta_actualizada.append(row)

    # Actualizar cantidades en comprados/aqp con costos editados por usuario
    comp_edit_map = {str(r["Código"]): safe_float(r.get("Costo",0)) for r in (datos_comprados or [])}

    comp_new, _ = get_materiales_comprados(data, cant_base_pt, cant_base_orig, codigo_pt_base)
    filas_comp = []
    for m in comp_new:
        costo = comp_edit_map.get(m["Código"], m["Costo"])
        filas_comp.append({"Tipo": m["Tipo"], "Código": m["Código"],
                           "Descripción": m["Descripción"], "Unidad": m["Unidad"],
                           "Cantidad": fmt_num(m["Cantidad"], 2),
                           "Costo":    round(costo, 6),
                           "Total":    fmt_num(m["Cantidad"]*costo, 2)})

    # Guardar costos calculados de 231 en store para actualización en vivo
    costos_231_map = {
        str(row.get("Código","")): float(str(row.get("Costo Unit.", 0) or 0).replace(",",""))
        for row in receta_actualizada
        if str(row.get("Es231","false")) == "true"
    }

    return (resultado, procesos_actualizados, receta_actualizada,
            filas_comp,
            f"✅ Calculado — {datetime.now().strftime('%H:%M:%S')}",
            costos_231_map)


# ── Callback 5: Exportar Excel ──────────────────────────────
@app.callback(
    Output("descarga-receta",  "data"),
    Input("btn-exportar",      "n_clicks"),
    State("tabla-receta",      "data"),
    State("tabla-procesos",    "data"),
    State("tabla-comprados",   "data"),
    State("input-codigo-nuevo","value"),
    State("input-desc-nuevo",  "value"),
    State("input-cant-base",   "value"),
    State("store-codigo-pt",   "data"),
    State("store-cant-base-orig","data"),
    prevent_initial_call=True,
)
def exportar(n_clicks, data, datos_procesos, datos_comprados,
             codigo_nuevo, desc_nuevo, cant_base, codigo_pt_base, cant_base_orig):
    from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    if not data or not codigo_pt_base: return None

    cant_base_pt   = float(cant_base or 8000)
    cant_base_orig = float(cant_base_orig or cant_base_pt)

    def safe_float(v):
        try: return float(str(v).replace(",",""))
        except: return 0.0

    df_tie_sim     = df_tie.copy()
    if datos_procesos:
        for row in datos_procesos:
            maq        = str(row.get("Maquina",""))
            try:
                nueva_base = float(str(row.get("Cantidad Base", 0) or 0).replace(",",""))
            except:
                nueva_base = 0
            if maq and nueva_base > 0:
                mask = df_tie_sim["Maquina"].astype(str).str.strip() == maq
                df_tie_sim.loc[mask, "Cantidad Base"] = nueva_base

    resumen_sim, _ = calcular_receta_simulada(data, cant_base_pt, df_tie_sim, codigo_pt_base)

    # Cálculo ACTUAL (PT base con cant_base_orig)
    from copy import deepcopy
    filas_orig    = get_nivel1_pt(codigo_pt_base)
    resumen_act, _ = calcular_receta_simulada(filas_orig, cant_base_orig, df_tie, codigo_pt_base)

    nombre = f"{codigo_nuevo or 'NUEVO'} — {desc_nuevo or 'Sin descripción'}"
    output = io.BytesIO()

    def build_resumen_dict(resumen, cant_base):
        d = {}; total = 0
        for proceso, v in resumen.items():
            for tipo, monto in [("CM",v["CM"]),("CIF",v["CIF"]),("MOD",v["MOD"])]:
                if monto > 0:
                    cu = monto / cant_base; total += cu
                    d[f"{tipo}_{proceso}"] = {"proceso": proceso, "tipo": tipo, "cu": cu, "lote": monto}
        d["TOTAL"] = {"proceso": "TOTAL", "tipo": "", "cu": total, "lote": total * cant_base}
        return d

    act_dict = build_resumen_dict(resumen_act, cant_base_orig)
    sim_dict = build_resumen_dict(resumen_sim, cant_base_pt)
    all_keys = list(dict.fromkeys(list(act_dict.keys()) + list(sim_dict.keys())))

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # ── Hoja 1: Comparativo Resumen ────────────────────
        # KPIs: Costo Unit., CM, CIF, MOD
        total_act = act_dict.get("TOTAL",{}).get("cu",0)
        total_sim = sim_dict.get("TOTAL",{}).get("cu",0)
        cm_act  = sum(v["cu"] for k,v in act_dict.items() if v["tipo"]=="CM")
        cm_sim  = sum(v["cu"] for k,v in sim_dict.items() if v["tipo"]=="CM")
        cif_act = sum(v["cu"] for k,v in act_dict.items() if v["tipo"]=="CIF")
        cif_sim = sum(v["cu"] for k,v in sim_dict.items() if v["tipo"]=="CIF")
        mod_act = sum(v["cu"] for k,v in act_dict.items() if v["tipo"]=="MOD")
        mod_sim = sum(v["cu"] for k,v in sim_dict.items() if v["tipo"]=="MOD")

        def var(a, s):
            v = s - a
            p = (v/a*100) if a != 0 else 0
            return round(v,6), round(p,1)

        rows_res = []
        for concepto, a, s in [
            ("Costo Unitario", total_act, total_sim),
            ("CM Unitario",    cm_act,    cm_sim),
            ("CIF Unitario",   cif_act,   cif_sim),
            ("MOD Unitario",   mod_act,   mod_sim),
        ]:
            dv, dp = var(a, s)
            rows_res.append({
                "Concepto":        concepto,
                "Actual (S/)":     round(a, 6),
                "Simulado (S/)":   round(s, 6),
                "Variación (S/)":  dv,
                "Variación (%)":   f"{dp:+.1f}%",
            })
        pd.DataFrame(rows_res).to_excel(writer, sheet_name="Comparativo Resumen", index=False)

        # ── Hoja 2: Comparativo Detalle ─────────────────────
        rows_det = []
        for key in all_keys:
            a_row = act_dict.get(key, {})
            s_row = sim_dict.get(key, {})
            proceso = a_row.get("proceso") or s_row.get("proceso","")
            tipo    = a_row.get("tipo")    or s_row.get("tipo","")
            cu_a    = a_row.get("cu", 0)
            cu_s    = s_row.get("cu", 0)
            dv, dp  = var(cu_a, cu_s)
            rows_det.append({
                "Proceso":          proceso,
                "Tipo":             tipo,
                "Actual (S/)":      round(cu_a, 6),
                "Simulado (S/)":    round(cu_s, 6),
                "Variación (S/)":   dv,
                "Variación (%)":    f"{dp:+.1f}%",
            })
        pd.DataFrame(rows_det).to_excel(writer, sheet_name="Comparativo Detalle", index=False)

        # ── Formato ─────────────────────────────────────────
        fills = {
            "Comparativo Resumen": PatternFill("solid", fgColor="1F3864"),
            "Comparativo Detalle": PatternFill("solid", fgColor="1F5C2E"),
        }
        hf   = Font(bold=True, color="FFFFFF")
        info = (f"PT Base: {codigo_pt_base}  |  Simulación: {nombre}  |  "
                f"Cant. Base: {int(cant_base_pt):,}  |  {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        for sn, ws in writer.sheets.items():
            ws.insert_rows(1); ws["A1"] = info
            ws["A1"].font = Font(bold=True, color="00C8FF")
            ws["A1"].fill = PatternFill("solid", fgColor="0F1923")
            ws.insert_rows(2)
            for cell in ws[3]:
                cell.font = hf; cell.fill = fills.get(sn, fills["Comparativo Resumen"])
                cell.alignment = Alignment(horizontal="center")
            # Color variación negativa (ahorro) en verde, positiva en rojo
            var_col = None
            for i, cell in enumerate(ws[3]):
                if cell.value and "Variación (S/)" in str(cell.value):
                    var_col = i + 1
            if var_col:
                for row in ws.iter_rows(min_row=4, min_col=var_col, max_col=var_col):
                    for cell in row:
                        try:
                            v = float(str(cell.value).replace("+","").replace("%",""))
                            cell.font = Font(bold=True,
                                color="4CAF50" if v < 0 else ("FF5252" if v > 0 else "FFFFFF"))
                        except: pass
            for col in ws.columns:
                ml = max((len(str(c.value)) for c in col if c.value), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml+4, 40)

    output.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return dcc.send_bytes(output.read(), filename=f"Receta_{codigo_nuevo or 'nuevo'}_{ts}.xlsx")


if __name__ == "__main__":
    app.run(debug=False, host="0.0.0.0", port=int(os.environ.get("PORT", 8051)))
