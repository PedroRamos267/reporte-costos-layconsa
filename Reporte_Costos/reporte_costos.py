"""
=============================================================
  REPORTE AUTOMATIZADO DE COSTOS - EXPLOSIÓN DE MATERIALES
  Lógica (de abajo hacia arriba):
    CM  = solo componentes COMPRADOS × Costo Calculado
    CIF = (T.Maq / Cant.Base.Tiempos) × Cantidad Total Requerida del semi × Tarifa Maquina
    MOD = (T.MO  / Cant.Base.Tiempos) × Cantidad Total Requerida del semi × Tarifa MO
    Costo x Und = (suma CM + CIF + MOD) / Cantidad Total Requerida del semi
    Regla: CM de fabricados (231) ya está en su propio proceso, no se duplica
=============================================================
"""

import pandas as pd
import os
import sys
from datetime import datetime

# ─── CONFIGURACIÓN ─────────────────────────────────────────
ARCHIVO_DATOS    = r"D:\Users\pramos\Desktop\Reporte_Costos\Analisis de costos_PY.xlsx"
HOJA_EXPLOSION   = "Explosión"
HOJA_TIEMPOS     = "Tiempos"
CARPETA_OUTPUT   = r"D:\Users\pramos\Desktop\Reporte_Costos\output"
PREFIJO_FABRIC   = "231"
PROCESOS_EXCLUIR = []  # Sin exclusiones
# ───────────────────────────────────────────────────────────


def cargar_datos():
    print("\n📂 Cargando datos...")
    try:
        df_exp = pd.read_excel(ARCHIVO_DATOS, sheet_name=HOJA_EXPLOSION)
        df_tie = pd.read_excel(ARCHIVO_DATOS, sheet_name=HOJA_TIEMPOS)
    except Exception as e:
        print(f"❌ Error al leer Excel: {e}")
        sys.exit(1)

    df_exp.columns = df_exp.columns.str.strip()
    df_tie.columns = df_tie.columns.str.strip()

    for col in ["Código PT", "Código Semi", "Componente", "Familia"]:
        if col in df_exp.columns:
            df_exp[col] = df_exp[col].astype(str).str.strip()
    df_tie["Código Semi"] = df_tie["Código Semi"].astype(str).str.strip()

    for col in ["Cantidad Total Requerida", "Cantidad Base", "Costo estandar"]:
        df_exp[col] = pd.to_numeric(df_exp[col], errors="coerce").fillna(0)
    for col in ["Cantidad Base", "T.MO", "T.Maq", "Tarifa MO", "Tarifa Maquina"]:
        if col in df_tie.columns:
            df_tie[col] = pd.to_numeric(df_tie[col], errors="coerce").fillna(0)

    print(f"   ✅ Explosión: {len(df_exp)} filas | Tiempos: {len(df_tie)} filas")
    return df_exp, df_tie


def es_fabricado(familia):
    return str(familia).strip().startswith(PREFIJO_FABRIC)


def get_tiempos(codigo, df_tie):
    row = df_tie[df_tie["Código Semi"] == str(codigo)]
    return row.iloc[0] if not row.empty else None


def calcular_semi(codigo_semi, cantidad_req, df_exp, df_tie, cache, resumen_global, codigo_pt):
    """
    Calcula CM, CIF, MOD y Costo x Und de un semielaborado.
    cantidad_req = Cantidad Total Requerida con que aparece este semi en el nivel superior.
    CIF = (T.Maq / Cant.Base.Tiempos) × cantidad_req × Tarifa Maquina
    MOD = (T.MO  / Cant.Base.Tiempos) × cantidad_req × Tarifa MO
    Costo x Und = (CM + CIF + MOD) / cantidad_req
    """
    # Cache por codigo+cantidad para evitar recalcular con distintas cantidades
    cache_key = f"{codigo_semi}_{cantidad_req}"
    if cache_key in cache:
        return cache[cache_key]["costo_x_und"], []

    hijos = df_exp[(df_exp["Código Semi"] == str(codigo_semi)) & (df_exp["Código PT"] == str(codigo_pt))].copy()
    if hijos.empty:
        return 0, []

    desc_semi = hijos["Descripción Semi"].iloc[0] if "Descripción Semi" in hijos.columns else ""

    # Tiempos del semielaborado
    t           = get_tiempos(codigo_semi, df_tie)
    proceso     = str(t["Proceso"]).strip().upper() if t is not None else "SIN PROCESO"
    cant_base_t = float(t["Cantidad Base"])          if t is not None else 1
    tarifa_maq  = float(t["Tarifa Maquina"])         if t is not None else 0
    tarifa_mo   = float(t["Tarifa MO"])              if t is not None else 0
    t_maq       = float(t["T.Maq"])                  if t is not None else 0
    t_mo        = float(t["T.MO"])                   if t is not None else 0
    if cant_base_t == 0:
        cant_base_t = 1

    # CIF y MOD usando cantidad_req (Cantidad Total Requerida del nivel superior)
    cif = (t_maq / cant_base_t) * cantidad_req * tarifa_maq
    mod = (t_mo  / cant_base_t) * cantidad_req * tarifa_mo

    detalle      = []
    cm_total     = 0
    cm_comprados = 0

    for _, row in hijos.iterrows():
        componente = str(row["Componente"])
        desc_comp  = str(row.get("Descripción Componente", ""))
        cantidad   = float(row["Cantidad Total Requerida"])
        costo_std  = float(row["Costo estandar"])
        familia    = str(row.get("Familia", componente[:3])).strip()

        if es_fabricado(familia):
            # Bajar un nivel pasando la Cantidad Total Requerida del componente
            costo_calc, sub_det = calcular_semi(
                componente, cantidad, df_exp, df_tie, cache, resumen_global, codigo_pt
            )
            detalle.extend(sub_det)
            cm_comp = cantidad * costo_calc
            # NO suma a cm_comprados — ya está en su proceso
        else:
            costo_calc   = costo_std
            cm_comp      = cantidad * costo_calc
            cm_comprados += cm_comp

        cm_total += cm_comp

        detalle.append({
            "Código Semi":            codigo_semi,
            "Descripción Semi":       desc_semi,
            "Componente":             componente,
            "Descripción Componente": desc_comp,
            "Familia":                familia,
            "Tipo":                   "FABRICADO" if es_fabricado(familia) else "COMPRADO",
            "Proceso":                proceso,
            "Cantidad Total Req":     cantidad,
            "Costo Calculado":        costo_calc,
            "CM":                     cm_comp,
            "CIF":                    0,
            "MOD":                    0,
            "Total":                  cm_comp,
        })

    total_semi  = cm_total + cif + mod
    costo_x_und = total_semi / cantidad_req if cantidad_req != 0 else 0

    # Acumular en resumen global solo si el proceso no está excluido
    if proceso not in PROCESOS_EXCLUIR:
        if proceso not in resumen_global:
            resumen_global[proceso] = {"CM": 0, "CIF": 0, "MOD": 0}
        resumen_global[proceso]["CM"]  += cm_comprados
        resumen_global[proceso]["CIF"] += cif
        resumen_global[proceso]["MOD"] += mod

    # Fila resumen del proceso
    detalle.append({
        "Código Semi":            codigo_semi,
        "Descripción Semi":       desc_semi,
        "Componente":             f"[PROCESO] {codigo_semi}",
        "Descripción Componente": f"{proceso} — CIF + MOD",
        "Familia":                PREFIJO_FABRIC,
        "Tipo":                   "PROCESO",
        "Proceso":                proceso,
        "Cantidad Total Req":     cantidad_req,
        "Costo Calculado":        costo_x_und,
        "CM":                     cm_total,
        "CIF":                    cif,
        "MOD":                    mod,
        "Total":                  cm_total + cif + mod,
    })

    cache[cache_key] = {"costo_x_und": costo_x_und}
    return costo_x_und, detalle


def explotar_pt(codigo_pt, df_exp, df_tie):
    nivel1 = df_exp[df_exp["Código Semi"] == str(codigo_pt)].copy()
    if nivel1.empty:
        return {}, [], 0

    cant_base_pt = float(nivel1["Cantidad Base"].iloc[0])
    if cant_base_pt == 0:
        cant_base_pt = 1
    desc_pt = nivel1["Descripción Semi"].iloc[0] if "Descripción Semi" in nivel1.columns else ""

    # Tiempos del PT
    t           = get_tiempos(codigo_pt, df_tie)
    proceso_pt  = str(t["Proceso"]).strip().upper() if t is not None else "ENCAJADO"
    cant_base_t = float(t["Cantidad Base"])          if t is not None else 1
    tarifa_maq  = float(t["Tarifa Maquina"])         if t is not None else 0
    tarifa_mo   = float(t["Tarifa MO"])              if t is not None else 0
    t_maq       = float(t["T.Maq"])                  if t is not None else 0
    t_mo        = float(t["T.MO"])                   if t is not None else 0
    if cant_base_t == 0:
        cant_base_t = 1

    # CIF y MOD del PT usando cant_base_pt
    cif_pt = (t_maq / cant_base_t) * cant_base_pt * tarifa_maq
    mod_pt = (t_mo  / cant_base_t) * cant_base_pt * tarifa_mo

    cache          = {}
    detalle        = []
    resumen_global = {}
    cm_total       = 0
    cm_comprados   = 0

    for _, row in nivel1.iterrows():
        componente = str(row["Componente"])
        desc_comp  = str(row.get("Descripción Componente", ""))
        cantidad   = float(row["Cantidad Total Requerida"])
        costo_std  = float(row["Costo estandar"])
        familia    = str(row.get("Familia", componente[:3])).strip()

        if es_fabricado(familia):
            # Pasar la Cantidad Total Requerida del componente
            costo_calc, sub_det = calcular_semi(
                componente, cantidad, df_exp, df_tie, cache, resumen_global, codigo_pt
            )
            detalle.extend(sub_det)
            cm_comp = cantidad * costo_calc
        else:
            costo_calc   = costo_std
            cm_comp      = cantidad * costo_calc
            cm_comprados += cm_comp

        cm_total += cm_comp

        detalle.append({
            "Código Semi":            codigo_pt,
            "Descripción Semi":       desc_pt,
            "Componente":             componente,
            "Descripción Componente": desc_comp,
            "Familia":                familia,
            "Tipo":                   "FABRICADO" if es_fabricado(familia) else "COMPRADO",
            "Proceso":                proceso_pt,
            "Cantidad Total Req":     cantidad,
            "Costo Calculado":        costo_calc,
            "CM":                     cm_comp,
            "CIF":                    0,
            "MOD":                    0,
            "Total":                  cm_comp,
        })

    # Acumular proceso PT
    if proceso_pt not in resumen_global:
        resumen_global[proceso_pt] = {"CM": 0, "CIF": 0, "MOD": 0}
    resumen_global[proceso_pt]["CM"]  += cm_comprados
    resumen_global[proceso_pt]["CIF"] += cif_pt
    resumen_global[proceso_pt]["MOD"] += mod_pt

    total_pt    = cm_total + cif_pt + mod_pt
    costo_x_und = total_pt / cant_base_pt

    return resumen_global, detalle, costo_x_und


def generar_reporte(df_exp, df_tie):
    print("\n⚙️  Calculando costos por proceso...")

    lista_pt      = df_exp["Código PT"].unique()
    filas_resumen = []
    filas_detalle = []

    for codigo_pt in lista_pt:
        df_pt_rows = df_exp[df_exp["Código PT"] == codigo_pt]
        if df_pt_rows.empty:
            continue
        desc_pt = df_pt_rows["Descripción PT"].iloc[0]

        resumen, detalle, costo_x_und = explotar_pt(codigo_pt, df_exp, df_tie)

        total_general = sum(v["CM"] + v["CIF"] + v["MOD"] for v in resumen.values())
        if total_general == 0:
            continue

        cant_base_pt = float(df_exp[df_exp["Código Semi"] == codigo_pt]["Cantidad Base"].iloc[0]) \
                       if not df_exp[df_exp["Código Semi"] == codigo_pt].empty else 1
        if cant_base_pt == 0:
            cant_base_pt = 1

        for proceso, valores in resumen.items():
            for tipo, monto in [("CM",  valores["CM"]),
                                 ("CIF", valores["CIF"]),
                                 ("MOD", valores["MOD"])]:
                if monto > 0:
                    filas_resumen.append({
                        "Código PT":      codigo_pt,
                        "Descripción PT": desc_pt,
                        "Proceso":        proceso,
                        "Tipo de Costo":  f"{tipo} {proceso}",
                        "Costo Unitario": monto / cant_base_pt,
                        "Total PT":       total_general,
                    })

        for d in detalle:
            d["Código PT"]      = codigo_pt
            d["Descripción PT"] = desc_pt
            filas_detalle.append(d)

    df_resumen = pd.DataFrame(filas_resumen)
    df_detalle = pd.DataFrame(filas_detalle)

    if not df_resumen.empty:
        df_resumen["% del Total"] = (
            df_resumen["Costo Unitario"] /
            df_resumen.groupby("Código PT")["Costo Unitario"].transform("sum")
        )

    print(f"   ✅ {len(lista_pt)} productos terminados procesados.")
    return df_resumen, df_detalle


def exportar_excel(df_resumen, df_detalle):
    os.makedirs(CARPETA_OUTPUT, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta      = os.path.join(CARPETA_OUTPUT, f"Reporte_Costos_{timestamp}.xlsx")
    print(f"\n📊 Generando Excel...")

    with pd.ExcelWriter(ruta, engine="openpyxl") as writer:

        # ── Hoja 1: Resumen Global ──────────────────────────
        resumen_global = df_resumen[
            ["Código PT", "Descripción PT", "Proceso", "Tipo de Costo", "Costo Unitario", "% del Total"]
        ].copy()
        # Agregar fila TOTAL por PT
        totales = df_resumen.groupby(["Código PT", "Descripción PT"])["Costo Unitario"].sum().reset_index()
        totales["Proceso"]       = "TOTAL"
        totales["Tipo de Costo"] = "TOTAL PT"
        totales["% del Total"]   = 1.0
        resumen_global = pd.concat([resumen_global, totales], ignore_index=True)
        resumen_global = resumen_global.sort_values(["Código PT", "Tipo de Costo"])
        resumen_global["% del Total"]   = resumen_global["% del Total"].map("{:.1%}".format)
        resumen_global["Costo Unitario"] = resumen_global["Costo Unitario"].map("{:.6f}".format)
        resumen_global.to_excel(writer, sheet_name="Resumen Global", index=False)

        # ── Hoja 2: Detalle Componentes ─────────────────────
        cols_det = ["Código PT", "Descripción PT", "Código Semi", "Descripción Semi",
                    "Componente", "Descripción Componente", "Familia", "Tipo",
                    "Proceso", "Cantidad Total Req", "Costo Calculado",
                    "CM", "CIF", "MOD", "Total"]
        cols_det = [c for c in cols_det if c in df_detalle.columns]
        df_detalle[cols_det].to_excel(writer, sheet_name="Detalle Componentes", index=False)

        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        header_font = Font(bold=True, color="FFFFFF")
        fills = {
            "Resumen Global":      PatternFill("solid", fgColor="1F3864"),
            "Detalle Componentes": PatternFill("solid", fgColor="2E75B6"),
        }

        for sheet_name, ws in writer.sheets.items():
            fill = fills.get(sheet_name, PatternFill("solid", fgColor="1F3864"))
            for cell in ws[1]:
                cell.font      = header_font
                cell.fill      = fill
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)

    print(f"   ✅ Excel guardado en: {ruta}")
    return ruta


def lanzar_dashboard(df_resumen, df_detalle, df_exp, df_tie):
    try:
        import plotly.graph_objects as go
        from dash import Dash, html, dcc, Input, Output, dash_table, State
    except ImportError:
        print("❌ Instala dash y plotly: pip install dash plotly")
        sys.exit(1)

    app     = Dash(__name__)
    COLORES = {
        "bg": "#0F1923", "card": "#1A2633", "text": "#E8EDF2",
        "accent": "#00C8FF", "CM": "#2196F3", "MOD": "#4CAF50",
        "CIF": "#FF9800", "TOTAL": "#E91E63",
    }

    lista_pt = df_resumen[["Código PT", "Descripción PT"]].drop_duplicates()

    def get_maquinas_inyeccion(codigo_pt):
        """
        Obtiene las máquinas de INYECCIÓN usadas por un PT.
        Una fila por máquina (no por componente).
        """
        visitados = set()
        maquinas  = {}  # {codigo_maquina: datos}

        def buscar(codigo):
            if codigo in visitados:
                return
            visitados.add(codigo)
            hijos = df_exp[df_exp["Código Semi"] == codigo]
            for _, row in hijos.iterrows():
                comp    = str(row["Componente"])
                familia = str(row.get("Familia", comp[:3])).strip()
                t       = df_tie[df_tie["Código Semi"] == comp]
                if not t.empty:
                    proc = str(t.iloc[0].get("Proceso", "")).strip().upper()
                    if "INYEC" in proc:
                        t_row  = t.iloc[0]
                        maq    = str(t_row.get("Maquina", comp))
                        if maq not in maquinas:
                            maquinas[maq] = {
                                "Maquina":     maq,
                                "T.Ciclo":     float(t_row.get("T.ciclo",   0) or 0),
                                "Cav.Oper":    float(t_row.get("Cav. Oper", 0) or 0),
                                "Cav.Tot":     float(t_row.get("Cav. Tot",  0) or 0),
                                "Tarifa Maq":  float(t_row.get("Tarifa Maquina", 0) or 0),
                                "Tarifa MO":   float(t_row.get("Tarifa MO",      0) or 0),
                            }
                if familia.startswith("231"):
                    buscar(comp)

        buscar(codigo_pt)
        return list(maquinas.values())

    app.layout = html.Div(
        style={"backgroundColor": COLORES["bg"], "minHeight": "100vh",
               "fontFamily": "'Segoe UI', sans-serif",
               "color": COLORES["text"], "padding": "20px"},
        children=[
            html.H1("📦 Reporte de Costos por Proceso",
                    style={"color": COLORES["accent"], "textAlign": "center",
                           "marginBottom": "5px"}),
            html.P(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                   style={"color": "#7A9BBF", "textAlign": "center",
                          "marginBottom": "25px"}),

            html.Div(style={"marginBottom": "25px"}, children=[
                html.Label("Selecciona un Producto Terminado:",
                           style={"color": COLORES["accent"], "fontWeight": "bold"}),
                dcc.Dropdown(
                    id="selector-pt",
                    options=[{"label": f"{r['Código PT']} — {r['Descripción PT']}",
                              "value": r["Código PT"]}
                             for _, r in lista_pt.iterrows()],
                    value=lista_pt["Código PT"].iloc[0],
                    style={"marginTop": "8px", "color": "#000"}
                ),
            ]),

            html.Div(id="kpis", style={"display": "flex", "gap": "15px",
                                        "marginBottom": "25px", "flexWrap": "wrap"}),

            # ── Simulador de Inyección ──────────────────────────
            html.Div(style={"backgroundColor": COLORES["card"], "borderRadius": "12px",
                            "padding": "15px", "marginBottom": "20px",
                            "border": "1px solid #00C8FF"}, children=[
                html.H3("🔧 Simulador de Inyección — Modifica T.Ciclo y Cav.Oper",
                        style={"color": COLORES["accent"], "fontSize": "16px",
                               "marginTop": 0, "marginBottom": "10px"}),
                html.P("Edita los valores en la tabla y presiona Recalcular para ver el impacto en costos.",
                       style={"color": "#7A9BBF", "fontSize": "12px", "marginBottom": "10px"}),
                dash_table.DataTable(
                    id="tabla-simulador",
                    columns=[
                        {"name": "Máquina",       "id": "Maquina",   "editable": False},
                        {"name": "T.Ciclo (s)",   "id": "T.Ciclo",   "editable": True,  "type": "numeric"},
                        {"name": "Cav.Oper",      "id": "Cav.Oper",  "editable": True,  "type": "numeric"},
                        {"name": "Cav.Tot",       "id": "Cav.Tot",   "editable": False},
                        {"name": "Cant.Base Calc","id": "Cant.Base", "editable": False},
                        {"name": "Tarifa Maq",    "id": "Tarifa Maq","editable": False},
                        {"name": "Tarifa MO",     "id": "Tarifa MO", "editable": False},
                    ],
                    style_header={"backgroundColor": "#1F3864", "color": "white",
                                  "fontWeight": "bold"},
                    style_cell={"backgroundColor": "#1E2D3D", "color": COLORES["text"],
                                "border": "1px solid #2A3F54", "padding": "8px",
                                "textAlign": "center"},
                    style_data_conditional=[
                        {"if": {"column_editable": True},
                         "backgroundColor": "#0D2137",
                         "border": "1px solid #00C8FF"},
                        {"if": {"row_index": "odd"}, "backgroundColor": "#162030"},
                    ],
                    editable=True,
                    row_selectable=False,
                    page_action="none",
                ),
                html.Div(style={"marginTop": "12px", "display": "flex",
                                "gap": "15px", "alignItems": "center"}, children=[
                    html.Button("🔄 Recalcular",
                        id="btn-recalcular",
                        style={"backgroundColor": COLORES["accent"],
                               "color": "#000", "fontWeight": "bold",
                               "border": "none", "borderRadius": "8px",
                               "padding": "10px 24px", "cursor": "pointer",
                               "fontSize": "14px"}),
                    html.Div(id="msg-simulador",
                             style={"color": "#4CAF50", "fontSize": "13px"}),
                ]),
            ]),

            html.Div(style={"display": "grid", "gridTemplateColumns": "1fr 1fr",
                            "gap": "20px", "marginBottom": "20px"}, children=[
                html.Div(style={"backgroundColor": COLORES["card"],
                                "borderRadius": "12px", "padding": "15px"}, children=[
                    html.H3("Cascada de Costos (S/)",
                            style={"color": COLORES["accent"],
                                   "fontSize": "16px", "marginTop": 0}),
                    dcc.Graph(id="grafico-cascada")
                ]),
                html.Div(style={"backgroundColor": COLORES["card"],
                                "borderRadius": "12px", "padding": "15px"}, children=[
                    html.H3("Cascada de Costos (%)",
                            style={"color": COLORES["accent"],
                                   "fontSize": "16px", "marginTop": 0}),
                    dcc.Graph(id="grafico-cascada-pct")
                ]),
            ]),
            html.Div(style={"backgroundColor": COLORES["card"], "borderRadius": "12px",
                            "padding": "15px", "marginBottom": "20px"}, children=[
                html.H3("Costo Total por Proceso",
                        style={"color": COLORES["accent"],
                               "fontSize": "16px", "marginTop": 0}),
                dcc.Graph(id="grafico-donut")
            ]),

            html.Div(style={"backgroundColor": COLORES["card"], "borderRadius": "12px",
                            "padding": "15px", "marginBottom": "20px"}, children=[
                html.H3("Resumen por Proceso",
                        style={"color": COLORES["accent"],
                               "fontSize": "16px", "marginTop": 0}),
                dash_table.DataTable(
                    id="tabla-resumen",
                    style_header={"backgroundColor": "#1F3864", "color": "white",
                                  "fontWeight": "bold"},
                    style_cell={"backgroundColor": "#1E2D3D", "color": COLORES["text"],
                                "border": "1px solid #2A3F54", "padding": "8px",
                                "textAlign": "center"},
                    style_data_conditional=[
                        {"if": {"row_index": "odd"}, "backgroundColor": "#162030"}
                    ],
                    page_size=15,
                )
            ]),

            html.Div(style={"backgroundColor": COLORES["card"], "borderRadius": "12px",
                            "padding": "15px"}, children=[
                html.H3("Detalle por Componente",
                        style={"color": COLORES["accent"],
                               "fontSize": "16px", "marginTop": 0}),
                dash_table.DataTable(
                    id="tabla-detalle",
                    style_header={"backgroundColor": "#1F3864", "color": "white",
                                  "fontWeight": "bold"},
                    style_cell={"backgroundColor": "#1E2D3D", "color": COLORES["text"],
                                "border": "1px solid #2A3F54", "padding": "8px",
                                "textAlign": "center"},
                    style_data_conditional=[
                        {"if": {"row_index": "odd"}, "backgroundColor": "#162030"}
                    ],
                    page_size=15,
                    filter_action="native",
                    sort_action="native",
                )
            ]),
        ]
    )

    # ── Callback 1: actualizar tabla simulador al cambiar PT ──
    @app.callback(
        Output("tabla-simulador", "data"),
        Input("selector-pt", "value"),
    )
    def cargar_simulador(codigo_pt):
        maquinas = get_maquinas_inyeccion(codigo_pt)
        rows     = []
        for m in maquinas:
            cant_base = round((3600 / m["T.Ciclo"]) * m["Cav.Oper"] * 24, 2)                         if m["T.Ciclo"] > 0 else 0
            rows.append({
                "Maquina":   m["Maquina"],
                "T.Ciclo":   m["T.Ciclo"],
                "Cav.Oper":  m["Cav.Oper"],
                "Cav.Tot":   m["Cav.Tot"],
                "Cant.Base": cant_base,
                "Tarifa Maq":m["Tarifa Maq"],
                "Tarifa MO": m["Tarifa MO"],
            })
        return rows

    # ── Callback 2: recalcular al presionar botón ──────────
    @app.callback(
        Output("kpis",                "children"),
        Output("grafico-cascada",     "figure"),
        Output("grafico-cascada-pct", "figure"),
        Output("grafico-donut",       "figure"),
        Output("tabla-resumen",       "data"),
        Output("tabla-resumen",       "columns"),
        Output("tabla-detalle",       "data"),
        Output("tabla-detalle",       "columns"),
        Output("msg-simulador",       "children"),
        Input("btn-recalcular",       "n_clicks"),
        Input("selector-pt",          "value"),
        State("tabla-simulador",      "data"),
    )
    def actualizar(n_clicks, codigo_pt, datos_simulador):
        # Aplicar parámetros del simulador a df_tie temporalmente
        df_tie_sim = df_tie.copy()
        if datos_simulador:
            for row in datos_simulador:
                maquina  = str(row.get("Maquina", ""))
                t_ciclo  = float(row.get("T.Ciclo", 0) or 0)
                cav_oper = float(row.get("Cav.Oper", 0) or 0)
                if t_ciclo > 0 and cav_oper > 0 and maquina:
                    nueva_base = (3600 / t_ciclo) * cav_oper * 24
                    # Aplica a todos los semis que usan esta máquina
                    mask = df_tie_sim["Maquina"].astype(str).str.strip() == maquina
                    df_tie_sim.loc[mask, "Cantidad Base"] = nueva_base
                    df_tie_sim.loc[mask, "T.ciclo"]      = t_ciclo
                    df_tie_sim.loc[mask, "Cav. Oper"]    = cav_oper

        # Recalcular costos con los nuevos parámetros
        from collections import defaultdict
        resumen_sim, detalle_sim, _ = explotar_pt(codigo_pt, df_exp, df_tie_sim)

        cant_base_pt = float(df_exp[df_exp["Código Semi"] == codigo_pt]["Cantidad Base"].iloc[0])                        if not df_exp[df_exp["Código Semi"] == codigo_pt].empty else 1
        if cant_base_pt == 0:
            cant_base_pt = 1

        filas = []
        for proceso, valores in resumen_sim.items():
            for tipo, monto in [("CM", valores["CM"]), ("CIF", valores["CIF"]), ("MOD", valores["MOD"])]:
                if monto > 0:
                    filas.append({
                        "Código PT":      codigo_pt,
                        "Descripción PT": "",
                        "Proceso":        proceso,
                        "Tipo de Costo":  f"{tipo} {proceso}",
                        "Costo Unitario": monto / cant_base_pt,
                        "Total PT":       0,
                    })

        df_pt = pd.DataFrame(filas)
        if df_pt.empty:
            df_pt = df_resumen[df_resumen["Código PT"] == codigo_pt].copy()
        else:
            total = df_pt["Costo Unitario"].sum()
            df_pt["% del Total"] = df_pt["Costo Unitario"] / total if total > 0 else 0

        msg    = f"✅ Recalculado — {datetime.now().strftime('%H:%M:%S')}" if n_clicks else ""
        df_det = df_detalle[df_detalle["Código PT"] == codigo_pt].copy()

        total   = df_pt["Costo Unitario"].sum()
        tot_cm  = df_pt[df_pt["Tipo de Costo"].str.startswith("CM")]["Costo Unitario"].sum()
        tot_mod = df_pt[df_pt["Tipo de Costo"].str.startswith("MOD")]["Costo Unitario"].sum()
        tot_cif = df_pt[df_pt["Tipo de Costo"].str.startswith("CIF")]["Costo Unitario"].sum()

        def kpi(titulo, valor, color):
            return html.Div(
                style={"backgroundColor": COLORES["card"],
                       "borderLeft": f"4px solid {color}",
                       "borderRadius": "10px", "padding": "15px 20px",
                       "flex": "1", "minWidth": "160px"},
                children=[
                    html.P(titulo, style={"margin": 0, "fontSize": "12px",
                                          "color": "#7A9BBF"}),
                    html.H2(f"S/ {valor:.6f}",
                            style={"margin": "5px 0 0 0", "color": color,
                                   "fontSize": "18px"}),
                ]
            )

        kpis_elem = [
            kpi("💰 Costo x Und", total,   COLORES["accent"]),
            kpi("🧱 CM Total",    tot_cm,  COLORES["CM"]),
            kpi("👷 MOD Total",   tot_mod, COLORES["MOD"]),
            kpi("⚙️ CIF Total",   tot_cif, COLORES["CIF"]),
        ]

        # ── Waterfall valores ──────────────────────────────────
        labels   = list(df_pt["Tipo de Costo"]) + ["TOTAL"]
        valores  = list(df_pt["Costo Unitario"]) + [total]
        measures = ["relative"] * len(df_pt) + ["total"]

        fig_cas = go.Figure(go.Waterfall(
            x            = labels,
            y            = valores,
            measure      = measures,
            text         = [f"S/ {v:.4f}" for v in valores],
            textposition = "outside",
            increasing   = dict(marker_color=COLORES["CM"]),
            totals       = dict(marker_color=COLORES["TOTAL"]),
            connector    = dict(line=dict(color="#4A5568", width=1)),
            hovertemplate="<b>%{x}</b><br>S/ %{y:.6f}<extra></extra>"
        ))
        fig_cas.update_layout(
            template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            margin=dict(l=10, r=10, t=30, b=80), xaxis_tickangle=-35,
            showlegend=False
        )

        # ── Waterfall porcentaje ──────────────────────────────
        pcts         = list(df_pt["% del Total"] * 100) + [100.0]
        measures_pct = ["relative"] * len(df_pt) + ["total"]

        fig_cas_pct = go.Figure(go.Waterfall(
            x            = labels,
            y            = pcts,
            measure      = measures_pct,
            text         = [f"{v:.1f}%" for v in pcts],
            textposition = "outside",
            increasing   = dict(marker_color=COLORES["MOD"]),
            totals       = dict(marker_color=COLORES["TOTAL"]),
            connector    = dict(line=dict(color="#4A5568", width=1)),
            hovertemplate="<b>%{x}</b><br>%{y:.1f}%<extra></extra>"
        ))
        fig_cas_pct.update_layout(
            template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            yaxis=dict(ticksuffix="%"),
            margin=dict(l=10, r=10, t=30, b=80), xaxis_tickangle=-35,
            showlegend=False
        )

        # ── Dona por proceso ──────────────────────────────────
        resumen_proc = df_pt.groupby("Proceso")["Costo Unitario"].sum().reset_index()
        paleta = ["#2196F3", "#4CAF50", "#FF9800", "#E91E63",
                  "#9C27B0", "#00BCD4", "#FF5722"]
        fig_don = go.Figure(go.Pie(
            labels       = resumen_proc["Proceso"],
            values       = resumen_proc["Costo Unitario"],
            hole         = 0.55,
            marker_colors= paleta[:len(resumen_proc)],
            hovertemplate="<b>%{label}</b><br>S/ %{value:.6f}<br>%{percent}<extra></extra>"
        ))
        fig_don.update_layout(
            template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)",
            margin=dict(l=10, r=10, t=10, b=10)
        )

        df_res_fmt = df_pt[["Tipo de Costo", "Costo Unitario", "% del Total"]].copy()
        df_res_fmt["Costo Unitario"] = df_res_fmt["Costo Unitario"].map("S/ {:.6f}".format)
        df_res_fmt["% del Total"]    = df_res_fmt["% del Total"].map("{:.1%}".format)
        cols_res = [{"name": c, "id": c} for c in df_res_fmt.columns]

        cols_show  = ["Código Semi", "Descripción Semi", "Componente",
                      "Descripción Componente", "Proceso", "Tipo",
                      "Cantidad Total Req", "Costo Calculado",
                      "CM", "CIF", "MOD", "Total"]
        cols_show  = [c for c in cols_show if c in df_det.columns]
        df_det_fmt = df_det[cols_show].copy()
        for c in ["Costo Calculado", "CM", "CIF", "MOD", "Total"]:
            if c in df_det_fmt.columns:
                df_det_fmt[c] = df_det_fmt[c].map("{:.6f}".format)
        cols_det = [{"name": c, "id": c} for c in df_det_fmt.columns]

        return (kpis_elem, fig_cas, fig_cas_pct, fig_don,
                df_res_fmt.to_dict("records"), cols_res,
                df_det_fmt.to_dict("records"), cols_det, msg)

    print("\n🚀 Iniciando Dashboard...")
    print("   Abre tu navegador en: http://127.0.0.1:8050")
    print("   Presiona Ctrl+C para detener.\n")
    app.run(debug=False)


def main():
    print("=" * 60)
    print("   REPORTE DE COSTOS POR PROCESO — EXPLOSIÓN BOM")
    print("=" * 60)

    df_exp, df_tie     = cargar_datos()
    df_resumen, df_det = generar_reporte(df_exp, df_tie)

    if df_resumen.empty:
        print("⚠️  No se generaron resultados. Verifica tu data.")
        sys.exit(1)

    print("\n¿Qué deseas generar?")
    print("  1 → Exportar a Excel")
    print("  2 → Lanzar Dashboard interactivo")
    print("  3 → Ambos")
    opcion = input("\nElige una opción (1/2/3): ").strip()

    if opcion in ("1", "3"):
        exportar_excel(df_resumen, df_det)
    if opcion in ("2", "3"):
        lanzar_dashboard(df_resumen, df_det, df_exp, df_tie)
    if opcion not in ("1", "2", "3"):
        print("❌ Opción no válida.")


if __name__ == "__main__":
    main()
